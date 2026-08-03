##"""
##Dairy Farm MS Pro - Fully Functional Mobile Version
##====================================================
##Production-ready dairy farm management for Android/iOS.
##
##Features:
##- Bottom Navigation (mobile-native)
##- Data Validation & Error Handling
##- Search & Filter across all modules
##- Charts & Analytics (Matplotlib)
##- CSV/JSON Export & Share
##- Farm Settings & Profile
##- Dark Mode Support
##- Auto-backup
##- Cow Detail Profiles
##- Production Trends
##- Expense Breakdown Charts
##- Income vs Expense Analytics
##- Recent Activity Feed
##- Responsive Touch Targets
##
##Dependencies:
##    pip install kivymd matplotlib
##
##Build Android:
##    buildozer android debug
##
##Run Desktop:
##    python dairy_farm_mobile_pro.py
##"""
##
##import json
##import os
##import csv
##import io
##import base64
##from datetime import datetime, timedelta
##from functools import partial
##from collections import defaultdict
##
##from kivy.config import Config
##nConfig = Config  # alias to avoid issues
##nConfig.set('graphics', 'width', '360')
##nConfig.set('graphics', 'height', '640')
##
##from kivy.core.window import Window
##from kivy.metrics import dp
##from kivy.properties import ObjectProperty, StringProperty, BooleanProperty
##from kivy.clock import Clock
##from kivy.uix.boxlayout import BoxLayout
##from kivy.uix.scrollview import ScrollView
##from kivy.uix.anchorlayout import AnchorLayout
##from kivy.animation import Animation
##from kivy.graphics import Color, Rectangle, RoundedRectangle
##
##from kivymd.app import MDApp
##from kivymd.uix.boxlayout import MDBoxLayout
##from kivymd.uix.button import (
##    MDRaisedButton, MDIconButton, MDFlatButton, 
##    MDFloatingActionButton, MDRoundFlatButton
##)
##from kivymd.uix.card import MDCard
##from kivymd.uix.chip import MDChip
##from kivymd.uix.dialog import MDDialog
##from kivymd.uix.label import MDLabel
##from kivymd.uix.list import (
##    MDList, OneLineIconListItem, TwoLineAvatarIconListItem,
##    IconLeftWidget, IconRightWidget
##)
##from kivymd.uix.menu import MDDropdownMenu
##from kivymd.uix.pickers import MDDatePicker
##from kivymd.uix.screen import MDScreen
##from kivymd.uix.screenmanager import MDScreenManager
##from kivy.uix.screenmanager import SlideTransition
##from kivymd.uix.textfield import MDTextField
##from kivymd.uix.toolbar import MDTopAppBar
### from kivymd.uix.bottomnavigation import MDBottomNavigation, MDBottomNavigationItem  # Using custom bottom bar
##from kivymd.uix.navigationdrawer import MDNavigationDrawer
##from kivymd.uix.snackbar import MDSnackbar
##from kivymd.uix.selectioncontrol import MDCheckbox
##from kivymd.uix.progressbar import MDProgressBar
##from kivymd.uix.gridlayout import MDGridLayout
##
### Chart imports
### Charts disabled: kivy-garden matplotlib is incompatible with Kivy 2.3.1
### All chart functions return None gracefully
##CHARTS_AVAILABLE = False
##
### Keep stub imports for the chart helper functions to reference
### but they will never create actual widgets
##try:
##    import matplotlib
##    matplotlib.use('Agg')
##    import matplotlib.pyplot as plt
##    from matplotlib.dates import DateFormatter
##except Exception:
##    pass
##
### ============================================================
### THEME & CONFIG
### ============================================================
##
##THEME = {
##    "primary": [0.102, 0.361, 0.227, 1],      # #1a5c3a
##    "accent": [0.176, 0.541, 0.369, 1],       # #2d8a5e
##    "light": [0.302, 0.722, 0.478, 1],        # #4db87a
##    "bg": [0.941, 0.957, 0.941, 1],           # #f0f4f0
##    "card": [1, 1, 1, 1],
##    "danger": [0.906, 0.298, 0.235, 1],       # #e74c3c
##    "warning": [0.953, 0.612, 0.071, 1],      # #f39c12
##    "success": [0.153, 0.682, 0.376, 1],      # #27ae60
##    "info": [0.204, 0.596, 0.859, 1],         # #3498db
##    "text": [0.173, 0.243, 0.314, 1],         # #2c3e50
##    "text_secondary": [0.498, 0.549, 0.553, 1], # #7f8c8d
##    "divider": [0.85, 0.85, 0.85, 1],
##}
##
##DARK_THEME = {
##    "primary": [0.302, 0.722, 0.478, 1],
##    "accent": [0.176, 0.541, 0.369, 1],
##    "light": [0.102, 0.361, 0.227, 1],
##    "bg": [0.12, 0.12, 0.12, 1],
##    "card": [0.18, 0.18, 0.18, 1],
##    "danger": [0.906, 0.298, 0.235, 1],
##    "warning": [0.953, 0.612, 0.071, 1],
##    "success": [0.153, 0.682, 0.376, 1],
##    "info": [0.204, 0.596, 0.859, 1],
##    "text": [0.9, 0.9, 0.9, 1],
##    "text_secondary": [0.6, 0.6, 0.6, 1],
##    "divider": [0.3, 0.3, 0.3, 1],
##}
##
##CURRENCY = "KES"
##
### ============================================================
### DATA LAYER (Robust & Validated)
### ============================================================
##
##def get_data_dir():
##    try:
##        app = MDApp.get_running_app()
##        if app:
##            return app.user_data_dir
##    except:
##        pass
##    base = os.path.expanduser("~/.local/share/DairyFarmMSPro")
##    os.makedirs(base, exist_ok=True)
##    return base
##
##
##def get_settings_path():
##    return os.path.join(get_data_dir(), "settings.json")
##
##
##def get_data_path():
##    return os.path.join(get_data_dir(), "dairy_farm_data.json")
##
##
##def get_backup_dir():
##    bd = os.path.join(get_data_dir(), "backups")
##    os.makedirs(bd, exist_ok=True)
##    return bd
##
##
##def load_settings():
##    path = get_settings_path()
##    defaults = {
##        "farm_name": "My Dairy Farm",
##        "farm_location": "",
##        "owner_name": "",
##        "currency": "KES",
##        "dark_mode": False,
##        "auto_backup": True,
##        "milk_price_default": "60",
##        "daily_rate_default": "500",
##    }
##    if os.path.exists(path):
##        try:
##            with open(path, "r", encoding="utf-8") as f:
##                loaded = json.load(f)
##                defaults.update(loaded)
##        except:
##            pass
##    return defaults
##
##
##def save_settings(settings):
##    with open(get_settings_path(), "w", encoding="utf-8") as f:
##        json.dump(settings, f, indent=2)
##
##
##def load_data():
##    path = get_data_path()
##    defaults = {
##        "farm": {"name": "My Dairy Farm", "location": "", "owner": ""},
##        "herd": [],
##        "milk_production": [],
##        "milk_sales": [],
##        "other_income": [],
##        "feed_expenses": [],
##        "vet_health": [],
##        "labour": [],
##        "operations": [],
##        "assets": [],
##        "breeding": [],
##        "activity_log": [],
##    }
##    if os.path.exists(path):
##        try:
##            with open(path, "r", encoding="utf-8") as f:
##                loaded = json.load(f)
##                # Merge with defaults for new fields
##                for k, v in defaults.items():
##                    if k not in loaded:
##                        loaded[k] = v
##                return loaded
##        except Exception as e:
##            print(f"Data load error: {e}")
##    return defaults
##
##
##def save_data(data):
##    try:
##        with open(get_data_path(), "w", encoding="utf-8") as f:
##            json.dump(data, f, indent=2, default=str)
##        # Auto-backup
##        settings = load_settings()
##        if settings.get("auto_backup", True):
##            backup_file = os.path.join(get_backup_dir(), f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
##            try:
##                with open(backup_file, "w", encoding="utf-8") as f:
##                    json.dump(data, f, indent=2, default=str)
##                # Keep only last 10 backups
##                backups = sorted([f for f in os.listdir(get_backup_dir()) if f.startswith("backup_")])
##                for old in backups[:-10]:
##                    os.remove(os.path.join(get_backup_dir(), old))
##            except:
##                pass
##        return True
##    except Exception as e:
##        print(f"Save error: {e}")
##        return False
##
##
##def log_activity(data, action, details=""):
##    data.setdefault("activity_log", []).insert(0, {
##        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
##        "action": action,
##        "details": details,
##    })
##    data["activity_log"] = data["activity_log"][:100]  # Keep last 100
##    save_data(data)
##
##
##def validate_date(date_str):
##    try:
##        datetime.strptime(str(date_str), "%Y-%m-%d")
##        return True
##    except:
##        return False
##
##
##def safe_float(val, default=0.0):
##    try:
##        return float(val) if val not in (None, "", "None") else default
##    except:
##        return default
##
##
##def format_currency(amount):
##    try:
##        return f"{CURRENCY} {float(amount):,.0f}"
##    except:
##        return f"{CURRENCY} 0"
##
##
### ============================================================
### CHART HELPERS
### ============================================================
##
##def create_pie_chart(labels, values, title, colors_list=None):
##    if not CHARTS_AVAILABLE:
##        return None
##    try:
##        fig, ax = plt.subplots(figsize=(3.2, 3.2), dpi=100)
##        fig.patch.set_alpha(0)
##        ax.set_facecolor('none')
##        if colors_list:
##            ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90, colors=colors_list,
##                   textprops={'color': 'white', 'fontsize': 8})
##        else:
##            ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90,
##                   textprops={'color': 'white', 'fontsize': 8})
##        ax.set_title(title, color='white', fontsize=10, pad=10)
##        plt.tight_layout()
##        canvas = FigureCanvasKivyAgg(fig)
##        return canvas
##    except Exception as e:
##        print(f"Chart error: {e}")
##        return None
##
##
##def create_line_chart(dates, values, title, ylabel="Amount (KES)"):
##    if not CHARTS_AVAILABLE or not dates:
##        return None
##    try:
##        fig, ax = plt.subplots(figsize=(3.2, 2.4), dpi=100)
##        fig.patch.set_alpha(0)
##        ax.set_facecolor('none')
##        ax.plot(dates, values, marker='o', linewidth=2, markersize=4, color='#4db87a')
##        ax.set_title(title, color='white', fontsize=10)
##        ax.set_ylabel(ylabel, color='white', fontsize=8)
##        ax.tick_params(colors='white', labelsize=7)
##        ax.spines['bottom'].set_color('white')
##        ax.spines['left'].set_color('white')
##        ax.spines['top'].set_visible(False)
##        ax.spines['right'].set_visible(False)
##        ax.grid(True, alpha=0.3, color='white')
##        plt.tight_layout()
##        return FigureCanvasKivyAgg(fig)
##    except Exception as e:
##        print(f"Chart error: {e}")
##        return None
##
##
##def create_bar_chart(labels, values, title, color='#2d8a5e'):
##    if not CHARTS_AVAILABLE or not labels:
##        return None
##    try:
##        fig, ax = plt.subplots(figsize=(3.2, 2.4), dpi=100)
##        fig.patch.set_alpha(0)
##        ax.set_facecolor('none')
##        bars = ax.bar(labels, values, color=color, alpha=0.8)
##        ax.set_title(title, color='white', fontsize=10)
##        ax.tick_params(colors='white', labelsize=7)
##        ax.spines['bottom'].set_color('white')
##        ax.spines['left'].set_color('white')
##        ax.spines['top'].set_visible(False)
##        ax.spines['right'].set_visible(False)
##        ax.grid(True, alpha=0.3, color='white', axis='y')
##        plt.tight_layout()
##        return FigureCanvasKivyAgg(fig)
##    except Exception as e:
##        print(f"Chart error: {e}")
##        return None
##
##
### ============================================================
### UI HELPERS
### ============================================================
##
##class ThemedCard(MDCard):
##    """Enhanced card with theme support."""
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.elevation = 2
##        self.radius = [dp(12),]
##        self.padding = dp(12)
##        self.md_bg_color = THEME["card"]
##
##
##class KPICard(ThemedCard):
##    def __init__(self, title, value, color, subtitle="", **kwargs):
##        super().__init__(**kwargs)
##        self.orientation = "vertical"
##        self.size_hint_y = None
##        self.height = dp(100)
##        self.add_widget(MDLabel(
##            text=title, theme_text_color="Custom",
##            text_color=THEME["text_secondary"], font_style="Caption",
##            halign="center", size_hint_y=None, height=dp(18)
##        ))
##        self.add_widget(MDLabel(
##            text=value, theme_text_color="Custom",
##            text_color=color, font_style="H5",
##            halign="center", bold=True, size_hint_y=None, height=dp(32)
##        ))
##        if subtitle:
##            self.add_widget(MDLabel(
##                text=subtitle, theme_text_color="Custom",
##                text_color=THEME["text_secondary"], font_style="Caption",
##                halign="center", size_hint_y=None, height=dp(18)
##            ))
##
##
##class ActionCard(ThemedCard):
##    def __init__(self, title, subtitle, on_edit=None, on_delete=None, on_tap=None, **kwargs):
##        super().__init__(**kwargs)
##        self.orientation = "vertical"
##        self.size_hint_y = None
##        self.height = dp(90)
##
##        main = MDBoxLayout()
##        text_area = MDBoxLayout(orientation="vertical", size_hint_x=0.78)
##        text_area.add_widget(MDLabel(
##            text=title, theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="Subtitle1",
##            bold=True, size_hint_y=None, height=dp(22)
##        ))
##        text_area.add_widget(MDLabel(
##            text=subtitle, theme_text_color="Secondary",
##            font_style="Caption", size_hint_y=None, height=dp(36)
##        ))
##        main.add_widget(text_area)
##
##        btn_area = MDBoxLayout(size_hint_x=0.22, spacing=dp(2))
##        if on_edit:
##            btn_area.add_widget(MDIconButton(
##                icon="pencil", theme_text_color="Custom",
##                text_color=THEME["info"], on_release=on_edit,
##                icon_size=dp(18)
##            ))
##        if on_delete:
##            btn_area.add_widget(MDIconButton(
##                icon="delete", theme_text_color="Custom",
##                text_color=THEME["danger"], on_release=on_delete,
##                icon_size=dp(18)
##            ))
##        main.add_widget(btn_area)
##        self.add_widget(main)
##
##        if on_tap:
##            self.bind(on_release=on_tap)
##
##
##class SearchBar(MDBoxLayout):
##    def __init__(self, on_search, **kwargs):
##        super().__init__(**kwargs)
##        self.size_hint_y = None
##        self.height = dp(50)
##        self.padding = [dp(4), dp(4), dp(4), dp(4)]
##        self.md_bg_color = THEME["card"]
##        self.radius = [dp(8),]
##
##        self.search_field = MDTextField(
##            hint_text="Search...",
##            icon_right="magnify",
##            mode="rectangle",
##            size_hint_x=0.85,
##            height=dp(40)
##        )
##        self.search_field.bind(text=lambda inst, val: on_search(val))
##        self.add_widget(self.search_field)
##
##        clear_btn = MDIconButton(
##            icon="close-circle", theme_text_color="Custom",
##            text_color=THEME["text_secondary"],
##            on_release=lambda x: [setattr(self.search_field, "text", ""), on_search("")],
##            size_hint_x=0.15
##        )
##        self.add_widget(clear_btn)
##
##
##class MobileDialog:
##    @staticmethod
##    def create(title, fields, on_save, on_cancel=None, size_hint=(0.92, None)):
##        content = MDBoxLayout(orientation="vertical", spacing=dp(6), size_hint_y=None,
##                              padding=[dp(8), dp(8), dp(8), dp(8)])
##        field_refs = {}
##        total_h = dp(16)
##
##        for field in fields:
##            h = dp(70) if field.get("multiline") else dp(55)
##            tf = MDTextField(
##                hint_text=field.get("hint", ""),
##                icon_right=field.get("icon", ""),
##                multiline=field.get("multiline", False),
##                mode="rectangle",
##                size_hint_y=None, height=h,
##                text=field.get("text", ""),
##                input_filter=field.get("input_filter", None),
##            )
##            field_refs[field.get("key", field["hint"])] = tf
##            content.add_widget(tf)
##            total_h += h + dp(6)
##
##        content.height = total_h
##
##        def _save(*args):
##            vals = {k: v.text for k, v in field_refs.items()}
##            # Validation
##            errors = []
##            for field in fields:
##                key = field.get("key", field["hint"])
##                if field.get("required") and not vals.get(key, "").strip():
##                    errors.append(f"{field['hint']} is required")
##                if field.get("is_date") and vals.get(key) and not validate_date(vals[key]):
##                    errors.append(f"{field['hint']} must be YYYY-MM-DD")
##                if field.get("is_number") and vals.get(key):
##                    try:
##                        float(vals[key])
##                    except:
##                        errors.append(f"{field['hint']} must be a number")
##
##            if errors:
##                show_snackbar("; ".join(errors[:2]), THEME["danger"])
##                return
##            on_save(vals)
##            dialog.dismiss()
##
##        def _cancel(*args):
##            if on_cancel: on_cancel()
##            dialog.dismiss()
##
##        dialog = MDDialog(
##            title=title, type="custom", content_cls=content,
##            size_hint=size_hint,
##            buttons=[
##                MDFlatButton(text="CANCEL", on_release=_cancel),
##                MDRaisedButton(text="SAVE", md_bg_color=THEME["success"], on_release=_save)
##            ],
##        )
##        return dialog, field_refs
##
##
##def show_snackbar(text, color=THEME["success"], duration=2):
##    try:
##        # KivyMD 2.0+ style
##        sb = MDSnackbar(
##            MDLabel(text=text, theme_text_color="Custom", text_color=[1,1,1,1], font_style="Body2"),
##            md_bg_color=color, duration=duration,
##        )
##    except TypeError:
##        # KivyMD 1.2.0 style
##        sb = MDSnackbar(text=text, md_bg_color=color, duration=duration)
##    sb.open()
##
##
##class ConfirmDialog:
##    @staticmethod
##    def show(title, text, on_confirm, confirm_text="DELETE", confirm_color=None):
##        if confirm_color is None:
##            confirm_color = THEME["danger"]
##        dialog = MDDialog(
##            title=title, text=text,
##            buttons=[
##                MDFlatButton(text="CANCEL", on_release=lambda x: dialog.dismiss()),
##                MDRaisedButton(text=confirm_text, md_bg_color=confirm_color,
##                               on_release=lambda x: [dialog.dismiss(), on_confirm()])
##            ]
##        )
##        dialog.open()
##
##
### ============================================================
### BASE SCREEN
### ============================================================
##
##class BaseScreen(MDScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.data = load_data()
##        self.settings = load_settings()
##        self.dialog = None
##        self.search_query = ""
##
##    def reload_data(self):
##        self.data = load_data()
##        self.settings = load_settings()
##
##    def refresh(self):
##        pass
##
##    def create_scroll_layout(self):
##        scroll = ScrollView()
##        content = MDBoxLayout(
##            orientation="vertical", size_hint_y=None,
##            padding=[dp(12), dp(8), dp(12), dp(80)],
##            spacing=dp(8)
##        )
##        content.bind(minimum_height=content.setter("height"))
##        scroll.add_widget(content)
##        return scroll, content
##
##    def add_screen_title(self, content, title):
##        content.add_widget(MDLabel(
##            text=title, theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H5",
##            halign="center", size_hint_y=None, height=dp(40), bold=True
##        ))
##
##    def add_fab(self, layout, callback, icon="plus"):
##        fab = MDFloatingActionButton(
##            icon=icon, md_bg_color=THEME["accent"],
##            pos_hint={"right": 0.95, "y": 0.04},
##            on_release=callback, elevation=4
##        )
##        layout.add_widget(fab)
##        return fab
##
##    def filter_items(self, items, query, keys):
##        if not query:
##            return items
##        q = query.lower()
##        return [item for item in items if any(q in str(item.get(k, "")).lower() for k in keys)]
##
##
### ============================================================
### DASHBOARD SCREEN
### ============================================================
##
##class DashboardScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##
##        # Header
##        farm_name = self.settings.get("farm_name", "My Dairy Farm")
##        content.add_widget(MDLabel(
##            text=farm_name.upper(), theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H4",
##            halign="center", size_hint_y=None, height=dp(44), bold=True
##        ))
##        content.add_widget(MDLabel(
##            text=datetime.now().strftime("%A, %d %B %Y"),
##            theme_text_color="Custom", text_color=THEME["text_secondary"],
##            font_style="Caption", halign="center", size_hint_y=None, height=dp(22)
##        ))
##
##        # KPIs
##        income = self._calc_income()
##        expenses = self._calc_expenses()
##        net = income - expenses
##        milking = len([c for c in self.data["herd"] if c.get("status", "").lower() == "milking"])
##        total_herd = len(self.data["herd"])
##
##        grid = MDGridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(220))
##        grid.add_widget(KPICard("INCOME", format_currency(income), THEME["success"]))
##        grid.add_widget(KPICard("EXPENSES", format_currency(expenses), THEME["danger"]))
##        net_color = THEME["success"] if net >= 0 else THEME["danger"]
##        grid.add_widget(KPICard("NET P&L", format_currency(net), net_color))
##        grid.add_widget(KPICard("HERD", f"{milking}/{total_herd}", THEME["accent"], "Milking/Total"))
##        content.add_widget(grid)
##
##        # Today's Milk
##        today = datetime.now().strftime("%Y-%m-%d")
##        today_milk = sum(safe_float(x.get("quantity")) for x in self.data["milk_production"] if x.get("date", "") == today)
##        content.add_widget(KPICard("TODAY'S MILK", f"{today_milk:.1f} L", THEME["info"], f"{len([x for x in self.data['milk_production'] if x.get('date')==today])} records"))
##
##        # Quick Actions
##        content.add_widget(MDLabel(text="QUICK ACTIONS", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        actions = MDBoxLayout(size_hint_y=None, height=dp(44), spacing=dp(6))
##        actions.add_widget(MDRaisedButton(text="+ Milk", md_bg_color=THEME["info"],
##            on_release=lambda x: setattr(self.manager, "current", "milk_production")))
##        actions.add_widget(MDRaisedButton(text="+ Sale", md_bg_color=THEME["success"],
##            on_release=lambda x: setattr(self.manager, "current", "milk_sales")))
##        actions.add_widget(MDRaisedButton(text="+ Cow", md_bg_color=THEME["accent"],
##            on_release=lambda x: setattr(self.manager, "current", "herd")))
##        actions.add_widget(MDRaisedButton(text="+ Expense", md_bg_color=THEME["danger"],
##            on_release=lambda x: setattr(self.manager, "current", "feed_expenses")))
##        content.add_widget(actions)
##
##        # Recent Activity
##        content.add_widget(MDLabel(text="RECENT ACTIVITY", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        logs = self.data.get("activity_log", [])[:10]
##        if not logs:
##            content.add_widget(MDLabel(text="No recent activity.", theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(50)))
##        for log in logs:
##            card = ThemedCard(orientation="horizontal", size_hint_y=None, height=dp(50), padding=dp(8))
##            card.add_widget(MDLabel(
##                text=f"{log.get('timestamp', '-')[11:16]}  {log.get('action', '')}",
##                theme_text_color="Custom", text_color=THEME["text"],
##                font_style="Caption", size_hint_x=0.7
##            ))
##            card.add_widget(MDLabel(
##                text=log.get('details', '')[:20],
##                theme_text_color="Custom", text_color=THEME["text_secondary"],
##                font_style="Caption", halign="right", size_hint_x=0.3
##            ))
##            content.add_widget(card)
##
##        layout.add_widget(scroll)
##        self.add_widget(layout)
##
##    def _calc_income(self):
##        milk = sum(safe_float(x.get("amount")) for x in self.data.get("milk_sales", []))
##        other = sum(safe_float(x.get("amount")) for x in self.data.get("other_income", []))
##        return milk + other
##
##    def _calc_expenses(self):
##        feed = sum(safe_float(x.get("amount")) for x in self.data.get("feed_expenses", []))
##        vet = sum(safe_float(x.get("cost")) for x in self.data.get("vet_health", []))
##        lab = sum(safe_float(x.get("amount")) for x in self.data.get("labour", []))
##        ops = sum(safe_float(x.get("amount")) for x in self.data.get("operations", []))
##        return feed + vet + lab + ops
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### HERD SCREEN
### ============================================================
##
##class HerdScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "HERD REGISTER")
##
##        # Stats
##        total = len(self.data["herd"])
##        milking = len([c for c in self.data["herd"] if c.get("status", "").lower() == "milking"])
##        dry = len([c for c in self.data["herd"] if c.get("status", "").lower() == "dry"])
##        heifer = len([c for c in self.data["herd"] if c.get("status", "").lower() == "heifer"])
##
##        stats = MDBoxLayout(size_hint_y=None, height=dp(32), spacing=dp(6))
##        for label, count, color in [("Total", total, THEME["primary"]), 
##                                     ("Milking", milking, THEME["success"]),
##                                     ("Dry", dry, THEME["warning"]),
##                                     ("Heifer", heifer, THEME["info"])]:
##            chip = MDChip(text=f"{label}: {count}", md_bg_color=color)
##            chip.text_color = [1, 1, 1, 1]
##            stats.add_widget(chip)
##        content.add_widget(stats)
##
##        # Search
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        # List
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        items = self.filter_items(self.data["herd"], self.search_query, ["tag_no", "name", "breed", "status"])
##
##        if not items:
##            msg = "No cows found." if self.search_query else "No cows registered yet.\nTap + to add your first cow."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(100)))
##            return
##
##        for i, cow in enumerate(items):
##            orig_idx = self.data["herd"].index(cow)
##            status_colors = {
##                "milking": THEME["success"], "dry": THEME["warning"],
##                "heifer": THEME["info"], "sold": THEME["danger"], "dead": THEME["danger"]
##            }
##            status_color = status_colors.get(cow.get("status", "").lower(), THEME["text_secondary"])
##
##            # Calculate age
##            age_text = ""
##            if cow.get("dob"):
##                try:
##                    dob = datetime.strptime(cow["dob"], "%Y-%m-%d")
##                    age_days = (datetime.now() - dob).days
##                    age_text = f"Age: {age_days // 365}y {(age_days % 365) // 30}m"
##                except:
##                    pass
##
##            card = ActionCard(
##                title=f"{cow.get('tag_no', 'N/A')} - {cow.get('name', 'Unknown')}",
##                subtitle=f"Breed: {cow.get('breed', '-')} | Status: {cow.get('status', '-')} | {age_text}",
##                on_edit=partial(self.edit_cow, orig_idx),
##                on_delete=partial(self.delete_cow, orig_idx),
##                on_tap=partial(self.view_cow, orig_idx)
##            )
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        dialog, fields = MobileDialog.create("Add New Cow", [
##            {"key": "tag", "hint": "Tag Number *", "icon": "identifier", "required": True},
##            {"key": "name", "hint": "Cow Name", "icon": "cow"},
##            {"key": "breed", "hint": "Breed", "icon": "dna"},
##            {"key": "status", "hint": "Status (Milking/Dry/Heifer)", "icon": "information", "text": "Milking"},
##            {"key": "dob", "hint": "Date of Birth (YYYY-MM-DD)", "icon": "calendar", "is_date": True},
##            {"key": "cost", "hint": "Purchase Cost", "icon": "cash", "input_filter": "float"},
##            {"key": "notes", "hint": "Notes", "icon": "note-text", "multiline": True},
##        ], on_save=self.save_cow)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_cow(self, values):
##        self.data["herd"].append({
##            "tag_no": values["tag"], "name": values["name"],
##            "breed": values["breed"], "status": values["status"] or "Milking",
##            "dob": values["dob"], "purchase_date": datetime.now().strftime("%Y-%m-%d"),
##            "purchase_cost": values["cost"] or "0", "notes": values["notes"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Added cow", values["tag"])
##        show_snackbar(f"Cow {values['tag']} added!")
##        self.refresh()
##
##    def edit_cow(self, index, *args):
##        cow = self.data["herd"][index]
##        dialog, fields = MobileDialog.create("Edit Cow", [
##            {"key": "tag", "hint": "Tag Number", "icon": "identifier", "text": cow.get("tag_no", "")},
##            {"key": "name", "hint": "Cow Name", "icon": "cow", "text": cow.get("name", "")},
##            {"key": "breed", "hint": "Breed", "icon": "dna", "text": cow.get("breed", "")},
##            {"key": "status", "hint": "Status", "icon": "information", "text": cow.get("status", "Milking")},
##            {"key": "dob", "hint": "Date of Birth", "icon": "calendar", "text": cow.get("dob", ""), "is_date": True},
##            {"key": "cost", "hint": "Purchase Cost", "icon": "cash", "text": str(cow.get("purchase_cost", "")), "input_filter": "float"},
##            {"key": "notes", "hint": "Notes", "icon": "note-text", "text": cow.get("notes", ""), "multiline": True},
##        ], on_save=lambda v: self.update_cow(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_cow(self, index, values):
##        self.data["herd"][index].update({
##            "tag_no": values["tag"], "name": values["name"],
##            "breed": values["breed"], "status": values["status"],
##            "dob": values["dob"], "purchase_cost": values["cost"] or "0",
##            "notes": values["notes"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Updated cow", values["tag"])
##        show_snackbar("Cow updated!")
##        self.refresh()
##
##    def delete_cow(self, index, *args):
##        tag = self.data["herd"][index].get("tag_no", "Unknown")
##        def confirm():
##            self.data["herd"].pop(index)
##            save_data(self.data)
##            log_activity(self.data, "Deleted cow", tag)
##            show_snackbar(f"Cow {tag} removed")
##            self.refresh()
##        ConfirmDialog.show("Delete Cow?", f"Remove {tag} permanently?", confirm)
##
##    def view_cow(self, index, *args):
##        cow = self.data["herd"][index]
##        # Production stats
##        cow_milk = sum(safe_float(x.get("quantity")) for x in self.data["milk_production"] if x.get("cow_tag") == cow.get("tag_no"))
##        cow_vet = sum(safe_float(x.get("cost")) for x in self.data["vet_health"] if x.get("cow_tag") == cow.get("tag_no"))
##
##        content = MDBoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None)
##        content.height = dp(200)
##        info_lines = [
##            f"Tag: {cow.get('tag_no', '-')}",
##            f"Name: {cow.get('name', '-')}",
##            f"Breed: {cow.get('breed', '-')}",
##            f"Status: {cow.get('status', '-')}",
##            f"DOB: {cow.get('dob', '-')}",
##            f"Purchased: {cow.get('purchase_date', '-')}",
##            f"Total Milk: {cow_milk:.1f} L",
##            f"Vet Costs: {format_currency(cow_vet)}",
##        ]
##        for line in info_lines:
##            content.add_widget(MDLabel(text=line, theme_text_color="Custom",
##                text_color=THEME["text"], font_style="Body1", size_hint_y=None, height=dp(22)))
##
##        dialog = MDDialog(title="Cow Profile", type="custom", content_cls=content,
##            buttons=[MDFlatButton(text="CLOSE", on_release=lambda x: dialog.dismiss())])
##        dialog.open()
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### MILK PRODUCTION SCREEN
### ============================================================
##
##class MilkProductionScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "MILK PRODUCTION")
##
##        today = datetime.now().strftime("%Y-%m-%d")
##        today_recs = [x for x in self.data["milk_production"] if x.get("date", "") == today]
##        today_total = sum(safe_float(x.get("quantity")) for x in today_recs)
##
##        content.add_widget(KPICard("TODAY'S TOTAL", f"{today_total:.1f} L", THEME["info"], f"{len(today_recs)} records"))
##
##        # 7-day trend chart
##        if CHARTS_AVAILABLE:
##            dates = []
##            amounts = []
##            for i in range(6, -1, -1):
##                d = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
##                amt = sum(safe_float(x.get("quantity")) for x in self.data["milk_production"] if x.get("date") == d)
##                dates.append(d[5:])  # MM-DD
##                amounts.append(amt)
##            chart = create_bar_chart(dates, amounts, "Last 7 Days (L)", color='#4db87a')
##            if chart:
##                try:
##                    chart.size_hint_y = None
##                    chart.height = dp(200)
##                    content.add_widget(chart)
##                except Exception as e:
##                    print(f"Chart display error: {e}")
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        all_items = sorted(self.data["milk_production"], key=lambda x: x.get("date", ""), reverse=True)
##        items = self.filter_items(all_items, self.search_query, ["date", "cow_tag", "session", "quality"])
##
##        if not items:
##            msg = "No records found." if self.search_query else "No milk production records yet."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(80)))
##            return
##
##        for i, rec in enumerate(items[:50]):  # Limit to 50 for performance
##            orig_idx = self.data["milk_production"].index(rec)
##            card = ActionCard(
##                title=f"{rec.get('date', '-')} | {rec.get('session', 'AM')}",
##                subtitle=f"Cow: {rec.get('cow_tag', '-')} | {rec.get('quantity', '0')} L | {rec.get('quality', 'Good')}",
##                on_edit=partial(self.edit_record, orig_idx),
##                on_delete=partial(self.delete_record, orig_idx)
##            )
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        cow_tags = [c.get("tag_no", "") for c in self.data["herd"] if c.get("tag_no")]
##        cow_hint = f"Cow Tag ({', '.join(cow_tags[:3])}...)" if cow_tags else "Cow Tag"
##        dialog, fields = MobileDialog.create("Record Milk Production", [
##            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##            {"key": "session", "hint": "Session (AM/PM/Evening)", "icon": "clock", "text": "AM", "required": True},
##            {"key": "cow", "hint": cow_hint, "icon": "cow", "required": True},
##            {"key": "qty", "hint": "Quantity (Liters)", "icon": "cup-water", "input_filter": "float", "required": True, "is_number": True},
##            {"key": "quality", "hint": "Quality", "icon": "star", "text": "Good"},
##            {"key": "notes", "hint": "Notes", "icon": "note-text", "multiline": True},
##        ], on_save=self.save_record)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_record(self, values):
##        self.data["milk_production"].append({
##            "date": values["date"], "session": values["session"],
##            "cow_tag": values["cow"], "quantity": values["qty"],
##            "quality": values["quality"], "notes": values["notes"],
##            "recorded_at": datetime.now().strftime("%Y-%m-%d %H:%M")
##        })
##        save_data(self.data)
##        log_activity(self.data, "Milk recorded", f"{values['cow']} - {values['qty']}L")
##        show_snackbar("Milk record saved!")
##        self.refresh()
##
##    def edit_record(self, index, *args):
##        rec = self.data["milk_production"][index]
##        dialog, fields = MobileDialog.create("Edit Milk Record", [
##            {"key": "date", "hint": "Date", "icon": "calendar", "text": rec.get("date", ""), "required": True, "is_date": True},
##            {"key": "session", "hint": "Session", "icon": "clock", "text": rec.get("session", "")},
##            {"key": "cow", "hint": "Cow Tag", "icon": "cow", "text": rec.get("cow_tag", "")},
##            {"key": "qty", "hint": "Quantity", "icon": "cup-water", "text": str(rec.get("quantity", "")), "input_filter": "float", "is_number": True},
##            {"key": "quality", "hint": "Quality", "icon": "star", "text": rec.get("quality", "")},
##            {"key": "notes", "hint": "Notes", "icon": "note-text", "text": rec.get("notes", ""), "multiline": True},
##        ], on_save=lambda v: self.update_record(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_record(self, index, values):
##        self.data["milk_production"][index].update({
##            "date": values["date"], "session": values["session"],
##            "cow_tag": values["cow"], "quantity": values["qty"],
##            "quality": values["quality"], "notes": values["notes"],
##        })
##        save_data(self.data)
##        show_snackbar("Record updated!")
##        self.refresh()
##
##    def delete_record(self, index, *args):
##        rec = self.data["milk_production"][index]
##        def confirm():
##            self.data["milk_production"].pop(index)
##            save_data(self.data)
##            show_snackbar("Record deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete Record?", f"Remove {rec.get('date', '')} - {rec.get('cow_tag', '')}?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### MILK SALES SCREEN
### ============================================================
##
##class MilkSalesScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.showing = "sales"
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "MILK SALES & INCOME")
##
##        total_sales = sum(safe_float(x.get("amount")) for x in self.data["milk_sales"])
##        total_other = sum(safe_float(x.get("amount")) for x in self.data["other_income"])
##        total_liters = sum(safe_float(x.get("liters")) for x in self.data["milk_sales"])
##
##        content.add_widget(KPICard("TOTAL SALES", format_currency(total_sales), THEME["success"], f"{total_liters:.0f} liters"))
##        content.add_widget(KPICard("OTHER INCOME", format_currency(total_other), THEME["info"]))
##
##        # Tabs
##        tab_bar = MDBoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6))
##        self.sales_btn = MDRaisedButton(text="Milk Sales", md_bg_color=THEME["primary"],
##            on_release=lambda x: self.switch_tab("sales"))
##        self.other_btn = MDFlatButton(text="Other Income", on_release=lambda x: self.switch_tab("other"))
##        tab_bar.add_widget(self.sales_btn)
##        tab_bar.add_widget(self.other_btn)
##        content.add_widget(tab_bar)
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def switch_tab(self, tab):
##        self.showing = tab
##        self.search_query = ""
##        if tab == "sales":
##            self.sales_btn.md_bg_color = THEME["primary"]
##            self.other_btn.md_bg_color = [0.9, 0.9, 0.9, 1]
##        else:
##            self.other_btn.md_bg_color = THEME["primary"]
##            self.sales_btn.md_bg_color = [0.9, 0.9, 0.9, 1]
##        self.refresh_list()
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##
##        if self.showing == "sales":
##            all_items = sorted(self.data["milk_sales"], key=lambda x: x.get("date", ""), reverse=True)
##            items = self.filter_items(all_items, self.search_query, ["date", "buyer", "amount"])
##            if not items:
##                self.list_container.add_widget(MDLabel(text="No sales records.", theme_text_color="Hint",
##                    halign="center", size_hint_y=None, height=dp(60)))
##            for item in items[:50]:
##                idx = self.data["milk_sales"].index(item)
##                card = ActionCard(
##                    title=f"{item.get('date', '-')} | {item.get('buyer', 'Unknown')}",
##                    subtitle=f"{item.get('liters', '0')}L @ KES {item.get('price_per_liter', '0')}/L = {format_currency(item.get('amount', '0'))}",
##                    on_edit=partial(self.edit_item, idx, "sales"),
##                    on_delete=partial(self.delete_item, idx, "sales")
##                )
##                self.list_container.add_widget(card)
##        else:
##            all_items = sorted(self.data["other_income"], key=lambda x: x.get("date", ""), reverse=True)
##            items = self.filter_items(all_items, self.search_query, ["date", "source", "amount"])
##            if not items:
##                self.list_container.add_widget(MDLabel(text="No other income records.", theme_text_color="Hint",
##                    halign="center", size_hint_y=None, height=dp(60)))
##            for item in items[:50]:
##                idx = self.data["other_income"].index(item)
##                card = ActionCard(
##                    title=f"{item.get('date', '-')} | {item.get('source', 'Unknown')}",
##                    subtitle=f"{format_currency(item.get('amount', '0'))} | {item.get('notes', '-')}",
##                    on_edit=partial(self.edit_item, idx, "other"),
##                    on_delete=partial(self.delete_item, idx, "other")
##                )
##                self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        if self.showing == "sales":
##            default_price = self.settings.get("milk_price_default", "60")
##            dialog, fields = MobileDialog.create("Record Milk Sale", [
##                {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##                {"key": "buyer", "hint": "Buyer Name", "icon": "account", "required": True},
##                {"key": "liters", "hint": "Liters Sold", "icon": "cup-water", "input_filter": "float", "required": True, "is_number": True},
##                {"key": "price", "hint": "Price per Liter (KES)", "icon": "cash", "input_filter": "float", "text": default_price, "is_number": True},
##            ], on_save=self.save_sale)
##        else:
##            dialog, fields = MobileDialog.create("Record Other Income", [
##                {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##                {"key": "source", "hint": "Income Source", "icon": "tag", "required": True},
##                {"key": "amount", "hint": "Amount (KES)", "icon": "cash", "input_filter": "float", "required": True, "is_number": True},
##                {"key": "notes", "hint": "Notes", "icon": "note-text", "multiline": True},
##            ], on_save=self.save_other)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_sale(self, values):
##        liters = safe_float(values["liters"])
##        price = safe_float(values["price"])
##        self.data["milk_sales"].append({
##            "date": values["date"], "buyer": values["buyer"],
##            "liters": str(liters), "price_per_liter": str(price),
##            "amount": str(liters * price),
##        })
##        save_data(self.data)
##        log_activity(self.data, "Milk sale", f"{values['buyer']} - {liters}L")
##        show_snackbar("Sale recorded!")
##        self.refresh()
##
##    def save_other(self, values):
##        self.data["other_income"].append({
##            "date": values["date"], "source": values["source"],
##            "amount": values["amount"], "notes": values["notes"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Other income", f"{values['source']} - {values['amount']}")
##        show_snackbar("Income recorded!")
##        self.refresh()
##
##    def edit_item(self, index, type, *args):
##        if type == "sales":
##            item = self.data["milk_sales"][index]
##            dialog, fields = MobileDialog.create("Edit Sale", [
##                {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "required": True, "is_date": True},
##                {"key": "buyer", "hint": "Buyer", "icon": "account", "text": item.get("buyer", "")},
##                {"key": "liters", "hint": "Liters", "icon": "cup-water", "text": str(item.get("liters", "")), "input_filter": "float", "is_number": True},
##                {"key": "price", "hint": "Price/Liter", "icon": "cash", "text": str(item.get("price_per_liter", "")), "input_filter": "float", "is_number": True},
##            ], on_save=lambda v: self.update_sale(index, v))
##        else:
##            item = self.data["other_income"][index]
##            dialog, fields = MobileDialog.create("Edit Income", [
##                {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "required": True, "is_date": True},
##                {"key": "source", "hint": "Source", "icon": "tag", "text": item.get("source", "")},
##                {"key": "amount", "hint": "Amount", "icon": "cash", "text": str(item.get("amount", "")), "input_filter": "float", "is_number": True},
##                {"key": "notes", "hint": "Notes", "icon": "note-text", "text": item.get("notes", ""), "multiline": True},
##            ], on_save=lambda v: self.update_other(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_sale(self, index, values):
##        liters = safe_float(values["liters"])
##        price = safe_float(values["price"])
##        self.data["milk_sales"][index].update({
##            "date": values["date"], "buyer": values["buyer"],
##            "liters": str(liters), "price_per_liter": str(price),
##            "amount": str(liters * price),
##        })
##        save_data(self.data)
##        show_snackbar("Sale updated!")
##        self.refresh()
##
##    def update_other(self, index, values):
##        self.data["other_income"][index].update({
##            "date": values["date"], "source": values["source"],
##            "amount": values["amount"], "notes": values["notes"],
##        })
##        save_data(self.data)
##        show_snackbar("Income updated!")
##        self.refresh()
##
##    def delete_item(self, index, type, *args):
##        key = "milk_sales" if type == "sales" else "other_income"
##        item = self.data[key][index]
##        def confirm():
##            self.data[key].pop(index)
##            save_data(self.data)
##            show_snackbar("Deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete?", "Remove this record permanently?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### FEED EXPENSES SCREEN
### ============================================================
##
##class FeedExpensesScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "FEED EXPENSES")
##
##        total = sum(safe_float(x.get("amount")) for x in self.data["feed_expenses"])
##        content.add_widget(KPICard("TOTAL FEED COSTS", format_currency(total), THEME["danger"]))
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        all_items = sorted(self.data["feed_expenses"], key=lambda x: x.get("date", ""), reverse=True)
##        items = self.filter_items(all_items, self.search_query, ["date", "feed_type", "supplier"])
##
##        if not items:
##            msg = "No records found." if self.search_query else "No feed expenses recorded."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(80)))
##            return
##
##        for item in items[:50]:
##            idx = self.data["feed_expenses"].index(item)
##            card = ActionCard(
##                title=f"{item.get('date', '-')} | {item.get('feed_type', 'Unknown')}",
##                subtitle=f"{item.get('supplier', '-')} | {item.get('quantity', '-')} {item.get('unit', 'kg')} | {format_currency(item.get('amount', '0'))}",
##                on_edit=partial(self.edit_item, idx),
##                on_delete=partial(self.delete_item, idx)
##            )
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        dialog, fields = MobileDialog.create("Add Feed Expense", [
##            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##            {"key": "type", "hint": "Feed Type (e.g. Dairy Meal)", "icon": "corn", "required": True},
##            {"key": "supplier", "hint": "Supplier Name", "icon": "truck-delivery"},
##            {"key": "qty", "hint": "Quantity", "icon": "numeric", "input_filter": "float", "is_number": True},
##            {"key": "unit", "hint": "Unit (kg/bale/bag)", "icon": "scale", "text": "kg"},
##            {"key": "amount", "hint": "Total Cost (KES)", "icon": "cash", "input_filter": "float", "required": True, "is_number": True},
##        ], on_save=self.save_item)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_item(self, values):
##        self.data["feed_expenses"].append({
##            "date": values["date"], "feed_type": values["type"],
##            "supplier": values["supplier"], "quantity": values["qty"],
##            "unit": values["unit"], "amount": values["amount"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Feed expense", f"{values['type']} - {values['amount']}")
##        show_snackbar("Feed expense saved!")
##        self.refresh()
##
##    def edit_item(self, index, *args):
##        item = self.data["feed_expenses"][index]
##        dialog, fields = MobileDialog.create("Edit Feed Expense", [
##            {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "is_date": True},
##            {"key": "type", "hint": "Feed Type", "icon": "corn", "text": item.get("feed_type", "")},
##            {"key": "supplier", "hint": "Supplier", "icon": "truck-delivery", "text": item.get("supplier", "")},
##            {"key": "qty", "hint": "Quantity", "icon": "numeric", "text": str(item.get("quantity", "")), "input_filter": "float", "is_number": True},
##            {"key": "unit", "hint": "Unit", "icon": "scale", "text": item.get("unit", "")},
##            {"key": "amount", "hint": "Amount", "icon": "cash", "text": str(item.get("amount", "")), "input_filter": "float", "is_number": True},
##        ], on_save=lambda v: self.update_item(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_item(self, index, values):
##        self.data["feed_expenses"][index].update({
##            "date": values["date"], "feed_type": values["type"],
##            "supplier": values["supplier"], "quantity": values["qty"],
##            "unit": values["unit"], "amount": values["amount"],
##        })
##        save_data(self.data)
##        show_snackbar("Updated!")
##        self.refresh()
##
##    def delete_item(self, index, *args):
##        item = self.data["feed_expenses"][index]
##        def confirm():
##            self.data["feed_expenses"].pop(index)
##            save_data(self.data)
##            show_snackbar("Deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete?", f"Remove {item.get('feed_type', 'this')} expense?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### VET & HEALTH SCREEN
### ============================================================
##
##class VetHealthScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "VET & HEALTH")
##
##        total = sum(safe_float(x.get("cost")) for x in self.data["vet_health"])
##        content.add_widget(KPICard("TOTAL VET COSTS", format_currency(total), THEME["danger"]))
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        all_items = sorted(self.data["vet_health"], key=lambda x: x.get("date", ""), reverse=True)
##        items = self.filter_items(all_items, self.search_query, ["date", "cow_tag", "treatment", "vet_name"])
##
##        if not items:
##            msg = "No records found." if self.search_query else "No veterinary records."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(80)))
##            return
##
##        for item in items[:50]:
##            idx = self.data["vet_health"].index(item)
##            card = ActionCard(
##                title=f"{item.get('date', '-')} | {item.get('cow_tag', 'Unknown')}",
##                subtitle=f"{item.get('treatment', '-')} | {item.get('vet_name', '-')} | {format_currency(item.get('cost', '0'))}",
##                on_edit=partial(self.edit_item, idx),
##                on_delete=partial(self.delete_item, idx)
##            )
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        dialog, fields = MobileDialog.create("Add Vet/Health Record", [
##            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##            {"key": "cow", "hint": "Cow Tag Number", "icon": "cow", "required": True},
##            {"key": "treatment", "hint": "Treatment/Diagnosis", "icon": "needle", "required": True},
##            {"key": "vet", "hint": "Veterinarian Name", "icon": "doctor"},
##            {"key": "cost", "hint": "Cost (KES)", "icon": "cash", "input_filter": "float", "is_number": True},
##            {"key": "notes", "hint": "Notes/Medications", "icon": "note-text", "multiline": True},
##        ], on_save=self.save_item)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_item(self, values):
##        self.data["vet_health"].append({
##            "date": values["date"], "cow_tag": values["cow"],
##            "treatment": values["treatment"], "vet_name": values["vet"],
##            "cost": values["cost"], "notes": values["notes"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Vet record", f"{values['cow']} - {values['treatment']}")
##        show_snackbar("Health record saved!")
##        self.refresh()
##
##    def edit_item(self, index, *args):
##        item = self.data["vet_health"][index]
##        dialog, fields = MobileDialog.create("Edit Health Record", [
##            {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "is_date": True},
##            {"key": "cow", "hint": "Cow Tag", "icon": "cow", "text": item.get("cow_tag", "")},
##            {"key": "treatment", "hint": "Treatment", "icon": "needle", "text": item.get("treatment", "")},
##            {"key": "vet", "hint": "Vet Name", "icon": "doctor", "text": item.get("vet_name", "")},
##            {"key": "cost", "hint": "Cost", "icon": "cash", "text": str(item.get("cost", "")), "input_filter": "float", "is_number": True},
##            {"key": "notes", "hint": "Notes", "icon": "note-text", "text": item.get("notes", ""), "multiline": True},
##        ], on_save=lambda v: self.update_item(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_item(self, index, values):
##        self.data["vet_health"][index].update({
##            "date": values["date"], "cow_tag": values["cow"],
##            "treatment": values["treatment"], "vet_name": values["vet"],
##            "cost": values["cost"], "notes": values["notes"],
##        })
##        save_data(self.data)
##        show_snackbar("Updated!")
##        self.refresh()
##
##    def delete_item(self, index, *args):
##        item = self.data["vet_health"][index]
##        def confirm():
##            self.data["vet_health"].pop(index)
##            save_data(self.data)
##            show_snackbar("Deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete?", f"Remove health record for {item.get('cow_tag', 'cow')}?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### LABOUR SCREEN
### ============================================================
##
##class LabourScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "LABOUR MANAGEMENT")
##
##        total = sum(safe_float(x.get("amount")) for x in self.data["labour"])
##        content.add_widget(KPICard("TOTAL LABOUR COSTS", format_currency(total), THEME["danger"]))
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        all_items = sorted(self.data["labour"], key=lambda x: x.get("date", ""), reverse=True)
##        items = self.filter_items(all_items, self.search_query, ["date", "worker_name", "task"])
##
##        if not items:
##            msg = "No records found." if self.search_query else "No labour records."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(80)))
##            return
##
##        for item in items[:50]:
##            idx = self.data["labour"].index(item)
##            card = ActionCard(
##                title=f"{item.get('date', '-')} | {item.get('worker_name', 'Unknown')}",
##                subtitle=f"{item.get('task', '-')} | {item.get('days', '-')} days @ KES {item.get('daily_rate', '0')}/day = {format_currency(item.get('amount', '0'))}",
##                on_edit=partial(self.edit_item, idx),
##                on_delete=partial(self.delete_item, idx)
##            )
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        default_rate = self.settings.get("daily_rate_default", "500")
##        dialog, fields = MobileDialog.create("Add Labour Record", [
##            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##            {"key": "worker", "hint": "Worker Name", "icon": "account", "required": True},
##            {"key": "task", "hint": "Task/Role", "icon": "hammer-wrench"},
##            {"key": "days", "hint": "Days Worked", "icon": "calendar-clock", "input_filter": "float", "is_number": True},
##            {"key": "rate", "hint": "Daily Rate (KES)", "icon": "cash", "input_filter": "float", "text": default_rate, "is_number": True},
##        ], on_save=self.save_item)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_item(self, values):
##        days = safe_float(values["days"])
##        rate = safe_float(values["rate"])
##        self.data["labour"].append({
##            "date": values["date"], "worker_name": values["worker"],
##            "task": values["task"], "days": str(days),
##            "daily_rate": str(rate), "amount": str(days * rate),
##        })
##        save_data(self.data)
##        log_activity(self.data, "Labour", f"{values['worker']} - {days} days")
##        show_snackbar("Labour record saved!")
##        self.refresh()
##
##    def edit_item(self, index, *args):
##        item = self.data["labour"][index]
##        dialog, fields = MobileDialog.create("Edit Labour Record", [
##            {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "is_date": True},
##            {"key": "worker", "hint": "Worker", "icon": "account", "text": item.get("worker_name", "")},
##            {"key": "task", "hint": "Task", "icon": "hammer-wrench", "text": item.get("task", "")},
##            {"key": "days", "hint": "Days", "icon": "calendar-clock", "text": str(item.get("days", "")), "input_filter": "float", "is_number": True},
##            {"key": "rate", "hint": "Daily Rate", "icon": "cash", "text": str(item.get("daily_rate", "")), "input_filter": "float", "is_number": True},
##        ], on_save=lambda v: self.update_item(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_item(self, index, values):
##        days = safe_float(values["days"])
##        rate = safe_float(values["rate"])
##        self.data["labour"][index].update({
##            "date": values["date"], "worker_name": values["worker"],
##            "task": values["task"], "days": str(days),
##            "daily_rate": str(rate), "amount": str(days * rate),
##        })
##        save_data(self.data)
##        show_snackbar("Updated!")
##        self.refresh()
##
##    def delete_item(self, index, *args):
##        item = self.data["labour"][index]
##        def confirm():
##            self.data["labour"].pop(index)
##            save_data(self.data)
##            show_snackbar("Deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete?", f"Remove record for {item.get('worker_name', 'worker')}?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### OPERATIONS SCREEN
### ============================================================
##
##class OperationsScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "OPERATIONS")
##
##        total = sum(safe_float(x.get("amount")) for x in self.data["operations"])
##        content.add_widget(KPICard("TOTAL OPERATIONS", format_currency(total), THEME["warning"]))
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        all_items = sorted(self.data["operations"], key=lambda x: x.get("date", ""), reverse=True)
##        items = self.filter_items(all_items, self.search_query, ["date", "category", "description"])
##
##        if not items:
##            msg = "No records found." if self.search_query else "No operation records."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(80)))
##            return
##
##        for item in items[:50]:
##            idx = self.data["operations"].index(item)
##            card = ActionCard(
##                title=f"{item.get('date', '-')} | {item.get('category', 'General')}",
##                subtitle=f"{item.get('description', '-')} | {format_currency(item.get('amount', '0'))}",
##                on_edit=partial(self.edit_item, idx),
##                on_delete=partial(self.delete_item, idx)
##            )
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        dialog, fields = MobileDialog.create("Add Operation Expense", [
##            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##            {"key": "category", "hint": "Category (Fuel/Repair/Utilities/Equipment)", "icon": "folder", "required": True},
##            {"key": "desc", "hint": "Description", "icon": "text"},
##            {"key": "amount", "hint": "Amount (KES)", "icon": "cash", "input_filter": "float", "required": True, "is_number": True},
##        ], on_save=self.save_item)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_item(self, values):
##        self.data["operations"].append({
##            "date": values["date"], "category": values["category"],
##            "description": values["desc"], "amount": values["amount"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Operation", f"{values['category']} - {values['amount']}")
##        show_snackbar("Operation saved!")
##        self.refresh()
##
##    def edit_item(self, index, *args):
##        item = self.data["operations"][index]
##        dialog, fields = MobileDialog.create("Edit Operation", [
##            {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "is_date": True},
##            {"key": "category", "hint": "Category", "icon": "folder", "text": item.get("category", "")},
##            {"key": "desc", "hint": "Description", "icon": "text", "text": item.get("description", "")},
##            {"key": "amount", "hint": "Amount", "icon": "cash", "text": str(item.get("amount", "")), "input_filter": "float", "is_number": True},
##        ], on_save=lambda v: self.update_item(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_item(self, index, values):
##        self.data["operations"][index].update({
##            "date": values["date"], "category": values["category"],
##            "description": values["desc"], "amount": values["amount"],
##        })
##        save_data(self.data)
##        show_snackbar("Updated!")
##        self.refresh()
##
##    def delete_item(self, index, *args):
##        item = self.data["operations"][index]
##        def confirm():
##            self.data["operations"].pop(index)
##            save_data(self.data)
##            show_snackbar("Deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete?", f"Remove {item.get('category', 'this')} operation?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
##
### ============================================================
### CHATS / NOTES SCREEN
### ============================================================
##
##class ChatsScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "FARM NOTES & CHATS")
##
##        total_notes = len(self.data.get("chats", []))
##        content.add_widget(KPICard("TOTAL NOTES", str(total_notes), THEME["info"]))
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(8))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog, icon="message-plus")
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        all_items = sorted(self.data.get("chats", []), key=lambda x: x.get("timestamp", ""), reverse=True)
##        items = self.filter_items(all_items, self.search_query, ["author", "message", "category"])
##
##        if not items:
##            msg = "No notes found." if self.search_query else "No notes yet.\nTap + to add a farm note or chat."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(100)))
##            return
##
##        for item in items:
##            idx = self.data["chats"].index(item)
##            card = ThemedCard(orientation="vertical", size_hint_y=None, height=dp(120), padding=dp(12))
##
##            header = MDBoxLayout(size_hint_y=None, height=dp(24))
##            header.add_widget(MDLabel(
##                text=item.get("author", "Unknown"),
##                theme_text_color="Custom", text_color=THEME["primary"],
##                font_style="Subtitle1", bold=True, size_hint_x=0.6
##            ))
##            header.add_widget(MDLabel(
##                text=item.get("timestamp", "")[:16],
##                theme_text_color="Custom", text_color=THEME["text_secondary"],
##                font_style="Caption", halign="right", size_hint_x=0.4
##            ))
##            card.add_widget(header)
##
##            if item.get("category"):
##                card.add_widget(MDLabel(
##                    text=f"Category: {item['category']}",
##                    theme_text_color="Custom", text_color=THEME["accent"],
##                    font_style="Caption", size_hint_y=None, height=dp(18)
##                ))
##
##            card.add_widget(MDLabel(
##                text=item.get("message", ""),
##                theme_text_color="Custom", text_color=THEME["text"],
##                font_style="Body1", size_hint_y=None, height=dp(50)
##            ))
##
##            btn_row = MDBoxLayout(size_hint_y=None, height=dp(32), spacing=dp(4))
##            btn_row.add_widget(Widget(size_hint_x=0.6))
##            btn_row.add_widget(MDIconButton(
##                icon="pencil", theme_text_color="Custom", text_color=THEME["info"],
##                on_release=partial(self.edit_note, idx), icon_size=dp(18)
##            ))
##            btn_row.add_widget(MDIconButton(
##                icon="delete", theme_text_color="Custom", text_color=THEME["danger"],
##                on_release=partial(self.delete_note, idx), icon_size=dp(18)
##            ))
##            card.add_widget(btn_row)
##
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        dialog, fields = MobileDialog.create("Add Farm Note", [
##            {"key": "author", "hint": "Your Name", "icon": "account", "required": True, "text": self.settings.get("owner_name", "")},
##            {"key": "category", "hint": "Category (General/Task/Reminder)", "icon": "tag", "text": "General"},
##            {"key": "message", "hint": "Message / Note", "icon": "message-text", "required": True, "multiline": True},
##        ], on_save=self.save_note)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_note(self, values):
##        self.data.setdefault("chats", []).insert(0, {
##            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
##            "author": values["author"],
##            "category": values["category"],
##            "message": values["message"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Note added", values["author"])
##        show_snackbar("Note saved!")
##        self.refresh()
##
##    def edit_note(self, index, *args):
##        item = self.data["chats"][index]
##        dialog, fields = MobileDialog.create("Edit Note", [
##            {"key": "author", "hint": "Author", "icon": "account", "text": item.get("author", "")},
##            {"key": "category", "hint": "Category", "icon": "tag", "text": item.get("category", "")},
##            {"key": "message", "hint": "Message", "icon": "message-text", "text": item.get("message", ""), "multiline": True},
##        ], on_save=lambda v: self.update_note(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_note(self, index, values):
##        self.data["chats"][index].update({
##            "author": values["author"],
##            "category": values["category"],
##            "message": values["message"],
##        })
##        save_data(self.data)
##        show_snackbar("Note updated!")
##        self.refresh()
##
##    def delete_note(self, index, *args):
##        def confirm():
##            self.data["chats"].pop(index)
##            save_data(self.data)
##            show_snackbar("Note deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete Note?", "Remove this note permanently?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
### ============================================================
### REPORTS SCREEN
### ============================================================
##
##class ReportsScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "REPORTS & ANALYTICS")
##
##        income = self._calc_income()
##        expenses = self._calc_expenses()
##        net = income - expenses
##
##        content.add_widget(KPICard("NET PROFIT/LOSS", format_currency(net), THEME["success"] if net >= 0 else THEME["danger"]))
##
##        # Expense breakdown pie chart
##        feed = sum(safe_float(x.get("amount")) for x in self.data["feed_expenses"])
##        vet = sum(safe_float(x.get("cost")) for x in self.data["vet_health"])
##        lab = sum(safe_float(x.get("amount")) for x in self.data["labour"])
##        ops = sum(safe_float(x.get("amount")) for x in self.data["operations"])
##
##        if CHARTS_AVAILABLE and expenses > 0:
##            labels = ["Feed", "Vet", "Labour", "Ops"]
##            values = [feed, vet, lab, ops]
##            colors = ['#f39c12', '#e74c3c', '#3498db', '#7f8c8d']
##            chart = create_pie_chart(labels, values, "Expense Breakdown", colors)
##            if chart:
##                try:
##                    chart.size_hint_y = None
##                    chart.height = dp(240)
##                    content.add_widget(chart)
##                except Exception as e:
##                    print(f"Chart display error: {e}")
##
##        content.add_widget(MDLabel(text="EXPENSE BREAKDOWN", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        breakdown = [("Feed", feed, THEME["warning"]), ("Vet & Health", vet, THEME["danger"]),
##                     ("Labour", lab, THEME["info"]), ("Operations", ops, THEME["text_secondary"])]
##        for name, amount, color in breakdown:
##            pct = (amount / expenses * 100) if expenses > 0 else 0
##            content.add_widget(MDLabel(
##                text=f"{name}: {format_currency(amount)} ({pct:.1f}%)",
##                theme_text_color="Custom", text_color=color,
##                font_style="Body1", size_hint_y=None, height=dp(26)
##            ))
##
##        # Monthly income vs expense chart
##        if CHARTS_AVAILABLE:
##            months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
##            inc_vals = []
##            exp_vals = []
##            for i in range(12):
##                m_income = sum(safe_float(x.get("amount")) for x in self.data["milk_sales"] if self._get_month(x.get("date", "")) == i)
##                m_income += sum(safe_float(x.get("amount")) for x in self.data["other_income"] if self._get_month(x.get("date", "")) == i)
##                m_exp = (sum(safe_float(x.get("amount")) for x in self.data["feed_expenses"] if self._get_month(x.get("date", "")) == i) +
##                         sum(safe_float(x.get("cost")) for x in self.data["vet_health"] if self._get_month(x.get("date", "")) == i) +
##                         sum(safe_float(x.get("amount")) for x in self.data["labour"] if self._get_month(x.get("date", "")) == i) +
##                         sum(safe_float(x.get("amount")) for x in self.data["operations"] if self._get_month(x.get("date", "")) == i))
##                inc_vals.append(m_income)
##                exp_vals.append(m_exp)
##
##            try:
##                fig, ax = plt.subplots(figsize=(3.2, 2.4), dpi=100)
##                fig.patch.set_alpha(0)
##                ax.set_facecolor('none')
##                x = range(12)
##                ax.bar([i - 0.2 for i in x], inc_vals, 0.4, label='Income', color='#27ae60', alpha=0.8)
##                ax.bar([i + 0.2 for i in x], exp_vals, 0.4, label='Expense', color='#e74c3c', alpha=0.8)
##                ax.set_xticks(x)
##                ax.set_xticklabels(months, fontsize=6, color='white')
##                ax.tick_params(colors='white', labelsize=7)
##                ax.set_title("Monthly Income vs Expense", color='white', fontsize=10)
##                ax.legend(facecolor='none', edgecolor='white', labelcolor='white', fontsize=7)
##                ax.spines['bottom'].set_color('white')
##                ax.spines['left'].set_color('white')
##                ax.spines['top'].set_visible(False)
##                ax.spines['right'].set_visible(False)
##                ax.grid(True, alpha=0.3, color='white', axis='y')
##                plt.tight_layout()
##                chart = FigureCanvasKivyAgg(fig)
##                chart.size_hint_y = None
##                chart.height = dp(200)
##                content.add_widget(chart)
##            except Exception as e:
##                print(f"Chart display error: {e}")
##
##        # Herd Summary
##        content.add_widget(MDLabel(text="HERD SUMMARY", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        breeds = defaultdict(int)
##        for cow in self.data["herd"]:
##            breeds[cow.get("breed", "Unknown")] += 1
##        if not breeds:
##            content.add_widget(MDLabel(text="No herd data.", theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(40)))
##        else:
##            for breed, count in breeds.items():
##                content.add_widget(MDLabel(
##                    text=f"{breed}: {count} cow(s)",
##                    theme_text_color="Custom", text_color=THEME["text"],
##                    font_style="Body1", size_hint_y=None, height=dp(24)
##                ))
##
##        layout.add_widget(scroll)
##        self.add_widget(layout)
##
##    def _calc_income(self):
##        return (sum(safe_float(x.get("amount")) for x in self.data.get("milk_sales", [])) +
##                sum(safe_float(x.get("amount")) for x in self.data.get("other_income", [])))
##
##    def _calc_expenses(self):
##        return (sum(safe_float(x.get("amount")) for x in self.data.get("feed_expenses", [])) +
##                sum(safe_float(x.get("cost")) for x in self.data.get("vet_health", [])) +
##                sum(safe_float(x.get("amount")) for x in self.data.get("labour", [])) +
##                sum(safe_float(x.get("amount")) for x in self.data.get("operations", [])))
##
##    def _get_month(self, date_str):
##        try:
##            return datetime.strptime(str(date_str), "%Y-%m-%d").month - 1
##        except:
##            return -1
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### SETTINGS SCREEN
### ============================================================
##
##class SettingsScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "SETTINGS & TOOLS")
##
##        # Farm Profile
##        content.add_widget(MDLabel(text="FARM PROFILE", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        profile_card = ThemedCard(orientation="vertical", size_hint_y=None, height=dp(200), spacing=dp(8))
##        self.farm_name_field = MDTextField(hint_text="Farm Name", text=self.settings.get("farm_name", ""), mode="rectangle")
##        self.farm_loc_field = MDTextField(hint_text="Location", text=self.settings.get("farm_location", ""), mode="rectangle")
##        self.owner_field = MDTextField(hint_text="Owner Name", text=self.settings.get("owner_name", ""), mode="rectangle")
##        profile_card.add_widget(self.farm_name_field)
##        profile_card.add_widget(self.farm_loc_field)
##        profile_card.add_widget(self.owner_field)
##        profile_card.add_widget(MDRaisedButton(text="SAVE PROFILE", md_bg_color=THEME["success"],
##            on_release=self.save_profile, size_hint=(1, None), height=dp(40)))
##        content.add_widget(profile_card)
##
##        # Defaults
##        content.add_widget(MDLabel(text="DEFAULTS", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        defaults_card = ThemedCard(orientation="vertical", size_hint_y=None, height=dp(160), spacing=dp(8))
##        self.price_field = MDTextField(hint_text="Default Milk Price/Liter (KES)",
##            text=self.settings.get("milk_price_default", "60"), mode="rectangle", input_filter="float")
##        self.rate_field = MDTextField(hint_text="Default Daily Labour Rate (KES)",
##            text=self.settings.get("daily_rate_default", "500"), mode="rectangle", input_filter="float")
##        defaults_card.add_widget(self.price_field)
##        defaults_card.add_widget(self.rate_field)
##        defaults_card.add_widget(MDRaisedButton(text="SAVE DEFAULTS", md_bg_color=THEME["success"],
##            on_release=self.save_defaults, size_hint=(1, None), height=dp(40)))
##        content.add_widget(defaults_card)
##
##        # Data Management
##        content.add_widget(MDLabel(text="DATA MANAGEMENT", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        data_card = ThemedCard(orientation="vertical", size_hint_y=None, height=dp(220), spacing=dp(8))
##        data_card.add_widget(MDRaisedButton(text="EXPORT TO CSV", md_bg_color=THEME["info"],
##            on_release=self.export_csv, size_hint=(1, None), height=dp(44)))
##        data_card.add_widget(MDRaisedButton(text="EXPORT TO JSON", md_bg_color=THEME["info"],
##            on_release=self.export_json, size_hint=(1, None), height=dp(44)))
##        data_card.add_widget(MDRaisedButton(text="VIEW BACKUPS", md_bg_color=THEME["accent"],
##            on_release=self.view_backups, size_hint=(1, None), height=dp(44)))
##        data_card.add_widget(MDRaisedButton(text="CLEAR ALL DATA", md_bg_color=THEME["danger"],
##            on_release=self.clear_all_data, size_hint=(1, None), height=dp(44)))
##        content.add_widget(data_card)
##
##        # App Info
##        content.add_widget(MDLabel(text="DAIRY FARM MS PRO v2.0", theme_text_color="Custom",
##            text_color=THEME["text_secondary"], font_style="Caption",
##            halign="center", size_hint_y=None, height=dp(30)))
##
##        layout.add_widget(scroll)
##        self.add_widget(layout)
##
##    def save_profile(self, *args):
##        self.settings["farm_name"] = self.farm_name_field.text
##        self.settings["farm_location"] = self.farm_loc_field.text
##        self.settings["owner_name"] = self.owner_field.text
##        save_settings(self.settings)
##        show_snackbar("Profile saved!")
##
##    def save_defaults(self, *args):
##        self.settings["milk_price_default"] = self.price_field.text or "60"
##        self.settings["daily_rate_default"] = self.rate_field.text or "500"
##        save_settings(self.settings)
##        show_snackbar("Defaults saved!")
##
##    def export_csv(self, *args):
##        try:
##            import csv
##            export_dir = os.path.join(get_data_dir(), "exports")
##            os.makedirs(export_dir, exist_ok=True)
##            filename = os.path.join(export_dir, f"dairy_export_{datetime.now().strftime('%Y%m%d')}.csv")
##
##            with open(filename, 'w', newline='', encoding='utf-8') as f:
##                writer = csv.writer(f)
##                writer.writerow(["Type", "Date", "Category", "Description", "Amount", "Details"])
##                for item in self.data.get("milk_sales", []):
##                    writer.writerow(["Milk Sale", item.get("date"), "", item.get("buyer"), item.get("amount"), f"{item.get('liters')}L"])
##                for item in self.data.get("other_income", []):
##                    writer.writerow(["Other Income", item.get("date"), item.get("source"), item.get("notes"), item.get("amount"), ""])
##                for item in self.data.get("feed_expenses", []):
##                    writer.writerow(["Feed", item.get("date"), item.get("feed_type"), item.get("supplier"), item.get("amount"), ""])
##                for item in self.data.get("vet_health", []):
##                    writer.writerow(["Vet", item.get("date"), item.get("treatment"), item.get("cow_tag"), item.get("cost"), ""])
##                for item in self.data.get("labour", []):
##                    writer.writerow(["Labour", item.get("date"), item.get("task"), item.get("worker_name"), item.get("amount"), ""])
##                for item in self.data.get("operations", []):
##                    writer.writerow(["Operations", item.get("date"), item.get("category"), item.get("description"), item.get("amount"), ""])
##
##            show_snackbar(f"Exported to {filename}")
##        except Exception as e:
##            show_snackbar(f"Export failed: {str(e)}", THEME["danger"])
##
##    def export_json(self, *args):
##        try:
##            export_dir = os.path.join(get_data_dir(), "exports")
##            os.makedirs(export_dir, exist_ok=True)
##            filename = os.path.join(export_dir, f"dairy_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
##            with open(filename, "w", encoding="utf-8") as f:
##                json.dump(self.data, f, indent=2, default=str)
##            show_snackbar(f"JSON exported!")
##        except Exception as e:
##            show_snackbar(f"Export failed: {str(e)}", THEME["danger"])
##
##    def view_backups(self, *args):
##        backups = sorted([f for f in os.listdir(get_backup_dir()) if f.startswith("backup_")], reverse=True)
##        content = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(4))
##        content.height = dp(20) + len(backups) * dp(36)
##
##        if not backups:
##            content.add_widget(MDLabel(text="No backups found.", theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(40)))
##        else:
##            for b in backups[:20]:
##                content.add_widget(MDLabel(text=b, theme_text_color="Custom",
##                    text_color=THEME["text"], font_style="Caption", size_hint_y=None, height=dp(32)))
##
##        dialog = MDDialog(title="Backups", type="custom", content_cls=content,
##            buttons=[MDFlatButton(text="CLOSE", on_release=lambda x: dialog.dismiss())])
##        dialog.open()
##
##    def clear_all_data(self, *args):
##        def confirm():
##            self.data = {
##                "farm": {"name": "My Dairy Farm", "location": "", "owner": ""},
##                "herd": [], "milk_production": [], "milk_sales": [],
##                "other_income": [], "feed_expenses": [], "vet_health": [],
##                "labour": [], "operations": [], "assets": [], "breeding": [], "activity_log": [],
##            }
##            save_data(self.data)
##            show_snackbar("All data cleared!")
##            self.refresh()
##        ConfirmDialog.show("CLEAR ALL DATA?", "This will delete EVERYTHING. This cannot be undone!", confirm, "CLEAR", THEME["danger"])
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### MAIN APP
### ============================================================
##
##class DairyFarmApp(MDApp):
##    def build(self):
##        self.theme_cls.primary_palette = "Green"
##        self.theme_cls.theme_style = "Light"
##
##        # Load settings for theme
##        self.settings = load_settings()
##        if self.settings.get("dark_mode"):
##            self.theme_cls.theme_style = "Dark"
##            global THEME
##            THEME = DARK_THEME
##
##        sm = MDScreenManager(transition=SlideTransition(duration=0.2))
##        sm.add_widget(DashboardScreen(name="dashboard"))
##        sm.add_widget(HerdScreen(name="herd"))
##        sm.add_widget(MilkProductionScreen(name="milk_production"))
##        sm.add_widget(MilkSalesScreen(name="milk_sales"))
##        sm.add_widget(FeedExpensesScreen(name="feed_expenses"))
##        sm.add_widget(VetHealthScreen(name="vet_health"))
##        sm.add_widget(LabourScreen(name="labour"))
##        sm.add_widget(OperationsScreen(name="operations"))
##        sm.add_widget(ReportsScreen(name="reports"))
##        sm.add_widget(ChatsScreen(name="chats"))
##        sm.add_widget(SettingsScreen(name="settings"))
##
##        self.sm = sm
##
##        # Main layout with bottom navigation
##        layout = MDBoxLayout(orientation="vertical")
##
##        # Toolbar
##        self.toolbar = MDTopAppBar(
##            title="Dairy Farm MS",
##            elevation=4,
##            pos_hint={"top": 1},
##            md_bg_color=THEME["primary"],
##            left_action_items=[["menu", lambda x: self.nav_drawer.set_state("open")]],
##            right_action_items=[["export-variant", lambda x: self.show_export_menu()]],
##        )
##        layout.add_widget(self.toolbar)
##        layout.add_widget(sm)
##
##        # Custom Bottom Navigation Bar (reliable across KivyMD versions)
##        self.bottom_bar = MDBoxLayout(
##            orientation="horizontal",
##            size_hint_y=None,
##            height=dp(56),
##            md_bg_color=THEME["card"],
##            padding=[dp(4), dp(4), dp(4), dp(4)],
##            spacing=dp(2),
##        )
##
##        # Add shadow line above bottom bar
##        with self.bottom_bar.canvas.before:
##            Color(*THEME["divider"])
##            Rectangle(pos=(0, dp(56)), size=(Window.width, dp(1)))
##
##        self.nav_buttons = {}
##        nav_items = [
##            ("dashboard", "Dashboard", "chart-bar"),
##            ("herd", "Herd", "cow"),
##            ("milk_production", "Milk", "cup-water"),
##            ("milk_sales", "Sales", "cash-multiple"),
##            ("reports", "Reports", "file-chart"),
##        ]
##
##        for screen_name, label, icon in nav_items:
##            btn = MDBoxLayout(orientation="vertical", size_hint_x=1)
##            btn_icon = MDIconButton(
##                icon=icon,
##                theme_text_color="Custom",
##                text_color=THEME["text_secondary"],
##                pos_hint={"center_x": 0.5},
##                icon_size=dp(22),
##            )
##            btn_label = MDLabel(
##                text=label,
##                theme_text_color="Custom",
##                text_color=THEME["text_secondary"],
##                font_style="Caption",
##                halign="center",
##                size_hint_y=None,
##                height=dp(16),
##            )
##            btn.add_widget(btn_icon)
##            btn.add_widget(btn_label)
##
##            # Make the whole box layout clickable
##            btn.bind(on_touch_down=partial(self.on_bottom_button_touch, screen_name, btn_icon, btn_label))
##
##            self.nav_buttons[screen_name] = (btn, btn_icon, btn_label)
##            self.bottom_bar.add_widget(btn)
##
##        layout.add_widget(self.bottom_bar)
##        self.bottom_nav = self.bottom_bar  # alias for compatibility
##
##        # Side Navigation Drawer (for less-used screens)
##        self.nav_drawer = MDNavigationDrawer(
##            md_bg_color=THEME["bg"],
##            elevation=4,
##        )
##
##        nav_content = MDBoxLayout(orientation="vertical", padding=dp(8), spacing=dp(4))
##        nav_content.add_widget(MDLabel(
##            text="MENU",
##            theme_text_color="Custom",
##            text_color=THEME["primary"],
##            font_style="H6",
##            size_hint_y=None,
##            height=dp(40),
##            halign="center"
##        ))
##
##        drawer_items = [
##            ("feed_expenses", "Feed Expenses", "corn"),
##            ("vet_health", "Vet & Health", "needle"),
##            ("labour", "Labour", "account-hard-hat"),
##            ("operations", "Operations", "cogs"),
##            ("chats", "Farm Notes", "message-text"),
##            ("settings", "Settings & Tools", "cog"),
##        ]
##
##        nav_list = MDList()
##        for screen_name, label, icon in drawer_items:
##            item = OneLineIconListItem(
##                IconLeftWidget(icon=icon),
##                text=label,
##                on_release=lambda x, s=screen_name: self.switch_screen(s)
##            )
##            nav_list.add_widget(item)
##
##        nav_content.add_widget(nav_list)
##        self.nav_drawer.add_widget(nav_content)
##
##        root = MDScreen()
##        root.add_widget(layout)
##        root.add_widget(self.nav_drawer)
##
##        return root
##
##    def show_export_menu(self, *args):
##        menu_items = [
##            {"text": "Export to CSV", "icon": "file-delimited", "on_release": lambda: self.do_export("csv")},
##            {"text": "Export to JSON", "icon": "code-json", "on_release": lambda: self.do_export("json")},
##        ]
##        self.export_menu = MDDropdownMenu(
##            caller=self.toolbar,
##            items=menu_items,
##            width_mult=3,
##        )
##        self.export_menu.open()
##
##    def do_export(self, fmt):
##        self.export_menu.dismiss()
##        data = load_data()
##        export_dir = os.path.join(get_data_dir(), "exports")
##        os.makedirs(export_dir, exist_ok=True)
##        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
##
##        if fmt == "csv":
##            try:
##                filename = os.path.join(export_dir, f"dairy_export_{timestamp}.csv")
##                with open(filename, 'w', newline='', encoding='utf-8') as f:
##                    writer = csv.writer(f)
##                    writer.writerow(["Type", "Date", "Category", "Description", "Amount", "Details"])
##                    for item in data.get("milk_sales", []):
##                        writer.writerow(["Milk Sale", item.get("date"), "", item.get("buyer"), item.get("amount"), f"{item.get('liters')}L"])
##                    for item in data.get("other_income", []):
##                        writer.writerow(["Other Income", item.get("date"), item.get("source"), item.get("notes"), item.get("amount"), ""])
##                    for item in data.get("feed_expenses", []):
##                        writer.writerow(["Feed", item.get("date"), item.get("feed_type"), item.get("supplier"), item.get("amount"), ""])
##                    for item in data.get("vet_health", []):
##                        writer.writerow(["Vet", item.get("date"), item.get("treatment"), item.get("cow_tag"), item.get("cost"), ""])
##                    for item in data.get("labour", []):
##                        writer.writerow(["Labour", item.get("date"), item.get("task"), item.get("worker_name"), item.get("amount"), ""])
##                    for item in data.get("operations", []):
##                        writer.writerow(["Operations", item.get("date"), item.get("category"), item.get("description"), item.get("amount"), ""])
##                show_snackbar(f"CSV exported: {filename}")
##            except Exception as e:
##                show_snackbar(f"Export failed: {str(e)}", THEME["danger"])
##        else:
##            try:
##                filename = os.path.join(export_dir, f"dairy_backup_{timestamp}.json")
##                with open(filename, "w", encoding="utf-8") as f:
##                    json.dump(data, f, indent=2, default=str)
##                show_snackbar(f"JSON exported!")
##            except Exception as e:
##                show_snackbar(f"Export failed: {str(e)}", THEME["danger"])
##
##    def on_bottom_button_touch(self, screen_name, btn_icon, btn_label, widget, touch):
##        if widget.collide_point(*touch.pos):
##            if touch.button == "left" and not touch.is_double_tap:
##                self.switch_screen(screen_name)
##                return True
##        return False
##
##    def update_bottom_nav_highlight(self, active_name):
##        """Highlight the active bottom nav button."""
##        for name, (btn, icon, label) in self.nav_buttons.items():
##            if name == active_name:
##                icon.text_color = THEME["primary"]
##                label.text_color = THEME["primary"]
##            else:
##                icon.text_color = THEME["text_secondary"]
##                label.text_color = THEME["text_secondary"]
##
##    def switch_screen(self, name):
##        # Refresh the target screen
##        for screen in self.sm.screens:
##            if screen.name == name and hasattr(screen, "refresh"):
##                screen.refresh()
##
##        self.sm.current = name
##        self.toolbar.title = name.replace("_", " ").title()
##        self.nav_drawer.set_state("close")
##
##        # Update bottom nav highlight
##        self.update_bottom_nav_highlight(name)
##
##
##if __name__ == "__main__":
##    DairyFarmApp().run()
#################################################################################################################################

##"""
##Dairy Farm MS Pro - Fully Functional Mobile Version
##====================================================
##Production-ready dairy farm management for Android/iOS.
##
##Features:
##- Bottom Navigation (mobile-native)
##- Data Validation & Error Handling
##- Search & Filter across all modules
##- Charts & Analytics (Matplotlib)
##- CSV/JSON Export & Share
##- Farm Settings & Profile
##- Dark Mode Support
##- Auto-backup
##- Cow Detail Profiles
##- Production Trends
##- Expense Breakdown Charts
##- Income vs Expense Analytics
##- Recent Activity Feed
##- Responsive Touch Targets
##
##Dependencies:
##    pip install kivymd matplotlib
##
##Build Android:
##    buildozer android debug
##
##Run Desktop:
##    python dairy_farm_mobile_pro.py
##"""
##
##import json
##try:
##    from openpyxl import Workbook
##    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
##    EXCEL_AVAILABLE = True
##except ImportError:
##    EXCEL_AVAILABLE = False
##import os
##import csv
##import io
##import base64
##from datetime import datetime, timedelta
##from functools import partial
##from collections import defaultdict
##
##from kivy.config import Config
##nConfig = Config  # alias to avoid issues
##nConfig.set('graphics', 'width', '360')
##nConfig.set('graphics', 'height', '640')
##
##from kivy.core.window import Window
##from kivy.metrics import dp
##from kivy.properties import ObjectProperty, StringProperty, BooleanProperty
##from kivy.clock import Clock
##from kivy.uix.boxlayout import BoxLayout
##from kivy.uix.scrollview import ScrollView
##from kivy.uix.anchorlayout import AnchorLayout
##from kivy.animation import Animation
##from kivy.graphics import Color, Rectangle, RoundedRectangle
##
##from kivymd.app import MDApp
##from kivymd.uix.boxlayout import MDBoxLayout
##from kivymd.uix.button import (
##    MDRaisedButton, MDIconButton, MDFlatButton, 
##    MDFloatingActionButton, MDRoundFlatButton
##)
##from kivymd.uix.card import MDCard
##from kivymd.uix.chip import MDChip
##from kivymd.uix.dialog import MDDialog
##from kivymd.uix.label import MDLabel
##from kivymd.uix.list import (
##    MDList, OneLineIconListItem, TwoLineAvatarIconListItem,
##    IconLeftWidget, IconRightWidget
##)
##from kivymd.uix.menu import MDDropdownMenu
##from kivymd.uix.pickers import MDDatePicker
##from kivymd.uix.screen import MDScreen
##from kivymd.uix.screenmanager import MDScreenManager
##from kivy.uix.screenmanager import SlideTransition
##from kivymd.uix.textfield import MDTextField
##from kivymd.uix.toolbar import MDTopAppBar
### from kivymd.uix.bottomnavigation import MDBottomNavigation, MDBottomNavigationItem  # Using custom bottom bar
##from kivymd.uix.navigationdrawer import MDNavigationDrawer
##from kivymd.uix.snackbar import MDSnackbar
##from kivymd.uix.selectioncontrol import MDCheckbox
##from kivymd.uix.progressbar import MDProgressBar
##from kivymd.uix.gridlayout import MDGridLayout
##
### Chart imports
### Charts disabled: kivy-garden matplotlib is incompatible with Kivy 2.3.1
### All chart functions return None gracefully
##CHARTS_AVAILABLE = False
##
### Keep stub imports for the chart helper functions to reference
### but they will never create actual widgets
##try:
##    import matplotlib
##    matplotlib.use('Agg')
##    import matplotlib.pyplot as plt
##    from matplotlib.dates import DateFormatter
##except Exception:
##    pass
##
### ============================================================
### THEME & CONFIG
### ============================================================
##
##THEME = {
##    "primary": [0.102, 0.361, 0.227, 1],      # #1a5c3a
##    "accent": [0.176, 0.541, 0.369, 1],       # #2d8a5e
##    "light": [0.302, 0.722, 0.478, 1],        # #4db87a
##    "bg": [0.941, 0.957, 0.941, 1],           # #f0f4f0
##    "card": [1, 1, 1, 1],
##    "danger": [0.906, 0.298, 0.235, 1],       # #e74c3c
##    "warning": [0.953, 0.612, 0.071, 1],      # #f39c12
##    "success": [0.153, 0.682, 0.376, 1],      # #27ae60
##    "info": [0.204, 0.596, 0.859, 1],         # #3498db
##    "text": [0.173, 0.243, 0.314, 1],         # #2c3e50
##    "text_secondary": [0.498, 0.549, 0.553, 1], # #7f8c8d
##    "divider": [0.85, 0.85, 0.85, 1],
##}
##
##DARK_THEME = {
##    "primary": [0.302, 0.722, 0.478, 1],
##    "accent": [0.176, 0.541, 0.369, 1],
##    "light": [0.102, 0.361, 0.227, 1],
##    "bg": [0.12, 0.12, 0.12, 1],
##    "card": [0.18, 0.18, 0.18, 1],
##    "danger": [0.906, 0.298, 0.235, 1],
##    "warning": [0.953, 0.612, 0.071, 1],
##    "success": [0.153, 0.682, 0.376, 1],
##    "info": [0.204, 0.596, 0.859, 1],
##    "text": [0.9, 0.9, 0.9, 1],
##    "text_secondary": [0.6, 0.6, 0.6, 1],
##    "divider": [0.3, 0.3, 0.3, 1],
##}
##
##CURRENCY = "KES"
##
### ============================================================
### DATA LAYER (Robust & Validated)
### ============================================================
##
##def get_data_dir():
##    try:
##        app = MDApp.get_running_app()
##        if app:
##            return app.user_data_dir
##    except:
##        pass
##    base = os.path.expanduser("~/.local/share/DairyFarmMSPro")
##    os.makedirs(base, exist_ok=True)
##    return base
##
##
##def get_settings_path():
##    return os.path.join(get_data_dir(), "settings.json")
##
##
##def get_data_path():
##    return os.path.join(get_data_dir(), "dairy_farm_data.json")
##
##
##def get_backup_dir():
##    bd = os.path.join(get_data_dir(), "backups")
##    os.makedirs(bd, exist_ok=True)
##    return bd
##
##
##def load_settings():
##    path = get_settings_path()
##    defaults = {
##        "farm_name": "My Dairy Farm",
##        "farm_location": "",
##        "owner_name": "",
##        "currency": "KES",
##        "dark_mode": False,
##        "auto_backup": True,
##        "milk_price_default": "60",
##        "daily_rate_default": "500",
##    }
##    if os.path.exists(path):
##        try:
##            with open(path, "r", encoding="utf-8") as f:
##                loaded = json.load(f)
##                defaults.update(loaded)
##        except:
##            pass
##    return defaults
##
##
##def save_settings(settings):
##    with open(get_settings_path(), "w", encoding="utf-8") as f:
##        json.dump(settings, f, indent=2)
##
##
##def load_data():
##    path = get_data_path()
##    defaults = {
##        "farm": {"name": "My Dairy Farm", "location": "", "owner": ""},
##        "herd": [],
##        "milk_production": [],
##        "milk_sales": [],
##        "other_income": [],
##        "feed_expenses": [],
##        "vet_health": [],
##        "labour": [],
##        "operations": [],
##        "assets": [],
##        "breeding": [],
##        "activity_log": [],
##    }
##    if os.path.exists(path):
##        try:
##            with open(path, "r", encoding="utf-8") as f:
##                loaded = json.load(f)
##                # Merge with defaults for new fields
##                for k, v in defaults.items():
##                    if k not in loaded:
##                        loaded[k] = v
##                return loaded
##        except Exception as e:
##            print(f"Data load error: {e}")
##    return defaults
##
##
##def save_data(data):
##    try:
##        with open(get_data_path(), "w", encoding="utf-8") as f:
##            json.dump(data, f, indent=2, default=str)
##        # Auto-backup
##        settings = load_settings()
##        if settings.get("auto_backup", True):
##            backup_file = os.path.join(get_backup_dir(), f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
##            try:
##                with open(backup_file, "w", encoding="utf-8") as f:
##                    json.dump(data, f, indent=2, default=str)
##                # Keep only last 10 backups
##                backups = sorted([f for f in os.listdir(get_backup_dir()) if f.startswith("backup_")])
##                for old in backups[:-10]:
##                    os.remove(os.path.join(get_backup_dir(), old))
##            except:
##                pass
##        return True
##    except Exception as e:
##        print(f"Save error: {e}")
##        return False
##
##
##def log_activity(data, action, details=""):
##    data.setdefault("activity_log", []).insert(0, {
##        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
##        "action": action,
##        "details": details,
##    })
##    data["activity_log"] = data["activity_log"][:100]  # Keep last 100
##    save_data(data)
##
##
##def validate_date(date_str):
##    try:
##        datetime.strptime(str(date_str), "%Y-%m-%d")
##        return True
##    except:
##        return False
##
##
##def safe_float(val, default=0.0):
##    try:
##        return float(val) if val not in (None, "", "None") else default
##    except:
##        return default
##
##
##def format_currency(amount):
##    try:
##        return f"{CURRENCY} {float(amount):,.0f}"
##    except:
##        return f"{CURRENCY} 0"
##
##
### ============================================================
### CHART HELPERS
### ============================================================
##
##def create_pie_chart(labels, values, title, colors_list=None):
##    if not CHARTS_AVAILABLE:
##        return None
##    try:
##        fig, ax = plt.subplots(figsize=(3.2, 3.2), dpi=100)
##        fig.patch.set_alpha(0)
##        ax.set_facecolor('none')
##        if colors_list:
##            ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90, colors=colors_list,
##                   textprops={'color': 'white', 'fontsize': 8})
##        else:
##            ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90,
##                   textprops={'color': 'white', 'fontsize': 8})
##        ax.set_title(title, color='white', fontsize=10, pad=10)
##        plt.tight_layout()
##        canvas = FigureCanvasKivyAgg(fig)
##        return canvas
##    except Exception as e:
##        print(f"Chart error: {e}")
##        return None
##
##
##def create_line_chart(dates, values, title, ylabel="Amount (KES)"):
##    if not CHARTS_AVAILABLE or not dates:
##        return None
##    try:
##        fig, ax = plt.subplots(figsize=(3.2, 2.4), dpi=100)
##        fig.patch.set_alpha(0)
##        ax.set_facecolor('none')
##        ax.plot(dates, values, marker='o', linewidth=2, markersize=4, color='#4db87a')
##        ax.set_title(title, color='white', fontsize=10)
##        ax.set_ylabel(ylabel, color='white', fontsize=8)
##        ax.tick_params(colors='white', labelsize=7)
##        ax.spines['bottom'].set_color('white')
##        ax.spines['left'].set_color('white')
##        ax.spines['top'].set_visible(False)
##        ax.spines['right'].set_visible(False)
##        ax.grid(True, alpha=0.3, color='white')
##        plt.tight_layout()
##        return FigureCanvasKivyAgg(fig)
##    except Exception as e:
##        print(f"Chart error: {e}")
##        return None
##
##
##def create_bar_chart(labels, values, title, color='#2d8a5e'):
##    if not CHARTS_AVAILABLE or not labels:
##        return None
##    try:
##        fig, ax = plt.subplots(figsize=(3.2, 2.4), dpi=100)
##        fig.patch.set_alpha(0)
##        ax.set_facecolor('none')
##        bars = ax.bar(labels, values, color=color, alpha=0.8)
##        ax.set_title(title, color='white', fontsize=10)
##        ax.tick_params(colors='white', labelsize=7)
##        ax.spines['bottom'].set_color('white')
##        ax.spines['left'].set_color('white')
##        ax.spines['top'].set_visible(False)
##        ax.spines['right'].set_visible(False)
##        ax.grid(True, alpha=0.3, color='white', axis='y')
##        plt.tight_layout()
##        return FigureCanvasKivyAgg(fig)
##    except Exception as e:
##        print(f"Chart error: {e}")
##        return None
##
##
### ============================================================
### UI HELPERS
### ============================================================
##
##class ThemedCard(MDCard):
##    """Enhanced card with theme support."""
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.elevation = 2
##        self.radius = [dp(12),]
##        self.padding = dp(12)
##        self.md_bg_color = THEME["card"]
##
##
##class KPICard(ThemedCard):
##    def __init__(self, title, value, color, subtitle="", **kwargs):
##        super().__init__(**kwargs)
##        self.orientation = "vertical"
##        self.size_hint_y = None
##        self.height = dp(100)
##        self.add_widget(MDLabel(
##            text=title, theme_text_color="Custom",
##            text_color=THEME["text_secondary"], font_style="Caption",
##            halign="center", size_hint_y=None, height=dp(18)
##        ))
##        self.add_widget(MDLabel(
##            text=value, theme_text_color="Custom",
##            text_color=color, font_style="H5",
##            halign="center", bold=True, size_hint_y=None, height=dp(32)
##        ))
##        if subtitle:
##            self.add_widget(MDLabel(
##                text=subtitle, theme_text_color="Custom",
##                text_color=THEME["text_secondary"], font_style="Caption",
##                halign="center", size_hint_y=None, height=dp(18)
##            ))
##
##
##class ActionCard(ThemedCard):
##    def __init__(self, title, subtitle, on_edit=None, on_delete=None, on_tap=None, **kwargs):
##        super().__init__(**kwargs)
##        self.orientation = "vertical"
##        self.size_hint_y = None
##        self.height = dp(90)
##
##        main = MDBoxLayout()
##        text_area = MDBoxLayout(orientation="vertical", size_hint_x=0.78)
##        text_area.add_widget(MDLabel(
##            text=title, theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="Subtitle1",
##            bold=True, size_hint_y=None, height=dp(22)
##        ))
##        text_area.add_widget(MDLabel(
##            text=subtitle, theme_text_color="Secondary",
##            font_style="Caption", size_hint_y=None, height=dp(36)
##        ))
##        main.add_widget(text_area)
##
##        btn_area = MDBoxLayout(size_hint_x=0.22, spacing=dp(2))
##        if on_edit:
##            btn_area.add_widget(MDIconButton(
##                icon="pencil", theme_text_color="Custom",
##                text_color=THEME["info"], on_release=on_edit,
##                icon_size=dp(18)
##            ))
##        if on_delete:
##            btn_area.add_widget(MDIconButton(
##                icon="delete", theme_text_color="Custom",
##                text_color=THEME["danger"], on_release=on_delete,
##                icon_size=dp(18)
##            ))
##        main.add_widget(btn_area)
##        self.add_widget(main)
##
##        if on_tap:
##            self.bind(on_release=on_tap)
##
##
##class SearchBar(MDBoxLayout):
##    def __init__(self, on_search, **kwargs):
##        super().__init__(**kwargs)
##        self.size_hint_y = None
##        self.height = dp(50)
##        self.padding = [dp(4), dp(4), dp(4), dp(4)]
##        self.md_bg_color = THEME["card"]
##        self.radius = [dp(8),]
##
##        self.search_field = MDTextField(
##            hint_text="Search...",
##            icon_right="magnify",
##            mode="rectangle",
##            size_hint_x=0.85,
##            height=dp(40)
##        )
##        self.search_field.bind(text=lambda inst, val: on_search(val))
##        self.add_widget(self.search_field)
##
##        clear_btn = MDIconButton(
##            icon="close-circle", theme_text_color="Custom",
##            text_color=THEME["text_secondary"],
##            on_release=lambda x: [setattr(self.search_field, "text", ""), on_search("")],
##            size_hint_x=0.15
##        )
##        self.add_widget(clear_btn)
##
##
##class MobileDialog:
##    @staticmethod
##    def create(title, fields, on_save, on_cancel=None, size_hint=(0.92, None)):
##        content = MDBoxLayout(orientation="vertical", spacing=dp(6), size_hint_y=None,
##                              padding=[dp(8), dp(8), dp(8), dp(8)])
##        field_refs = {}
##        total_h = dp(16)
##
##        for field in fields:
##            h = dp(70) if field.get("multiline") else dp(55)
##            tf = MDTextField(
##                hint_text=field.get("hint", ""),
##                icon_right=field.get("icon", ""),
##                multiline=field.get("multiline", False),
##                mode="rectangle",
##                size_hint_y=None, height=h,
##                text=field.get("text", ""),
##                input_filter=field.get("input_filter", None),
##            )
##            field_refs[field.get("key", field["hint"])] = tf
##            content.add_widget(tf)
##            total_h += h + dp(6)
##
##        content.height = total_h
##
##        def _save(*args):
##            vals = {k: v.text for k, v in field_refs.items()}
##            # Validation
##            errors = []
##            for field in fields:
##                key = field.get("key", field["hint"])
##                if field.get("required") and not vals.get(key, "").strip():
##                    errors.append(f"{field['hint']} is required")
##                if field.get("is_date") and vals.get(key) and not validate_date(vals[key]):
##                    errors.append(f"{field['hint']} must be YYYY-MM-DD")
##                if field.get("is_number") and vals.get(key):
##                    try:
##                        float(vals[key])
##                    except:
##                        errors.append(f"{field['hint']} must be a number")
##
##            if errors:
##                show_snackbar("; ".join(errors[:2]), THEME["danger"])
##                return
##            on_save(vals)
##            dialog.dismiss()
##
##        def _cancel(*args):
##            if on_cancel: on_cancel()
##            dialog.dismiss()
##
##        dialog = MDDialog(
##            title=title, type="custom", content_cls=content,
##            size_hint=size_hint,
##            buttons=[
##                MDFlatButton(text="CANCEL", on_release=_cancel),
##                MDRaisedButton(text="SAVE", md_bg_color=THEME["success"], on_release=_save)
##            ],
##        )
##        return dialog, field_refs
##
##
##def show_snackbar(text, color=THEME["success"], duration=2):
##    try:
##        # KivyMD 2.0+ style
##        sb = MDSnackbar(
##            MDLabel(text=text, theme_text_color="Custom", text_color=[1,1,1,1], font_style="Body2"),
##            md_bg_color=color, duration=duration,
##        )
##    except TypeError:
##        # KivyMD 1.2.0 style
##        sb = MDSnackbar(text=text, md_bg_color=color, duration=duration)
##    sb.open()
##
##
##class ConfirmDialog:
##    @staticmethod
##    def show(title, text, on_confirm, confirm_text="DELETE", confirm_color=None):
##        if confirm_color is None:
##            confirm_color = THEME["danger"]
##        dialog = MDDialog(
##            title=title, text=text,
##            buttons=[
##                MDFlatButton(text="CANCEL", on_release=lambda x: dialog.dismiss()),
##                MDRaisedButton(text=confirm_text, md_bg_color=confirm_color,
##                               on_release=lambda x: [dialog.dismiss(), on_confirm()])
##            ]
##        )
##        dialog.open()
##
##
### ============================================================
### BASE SCREEN
### ============================================================
##
##class BaseScreen(MDScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.data = load_data()
##        self.settings = load_settings()
##        self.dialog = None
##        self.search_query = ""
##
##    def reload_data(self):
##        self.data = load_data()
##        self.settings = load_settings()
##
##    def refresh(self):
##        pass
##
##    def create_scroll_layout(self):
##        scroll = ScrollView()
##        content = MDBoxLayout(
##            orientation="vertical", size_hint_y=None,
##            padding=[dp(12), dp(8), dp(12), dp(80)],
##            spacing=dp(8)
##        )
##        content.bind(minimum_height=content.setter("height"))
##        scroll.add_widget(content)
##        return scroll, content
##
##    def add_screen_title(self, content, title):
##        content.add_widget(MDLabel(
##            text=title, theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H5",
##            halign="center", size_hint_y=None, height=dp(40), bold=True
##        ))
##
##    def add_fab(self, layout, callback, icon="plus"):
##        fab = MDFloatingActionButton(
##            icon=icon, md_bg_color=THEME["accent"],
##            pos_hint={"right": 0.95, "y": 0.04},
##            on_release=callback, elevation=4
##        )
##        layout.add_widget(fab)
##        return fab
##
##    def filter_items(self, items, query, keys):
##        if not query:
##            return items
##        q = query.lower()
##        return [item for item in items if any(q in str(item.get(k, "")).lower() for k in keys)]
##
##
### ============================================================
### DASHBOARD SCREEN
### ============================================================
##
##class DashboardScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##
##        # Header
##        farm_name = self.settings.get("farm_name", "My Dairy Farm")
##        content.add_widget(MDLabel(
##            text=farm_name.upper(), theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H4",
##            halign="center", size_hint_y=None, height=dp(44), bold=True
##        ))
##        content.add_widget(MDLabel(
##            text=datetime.now().strftime("%A, %d %B %Y"),
##            theme_text_color="Custom", text_color=THEME["text_secondary"],
##            font_style="Caption", halign="center", size_hint_y=None, height=dp(22)
##        ))
##
##        # KPIs
##        income = self._calc_income()
##        expenses = self._calc_expenses()
##        net = income - expenses
##        milking = len([c for c in self.data["herd"] if c.get("status", "").lower() == "milking"])
##        total_herd = len(self.data["herd"])
##
##        grid = MDGridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(220))
##        grid.add_widget(KPICard("INCOME", format_currency(income), THEME["success"]))
##        grid.add_widget(KPICard("EXPENSES", format_currency(expenses), THEME["danger"]))
##        net_color = THEME["success"] if net >= 0 else THEME["danger"]
##        grid.add_widget(KPICard("NET P&L", format_currency(net), net_color))
##        grid.add_widget(KPICard("HERD", f"{milking}/{total_herd}", THEME["accent"], "Milking/Total"))
##        content.add_widget(grid)
##
##        # Today's Milk
##        today = datetime.now().strftime("%Y-%m-%d")
##        today_milk = sum(safe_float(x.get("quantity")) for x in self.data["milk_production"] if x.get("date", "") == today)
##        content.add_widget(KPICard("TODAY'S MILK", f"{today_milk:.1f} L", THEME["info"], f"{len([x for x in self.data['milk_production'] if x.get('date')==today])} records"))
##
##        # Quick Actions
##        content.add_widget(MDLabel(text="QUICK ACTIONS", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        actions = MDBoxLayout(size_hint_y=None, height=dp(44), spacing=dp(6))
##        actions.add_widget(MDRaisedButton(text="+ Milk", md_bg_color=THEME["info"],
##            on_release=lambda x: setattr(self.manager, "current", "milk_production")))
##        actions.add_widget(MDRaisedButton(text="+ Sale", md_bg_color=THEME["success"],
##            on_release=lambda x: setattr(self.manager, "current", "milk_sales")))
##        actions.add_widget(MDRaisedButton(text="+ Cow", md_bg_color=THEME["accent"],
##            on_release=lambda x: setattr(self.manager, "current", "herd")))
##        actions.add_widget(MDRaisedButton(text="+ Expense", md_bg_color=THEME["danger"],
##            on_release=lambda x: setattr(self.manager, "current", "feed_expenses")))
##        content.add_widget(actions)
##
##        # Recent Activity
##        content.add_widget(MDLabel(text="RECENT ACTIVITY", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        logs = self.data.get("activity_log", [])[:10]
##        if not logs:
##            content.add_widget(MDLabel(text="No recent activity.", theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(50)))
##        for log in logs:
##            card = ThemedCard(orientation="horizontal", size_hint_y=None, height=dp(50), padding=dp(8))
##            card.add_widget(MDLabel(
##                text=f"{log.get('timestamp', '-')[11:16]}  {log.get('action', '')}",
##                theme_text_color="Custom", text_color=THEME["text"],
##                font_style="Caption", size_hint_x=0.7
##            ))
##            card.add_widget(MDLabel(
##                text=log.get('details', '')[:20],
##                theme_text_color="Custom", text_color=THEME["text_secondary"],
##                font_style="Caption", halign="right", size_hint_x=0.3
##            ))
##            content.add_widget(card)
##
##        layout.add_widget(scroll)
##        self.add_widget(layout)
##
##    def _calc_income(self):
##        milk = sum(safe_float(x.get("amount")) for x in self.data.get("milk_sales", []))
##        other = sum(safe_float(x.get("amount")) for x in self.data.get("other_income", []))
##        return milk + other
##
##    def _calc_expenses(self):
##        feed = sum(safe_float(x.get("amount")) for x in self.data.get("feed_expenses", []))
##        vet = sum(safe_float(x.get("cost")) for x in self.data.get("vet_health", []))
##        lab = sum(safe_float(x.get("amount")) for x in self.data.get("labour", []))
##        ops = sum(safe_float(x.get("amount")) for x in self.data.get("operations", []))
##        return feed + vet + lab + ops
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### HERD SCREEN
### ============================================================
##
##class HerdScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "HERD REGISTER")
##
##        # Stats
##        total = len(self.data["herd"])
##        milking = len([c for c in self.data["herd"] if c.get("status", "").lower() == "milking"])
##        dry = len([c for c in self.data["herd"] if c.get("status", "").lower() == "dry"])
##        heifer = len([c for c in self.data["herd"] if c.get("status", "").lower() == "heifer"])
##
##        stats = MDBoxLayout(size_hint_y=None, height=dp(32), spacing=dp(6))
##        for label, count, color in [("Total", total, THEME["primary"]), 
##                                     ("Milking", milking, THEME["success"]),
##                                     ("Dry", dry, THEME["warning"]),
##                                     ("Heifer", heifer, THEME["info"])]:
##            chip = MDChip(text=f"{label}: {count}", md_bg_color=color)
##            chip.text_color = [1, 1, 1, 1]
##            stats.add_widget(chip)
##        content.add_widget(stats)
##
##        # Search
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        # List
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        items = self.filter_items(self.data["herd"], self.search_query, ["tag_no", "name", "breed", "status"])
##
##        if not items:
##            msg = "No cows found." if self.search_query else "No cows registered yet.\nTap + to add your first cow."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(100)))
##            return
##
##        for i, cow in enumerate(items):
##            orig_idx = self.data["herd"].index(cow)
##            status_colors = {
##                "milking": THEME["success"], "dry": THEME["warning"],
##                "heifer": THEME["info"], "sold": THEME["danger"], "dead": THEME["danger"]
##            }
##            status_color = status_colors.get(cow.get("status", "").lower(), THEME["text_secondary"])
##
##            # Calculate age
##            age_text = ""
##            if cow.get("dob"):
##                try:
##                    dob = datetime.strptime(cow["dob"], "%Y-%m-%d")
##                    age_days = (datetime.now() - dob).days
##                    age_text = f"Age: {age_days // 365}y {(age_days % 365) // 30}m"
##                except:
##                    pass
##
##            card = ActionCard(
##                title=f"{cow.get('tag_no', 'N/A')} - {cow.get('name', 'Unknown')}",
##                subtitle=f"Breed: {cow.get('breed', '-')} | Status: {cow.get('status', '-')} | {age_text}",
##                on_edit=partial(self.edit_cow, orig_idx),
##                on_delete=partial(self.delete_cow, orig_idx),
##                on_tap=partial(self.view_cow, orig_idx)
##            )
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        dialog, fields = MobileDialog.create("Add New Cow", [
##            {"key": "tag", "hint": "Tag Number *", "icon": "identifier", "required": True},
##            {"key": "name", "hint": "Cow Name", "icon": "cow"},
##            {"key": "breed", "hint": "Breed", "icon": "dna"},
##            {"key": "status", "hint": "Status (Milking/Dry/Heifer)", "icon": "information", "text": "Milking"},
##            {"key": "dob", "hint": "Date of Birth (YYYY-MM-DD)", "icon": "calendar", "is_date": True},
##            {"key": "cost", "hint": "Purchase Cost", "icon": "cash", "input_filter": "float"},
##            {"key": "notes", "hint": "Notes", "icon": "note-text", "multiline": True},
##        ], on_save=self.save_cow)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_cow(self, values):
##        self.data["herd"].append({
##            "tag_no": values["tag"], "name": values["name"],
##            "breed": values["breed"], "status": values["status"] or "Milking",
##            "dob": values["dob"], "purchase_date": datetime.now().strftime("%Y-%m-%d"),
##            "purchase_cost": values["cost"] or "0", "notes": values["notes"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Added cow", values["tag"])
##        show_snackbar(f"Cow {values['tag']} added!")
##        self.refresh()
##
##    def edit_cow(self, index, *args):
##        cow = self.data["herd"][index]
##        dialog, fields = MobileDialog.create("Edit Cow", [
##            {"key": "tag", "hint": "Tag Number", "icon": "identifier", "text": cow.get("tag_no", "")},
##            {"key": "name", "hint": "Cow Name", "icon": "cow", "text": cow.get("name", "")},
##            {"key": "breed", "hint": "Breed", "icon": "dna", "text": cow.get("breed", "")},
##            {"key": "status", "hint": "Status", "icon": "information", "text": cow.get("status", "Milking")},
##            {"key": "dob", "hint": "Date of Birth", "icon": "calendar", "text": cow.get("dob", ""), "is_date": True},
##            {"key": "cost", "hint": "Purchase Cost", "icon": "cash", "text": str(cow.get("purchase_cost", "")), "input_filter": "float"},
##            {"key": "notes", "hint": "Notes", "icon": "note-text", "text": cow.get("notes", ""), "multiline": True},
##        ], on_save=lambda v: self.update_cow(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_cow(self, index, values):
##        self.data["herd"][index].update({
##            "tag_no": values["tag"], "name": values["name"],
##            "breed": values["breed"], "status": values["status"],
##            "dob": values["dob"], "purchase_cost": values["cost"] or "0",
##            "notes": values["notes"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Updated cow", values["tag"])
##        show_snackbar("Cow updated!")
##        self.refresh()
##
##    def delete_cow(self, index, *args):
##        tag = self.data["herd"][index].get("tag_no", "Unknown")
##        def confirm():
##            self.data["herd"].pop(index)
##            save_data(self.data)
##            log_activity(self.data, "Deleted cow", tag)
##            show_snackbar(f"Cow {tag} removed")
##            self.refresh()
##        ConfirmDialog.show("Delete Cow?", f"Remove {tag} permanently?", confirm)
##
##    def view_cow(self, index, *args):
##        cow = self.data["herd"][index]
##        # Production stats
##        cow_milk = sum(safe_float(x.get("quantity")) for x in self.data["milk_production"] if x.get("cow_tag") == cow.get("tag_no"))
##        cow_vet = sum(safe_float(x.get("cost")) for x in self.data["vet_health"] if x.get("cow_tag") == cow.get("tag_no"))
##
##        content = MDBoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None)
##        content.height = dp(200)
##        info_lines = [
##            f"Tag: {cow.get('tag_no', '-')}",
##            f"Name: {cow.get('name', '-')}",
##            f"Breed: {cow.get('breed', '-')}",
##            f"Status: {cow.get('status', '-')}",
##            f"DOB: {cow.get('dob', '-')}",
##            f"Purchased: {cow.get('purchase_date', '-')}",
##            f"Total Milk: {cow_milk:.1f} L",
##            f"Vet Costs: {format_currency(cow_vet)}",
##        ]
##        for line in info_lines:
##            content.add_widget(MDLabel(text=line, theme_text_color="Custom",
##                text_color=THEME["text"], font_style="Body1", size_hint_y=None, height=dp(22)))
##
##        dialog = MDDialog(title="Cow Profile", type="custom", content_cls=content,
##            buttons=[MDFlatButton(text="CLOSE", on_release=lambda x: dialog.dismiss())])
##        dialog.open()
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### MILK PRODUCTION SCREEN
### ============================================================
##
##class MilkProductionScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "MILK PRODUCTION")
##
##        today = datetime.now().strftime("%Y-%m-%d")
##        today_recs = [x for x in self.data["milk_production"] if x.get("date", "") == today]
##        today_total = sum(safe_float(x.get("quantity")) for x in today_recs)
##
##        content.add_widget(KPICard("TODAY'S TOTAL", f"{today_total:.1f} L", THEME["info"], f"{len(today_recs)} records"))
##
##        # 7-day trend chart
##        if CHARTS_AVAILABLE:
##            dates = []
##            amounts = []
##            for i in range(6, -1, -1):
##                d = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
##                amt = sum(safe_float(x.get("quantity")) for x in self.data["milk_production"] if x.get("date") == d)
##                dates.append(d[5:])  # MM-DD
##                amounts.append(amt)
##            chart = create_bar_chart(dates, amounts, "Last 7 Days (L)", color='#4db87a')
##            if chart:
##                try:
##                    chart.size_hint_y = None
##                    chart.height = dp(200)
##                    content.add_widget(chart)
##                except Exception as e:
##                    print(f"Chart display error: {e}")
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        all_items = sorted(self.data["milk_production"], key=lambda x: x.get("date", ""), reverse=True)
##        items = self.filter_items(all_items, self.search_query, ["date", "cow_tag", "session", "quality"])
##
##        if not items:
##            msg = "No records found." if self.search_query else "No milk production records yet."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(80)))
##            return
##
##        for i, rec in enumerate(items[:50]):  # Limit to 50 for performance
##            orig_idx = self.data["milk_production"].index(rec)
##            card = ActionCard(
##                title=f"{rec.get('date', '-')} | {rec.get('session', 'AM')}",
##                subtitle=f"Cow: {rec.get('cow_tag', '-')} | {rec.get('quantity', '0')} L | {rec.get('quality', 'Good')}",
##                on_edit=partial(self.edit_record, orig_idx),
##                on_delete=partial(self.delete_record, orig_idx)
##            )
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        cow_tags = [c.get("tag_no", "") for c in self.data["herd"] if c.get("tag_no")]
##        cow_hint = f"Cow Tag ({', '.join(cow_tags[:3])}...)" if cow_tags else "Cow Tag"
##        dialog, fields = MobileDialog.create("Record Milk Production", [
##            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##            {"key": "session", "hint": "Session (AM/PM/Evening)", "icon": "clock", "text": "AM", "required": True},
##            {"key": "cow", "hint": cow_hint, "icon": "cow", "required": True},
##            {"key": "qty", "hint": "Quantity (Liters)", "icon": "cup-water", "input_filter": "float", "required": True, "is_number": True},
##            {"key": "quality", "hint": "Quality", "icon": "star", "text": "Good"},
##            {"key": "notes", "hint": "Notes", "icon": "note-text", "multiline": True},
##        ], on_save=self.save_record)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_record(self, values):
##        self.data["milk_production"].append({
##            "date": values["date"], "session": values["session"],
##            "cow_tag": values["cow"], "quantity": values["qty"],
##            "quality": values["quality"], "notes": values["notes"],
##            "recorded_at": datetime.now().strftime("%Y-%m-%d %H:%M")
##        })
##        save_data(self.data)
##        log_activity(self.data, "Milk recorded", f"{values['cow']} - {values['qty']}L")
##        show_snackbar("Milk record saved!")
##        self.refresh()
##
##    def edit_record(self, index, *args):
##        rec = self.data["milk_production"][index]
##        dialog, fields = MobileDialog.create("Edit Milk Record", [
##            {"key": "date", "hint": "Date", "icon": "calendar", "text": rec.get("date", ""), "required": True, "is_date": True},
##            {"key": "session", "hint": "Session", "icon": "clock", "text": rec.get("session", "")},
##            {"key": "cow", "hint": "Cow Tag", "icon": "cow", "text": rec.get("cow_tag", "")},
##            {"key": "qty", "hint": "Quantity", "icon": "cup-water", "text": str(rec.get("quantity", "")), "input_filter": "float", "is_number": True},
##            {"key": "quality", "hint": "Quality", "icon": "star", "text": rec.get("quality", "")},
##            {"key": "notes", "hint": "Notes", "icon": "note-text", "text": rec.get("notes", ""), "multiline": True},
##        ], on_save=lambda v: self.update_record(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_record(self, index, values):
##        self.data["milk_production"][index].update({
##            "date": values["date"], "session": values["session"],
##            "cow_tag": values["cow"], "quantity": values["qty"],
##            "quality": values["quality"], "notes": values["notes"],
##        })
##        save_data(self.data)
##        show_snackbar("Record updated!")
##        self.refresh()
##
##    def delete_record(self, index, *args):
##        rec = self.data["milk_production"][index]
##        def confirm():
##            self.data["milk_production"].pop(index)
##            save_data(self.data)
##            show_snackbar("Record deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete Record?", f"Remove {rec.get('date', '')} - {rec.get('cow_tag', '')}?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### MILK SALES SCREEN
### ============================================================
##
##class MilkSalesScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.showing = "sales"
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "MILK SALES & INCOME")
##
##        total_sales = sum(safe_float(x.get("amount")) for x in self.data["milk_sales"])
##        total_other = sum(safe_float(x.get("amount")) for x in self.data["other_income"])
##        total_liters = sum(safe_float(x.get("liters")) for x in self.data["milk_sales"])
##
##        content.add_widget(KPICard("TOTAL SALES", format_currency(total_sales), THEME["success"], f"{total_liters:.0f} liters"))
##        content.add_widget(KPICard("OTHER INCOME", format_currency(total_other), THEME["info"]))
##
##        # Tabs
##        tab_bar = MDBoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6))
##        self.sales_btn = MDRaisedButton(text="Milk Sales", md_bg_color=THEME["primary"],
##            on_release=lambda x: self.switch_tab("sales"))
##        self.other_btn = MDFlatButton(text="Other Income", on_release=lambda x: self.switch_tab("other"))
##        tab_bar.add_widget(self.sales_btn)
##        tab_bar.add_widget(self.other_btn)
##        content.add_widget(tab_bar)
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def switch_tab(self, tab):
##        self.showing = tab
##        self.search_query = ""
##        if tab == "sales":
##            self.sales_btn.md_bg_color = THEME["primary"]
##            self.other_btn.md_bg_color = [0.9, 0.9, 0.9, 1]
##        else:
##            self.other_btn.md_bg_color = THEME["primary"]
##            self.sales_btn.md_bg_color = [0.9, 0.9, 0.9, 1]
##        self.refresh_list()
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##
##        if self.showing == "sales":
##            all_items = sorted(self.data["milk_sales"], key=lambda x: x.get("date", ""), reverse=True)
##            items = self.filter_items(all_items, self.search_query, ["date", "buyer", "amount"])
##            if not items:
##                self.list_container.add_widget(MDLabel(text="No sales records.", theme_text_color="Hint",
##                    halign="center", size_hint_y=None, height=dp(60)))
##            for item in items[:50]:
##                idx = self.data["milk_sales"].index(item)
##                card = ActionCard(
##                    title=f"{item.get('date', '-')} | {item.get('buyer', 'Unknown')}",
##                    subtitle=f"{item.get('liters', '0')}L @ KES {item.get('price_per_liter', '0')}/L = {format_currency(item.get('amount', '0'))}",
##                    on_edit=partial(self.edit_item, idx, "sales"),
##                    on_delete=partial(self.delete_item, idx, "sales")
##                )
##                self.list_container.add_widget(card)
##        else:
##            all_items = sorted(self.data["other_income"], key=lambda x: x.get("date", ""), reverse=True)
##            items = self.filter_items(all_items, self.search_query, ["date", "source", "amount"])
##            if not items:
##                self.list_container.add_widget(MDLabel(text="No other income records.", theme_text_color="Hint",
##                    halign="center", size_hint_y=None, height=dp(60)))
##            for item in items[:50]:
##                idx = self.data["other_income"].index(item)
##                card = ActionCard(
##                    title=f"{item.get('date', '-')} | {item.get('source', 'Unknown')}",
##                    subtitle=f"{format_currency(item.get('amount', '0'))} | {item.get('notes', '-')}",
##                    on_edit=partial(self.edit_item, idx, "other"),
##                    on_delete=partial(self.delete_item, idx, "other")
##                )
##                self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        if self.showing == "sales":
##            default_price = self.settings.get("milk_price_default", "60")
##            dialog, fields = MobileDialog.create("Record Milk Sale", [
##                {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##                {"key": "buyer", "hint": "Buyer Name", "icon": "account", "required": True},
##                {"key": "liters", "hint": "Liters Sold", "icon": "cup-water", "input_filter": "float", "required": True, "is_number": True},
##                {"key": "price", "hint": "Price per Liter (KES)", "icon": "cash", "input_filter": "float", "text": default_price, "is_number": True},
##            ], on_save=self.save_sale)
##        else:
##            dialog, fields = MobileDialog.create("Record Other Income", [
##                {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##                {"key": "source", "hint": "Income Source", "icon": "tag", "required": True},
##                {"key": "amount", "hint": "Amount (KES)", "icon": "cash", "input_filter": "float", "required": True, "is_number": True},
##                {"key": "notes", "hint": "Notes", "icon": "note-text", "multiline": True},
##            ], on_save=self.save_other)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_sale(self, values):
##        liters = safe_float(values["liters"])
##        price = safe_float(values["price"])
##        self.data["milk_sales"].append({
##            "date": values["date"], "buyer": values["buyer"],
##            "liters": str(liters), "price_per_liter": str(price),
##            "amount": str(liters * price),
##        })
##        save_data(self.data)
##        log_activity(self.data, "Milk sale", f"{values['buyer']} - {liters}L")
##        show_snackbar("Sale recorded!")
##        self.refresh()
##
##    def save_other(self, values):
##        self.data["other_income"].append({
##            "date": values["date"], "source": values["source"],
##            "amount": values["amount"], "notes": values["notes"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Other income", f"{values['source']} - {values['amount']}")
##        show_snackbar("Income recorded!")
##        self.refresh()
##
##    def edit_item(self, index, type, *args):
##        if type == "sales":
##            item = self.data["milk_sales"][index]
##            dialog, fields = MobileDialog.create("Edit Sale", [
##                {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "required": True, "is_date": True},
##                {"key": "buyer", "hint": "Buyer", "icon": "account", "text": item.get("buyer", "")},
##                {"key": "liters", "hint": "Liters", "icon": "cup-water", "text": str(item.get("liters", "")), "input_filter": "float", "is_number": True},
##                {"key": "price", "hint": "Price/Liter", "icon": "cash", "text": str(item.get("price_per_liter", "")), "input_filter": "float", "is_number": True},
##            ], on_save=lambda v: self.update_sale(index, v))
##        else:
##            item = self.data["other_income"][index]
##            dialog, fields = MobileDialog.create("Edit Income", [
##                {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "required": True, "is_date": True},
##                {"key": "source", "hint": "Source", "icon": "tag", "text": item.get("source", "")},
##                {"key": "amount", "hint": "Amount", "icon": "cash", "text": str(item.get("amount", "")), "input_filter": "float", "is_number": True},
##                {"key": "notes", "hint": "Notes", "icon": "note-text", "text": item.get("notes", ""), "multiline": True},
##            ], on_save=lambda v: self.update_other(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_sale(self, index, values):
##        liters = safe_float(values["liters"])
##        price = safe_float(values["price"])
##        self.data["milk_sales"][index].update({
##            "date": values["date"], "buyer": values["buyer"],
##            "liters": str(liters), "price_per_liter": str(price),
##            "amount": str(liters * price),
##        })
##        save_data(self.data)
##        show_snackbar("Sale updated!")
##        self.refresh()
##
##    def update_other(self, index, values):
##        self.data["other_income"][index].update({
##            "date": values["date"], "source": values["source"],
##            "amount": values["amount"], "notes": values["notes"],
##        })
##        save_data(self.data)
##        show_snackbar("Income updated!")
##        self.refresh()
##
##    def delete_item(self, index, type, *args):
##        key = "milk_sales" if type == "sales" else "other_income"
##        item = self.data[key][index]
##        def confirm():
##            self.data[key].pop(index)
##            save_data(self.data)
##            show_snackbar("Deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete?", "Remove this record permanently?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### FEED EXPENSES SCREEN
### ============================================================
##
##class FeedExpensesScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "FEED EXPENSES")
##
##        total = sum(safe_float(x.get("amount")) for x in self.data["feed_expenses"])
##        content.add_widget(KPICard("TOTAL FEED COSTS", format_currency(total), THEME["danger"]))
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        all_items = sorted(self.data["feed_expenses"], key=lambda x: x.get("date", ""), reverse=True)
##        items = self.filter_items(all_items, self.search_query, ["date", "feed_type", "supplier"])
##
##        if not items:
##            msg = "No records found." if self.search_query else "No feed expenses recorded."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(80)))
##            return
##
##        for item in items[:50]:
##            idx = self.data["feed_expenses"].index(item)
##            card = ActionCard(
##                title=f"{item.get('date', '-')} | {item.get('feed_type', 'Unknown')}",
##                subtitle=f"{item.get('supplier', '-')} | {item.get('quantity', '-')} {item.get('unit', 'kg')} | {format_currency(item.get('amount', '0'))}",
##                on_edit=partial(self.edit_item, idx),
##                on_delete=partial(self.delete_item, idx)
##            )
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        dialog, fields = MobileDialog.create("Add Feed Expense", [
##            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##            {"key": "type", "hint": "Feed Type (e.g. Dairy Meal)", "icon": "corn", "required": True},
##            {"key": "supplier", "hint": "Supplier Name", "icon": "truck-delivery"},
##            {"key": "qty", "hint": "Quantity", "icon": "numeric", "input_filter": "float", "is_number": True},
##            {"key": "unit", "hint": "Unit (kg/bale/bag)", "icon": "scale", "text": "kg"},
##            {"key": "amount", "hint": "Total Cost (KES)", "icon": "cash", "input_filter": "float", "required": True, "is_number": True},
##        ], on_save=self.save_item)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_item(self, values):
##        self.data["feed_expenses"].append({
##            "date": values["date"], "feed_type": values["type"],
##            "supplier": values["supplier"], "quantity": values["qty"],
##            "unit": values["unit"], "amount": values["amount"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Feed expense", f"{values['type']} - {values['amount']}")
##        show_snackbar("Feed expense saved!")
##        self.refresh()
##
##    def edit_item(self, index, *args):
##        item = self.data["feed_expenses"][index]
##        dialog, fields = MobileDialog.create("Edit Feed Expense", [
##            {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "is_date": True},
##            {"key": "type", "hint": "Feed Type", "icon": "corn", "text": item.get("feed_type", "")},
##            {"key": "supplier", "hint": "Supplier", "icon": "truck-delivery", "text": item.get("supplier", "")},
##            {"key": "qty", "hint": "Quantity", "icon": "numeric", "text": str(item.get("quantity", "")), "input_filter": "float", "is_number": True},
##            {"key": "unit", "hint": "Unit", "icon": "scale", "text": item.get("unit", "")},
##            {"key": "amount", "hint": "Amount", "icon": "cash", "text": str(item.get("amount", "")), "input_filter": "float", "is_number": True},
##        ], on_save=lambda v: self.update_item(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_item(self, index, values):
##        self.data["feed_expenses"][index].update({
##            "date": values["date"], "feed_type": values["type"],
##            "supplier": values["supplier"], "quantity": values["qty"],
##            "unit": values["unit"], "amount": values["amount"],
##        })
##        save_data(self.data)
##        show_snackbar("Updated!")
##        self.refresh()
##
##    def delete_item(self, index, *args):
##        item = self.data["feed_expenses"][index]
##        def confirm():
##            self.data["feed_expenses"].pop(index)
##            save_data(self.data)
##            show_snackbar("Deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete?", f"Remove {item.get('feed_type', 'this')} expense?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### VET & HEALTH SCREEN
### ============================================================
##
##class VetHealthScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "VET & HEALTH")
##
##        total = sum(safe_float(x.get("cost")) for x in self.data["vet_health"])
##        content.add_widget(KPICard("TOTAL VET COSTS", format_currency(total), THEME["danger"]))
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        all_items = sorted(self.data["vet_health"], key=lambda x: x.get("date", ""), reverse=True)
##        items = self.filter_items(all_items, self.search_query, ["date", "cow_tag", "treatment", "vet_name"])
##
##        if not items:
##            msg = "No records found." if self.search_query else "No veterinary records."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(80)))
##            return
##
##        for item in items[:50]:
##            idx = self.data["vet_health"].index(item)
##            card = ActionCard(
##                title=f"{item.get('date', '-')} | {item.get('cow_tag', 'Unknown')}",
##                subtitle=f"{item.get('treatment', '-')} | {item.get('vet_name', '-')} | {format_currency(item.get('cost', '0'))}",
##                on_edit=partial(self.edit_item, idx),
##                on_delete=partial(self.delete_item, idx)
##            )
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        dialog, fields = MobileDialog.create("Add Vet/Health Record", [
##            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##            {"key": "cow", "hint": "Cow Tag Number", "icon": "cow", "required": True},
##            {"key": "treatment", "hint": "Treatment/Diagnosis", "icon": "needle", "required": True},
##            {"key": "vet", "hint": "Veterinarian Name", "icon": "doctor"},
##            {"key": "cost", "hint": "Cost (KES)", "icon": "cash", "input_filter": "float", "is_number": True},
##            {"key": "notes", "hint": "Notes/Medications", "icon": "note-text", "multiline": True},
##        ], on_save=self.save_item)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_item(self, values):
##        self.data["vet_health"].append({
##            "date": values["date"], "cow_tag": values["cow"],
##            "treatment": values["treatment"], "vet_name": values["vet"],
##            "cost": values["cost"], "notes": values["notes"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Vet record", f"{values['cow']} - {values['treatment']}")
##        show_snackbar("Health record saved!")
##        self.refresh()
##
##    def edit_item(self, index, *args):
##        item = self.data["vet_health"][index]
##        dialog, fields = MobileDialog.create("Edit Health Record", [
##            {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "is_date": True},
##            {"key": "cow", "hint": "Cow Tag", "icon": "cow", "text": item.get("cow_tag", "")},
##            {"key": "treatment", "hint": "Treatment", "icon": "needle", "text": item.get("treatment", "")},
##            {"key": "vet", "hint": "Vet Name", "icon": "doctor", "text": item.get("vet_name", "")},
##            {"key": "cost", "hint": "Cost", "icon": "cash", "text": str(item.get("cost", "")), "input_filter": "float", "is_number": True},
##            {"key": "notes", "hint": "Notes", "icon": "note-text", "text": item.get("notes", ""), "multiline": True},
##        ], on_save=lambda v: self.update_item(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_item(self, index, values):
##        self.data["vet_health"][index].update({
##            "date": values["date"], "cow_tag": values["cow"],
##            "treatment": values["treatment"], "vet_name": values["vet"],
##            "cost": values["cost"], "notes": values["notes"],
##        })
##        save_data(self.data)
##        show_snackbar("Updated!")
##        self.refresh()
##
##    def delete_item(self, index, *args):
##        item = self.data["vet_health"][index]
##        def confirm():
##            self.data["vet_health"].pop(index)
##            save_data(self.data)
##            show_snackbar("Deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete?", f"Remove health record for {item.get('cow_tag', 'cow')}?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### LABOUR SCREEN
### ============================================================
##
##class LabourScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "LABOUR MANAGEMENT")
##
##        total = sum(safe_float(x.get("amount")) for x in self.data["labour"])
##        content.add_widget(KPICard("TOTAL LABOUR COSTS", format_currency(total), THEME["danger"]))
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        all_items = sorted(self.data["labour"], key=lambda x: x.get("date", ""), reverse=True)
##        items = self.filter_items(all_items, self.search_query, ["date", "worker_name", "task"])
##
##        if not items:
##            msg = "No records found." if self.search_query else "No labour records."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(80)))
##            return
##
##        for item in items[:50]:
##            idx = self.data["labour"].index(item)
##            card = ActionCard(
##                title=f"{item.get('date', '-')} | {item.get('worker_name', 'Unknown')}",
##                subtitle=f"{item.get('task', '-')} | {item.get('days', '-')} days @ KES {item.get('daily_rate', '0')}/day = {format_currency(item.get('amount', '0'))}",
##                on_edit=partial(self.edit_item, idx),
##                on_delete=partial(self.delete_item, idx)
##            )
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        default_rate = self.settings.get("daily_rate_default", "500")
##        dialog, fields = MobileDialog.create("Add Labour Record", [
##            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##            {"key": "worker", "hint": "Worker Name", "icon": "account", "required": True},
##            {"key": "task", "hint": "Task/Role", "icon": "hammer-wrench"},
##            {"key": "days", "hint": "Days Worked", "icon": "calendar-clock", "input_filter": "float", "is_number": True},
##            {"key": "rate", "hint": "Daily Rate (KES)", "icon": "cash", "input_filter": "float", "text": default_rate, "is_number": True},
##        ], on_save=self.save_item)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_item(self, values):
##        days = safe_float(values["days"])
##        rate = safe_float(values["rate"])
##        self.data["labour"].append({
##            "date": values["date"], "worker_name": values["worker"],
##            "task": values["task"], "days": str(days),
##            "daily_rate": str(rate), "amount": str(days * rate),
##        })
##        save_data(self.data)
##        log_activity(self.data, "Labour", f"{values['worker']} - {days} days")
##        show_snackbar("Labour record saved!")
##        self.refresh()
##
##    def edit_item(self, index, *args):
##        item = self.data["labour"][index]
##        dialog, fields = MobileDialog.create("Edit Labour Record", [
##            {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "is_date": True},
##            {"key": "worker", "hint": "Worker", "icon": "account", "text": item.get("worker_name", "")},
##            {"key": "task", "hint": "Task", "icon": "hammer-wrench", "text": item.get("task", "")},
##            {"key": "days", "hint": "Days", "icon": "calendar-clock", "text": str(item.get("days", "")), "input_filter": "float", "is_number": True},
##            {"key": "rate", "hint": "Daily Rate", "icon": "cash", "text": str(item.get("daily_rate", "")), "input_filter": "float", "is_number": True},
##        ], on_save=lambda v: self.update_item(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_item(self, index, values):
##        days = safe_float(values["days"])
##        rate = safe_float(values["rate"])
##        self.data["labour"][index].update({
##            "date": values["date"], "worker_name": values["worker"],
##            "task": values["task"], "days": str(days),
##            "daily_rate": str(rate), "amount": str(days * rate),
##        })
##        save_data(self.data)
##        show_snackbar("Updated!")
##        self.refresh()
##
##    def delete_item(self, index, *args):
##        item = self.data["labour"][index]
##        def confirm():
##            self.data["labour"].pop(index)
##            save_data(self.data)
##            show_snackbar("Deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete?", f"Remove record for {item.get('worker_name', 'worker')}?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### OPERATIONS SCREEN
### ============================================================
##
##class OperationsScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "OPERATIONS")
##
##        total = sum(safe_float(x.get("amount")) for x in self.data["operations"])
##        content.add_widget(KPICard("TOTAL OPERATIONS", format_currency(total), THEME["warning"]))
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog)
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        all_items = sorted(self.data["operations"], key=lambda x: x.get("date", ""), reverse=True)
##        items = self.filter_items(all_items, self.search_query, ["date", "category", "description"])
##
##        if not items:
##            msg = "No records found." if self.search_query else "No operation records."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(80)))
##            return
##
##        for item in items[:50]:
##            idx = self.data["operations"].index(item)
##            card = ActionCard(
##                title=f"{item.get('date', '-')} | {item.get('category', 'General')}",
##                subtitle=f"{item.get('description', '-')} | {format_currency(item.get('amount', '0'))}",
##                on_edit=partial(self.edit_item, idx),
##                on_delete=partial(self.delete_item, idx)
##            )
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        dialog, fields = MobileDialog.create("Add Operation Expense", [
##            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
##            {"key": "category", "hint": "Category (Fuel/Repair/Utilities/Equipment)", "icon": "folder", "required": True},
##            {"key": "desc", "hint": "Description", "icon": "text"},
##            {"key": "amount", "hint": "Amount (KES)", "icon": "cash", "input_filter": "float", "required": True, "is_number": True},
##        ], on_save=self.save_item)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_item(self, values):
##        self.data["operations"].append({
##            "date": values["date"], "category": values["category"],
##            "description": values["desc"], "amount": values["amount"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Operation", f"{values['category']} - {values['amount']}")
##        show_snackbar("Operation saved!")
##        self.refresh()
##
##    def edit_item(self, index, *args):
##        item = self.data["operations"][index]
##        dialog, fields = MobileDialog.create("Edit Operation", [
##            {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "is_date": True},
##            {"key": "category", "hint": "Category", "icon": "folder", "text": item.get("category", "")},
##            {"key": "desc", "hint": "Description", "icon": "text", "text": item.get("description", "")},
##            {"key": "amount", "hint": "Amount", "icon": "cash", "text": str(item.get("amount", "")), "input_filter": "float", "is_number": True},
##        ], on_save=lambda v: self.update_item(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_item(self, index, values):
##        self.data["operations"][index].update({
##            "date": values["date"], "category": values["category"],
##            "description": values["desc"], "amount": values["amount"],
##        })
##        save_data(self.data)
##        show_snackbar("Updated!")
##        self.refresh()
##
##    def delete_item(self, index, *args):
##        item = self.data["operations"][index]
##        def confirm():
##            self.data["operations"].pop(index)
##            save_data(self.data)
##            show_snackbar("Deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete?", f"Remove {item.get('category', 'this')} operation?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
##
### ============================================================
### CHATS / NOTES SCREEN
### ============================================================
##
##class ChatsScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "FARM NOTES & CHATS")
##
##        total_notes = len(self.data.get("chats", []))
##        content.add_widget(KPICard("TOTAL NOTES", str(total_notes), THEME["info"]))
##
##        content.add_widget(SearchBar(on_search=self.on_search))
##
##        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(8))
##        self.list_container.bind(minimum_height=self.list_container.setter("height"))
##        content.add_widget(self.list_container)
##        self.refresh_list()
##
##        layout.add_widget(scroll)
##        self.add_fab(layout, self.show_add_dialog, icon="message-plus")
##        self.add_widget(layout)
##
##    def on_search(self, query):
##        self.search_query = query
##        self.refresh_list()
##
##    def refresh_list(self):
##        self.list_container.clear_widgets()
##        all_items = sorted(self.data.get("chats", []), key=lambda x: x.get("timestamp", ""), reverse=True)
##        items = self.filter_items(all_items, self.search_query, ["author", "message", "category"])
##
##        if not items:
##            msg = "No notes found." if self.search_query else "No notes yet.\nTap + to add a farm note or chat."
##            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(100)))
##            return
##
##        for item in items:
##            idx = self.data["chats"].index(item)
##            card = ThemedCard(orientation="vertical", size_hint_y=None, height=dp(120), padding=dp(12))
##
##            header = MDBoxLayout(size_hint_y=None, height=dp(24))
##            header.add_widget(MDLabel(
##                text=item.get("author", "Unknown"),
##                theme_text_color="Custom", text_color=THEME["primary"],
##                font_style="Subtitle1", bold=True, size_hint_x=0.6
##            ))
##            header.add_widget(MDLabel(
##                text=item.get("timestamp", "")[:16],
##                theme_text_color="Custom", text_color=THEME["text_secondary"],
##                font_style="Caption", halign="right", size_hint_x=0.4
##            ))
##            card.add_widget(header)
##
##            if item.get("category"):
##                card.add_widget(MDLabel(
##                    text=f"Category: {item['category']}",
##                    theme_text_color="Custom", text_color=THEME["accent"],
##                    font_style="Caption", size_hint_y=None, height=dp(18)
##                ))
##
##            card.add_widget(MDLabel(
##                text=item.get("message", ""),
##                theme_text_color="Custom", text_color=THEME["text"],
##                font_style="Body1", size_hint_y=None, height=dp(50)
##            ))
##
##            btn_row = MDBoxLayout(size_hint_y=None, height=dp(32), spacing=dp(4))
##            btn_row.add_widget(Widget(size_hint_x=0.6))
##            btn_row.add_widget(MDIconButton(
##                icon="pencil", theme_text_color="Custom", text_color=THEME["info"],
##                on_release=partial(self.edit_note, idx), icon_size=dp(18)
##            ))
##            btn_row.add_widget(MDIconButton(
##                icon="delete", theme_text_color="Custom", text_color=THEME["danger"],
##                on_release=partial(self.delete_note, idx), icon_size=dp(18)
##            ))
##            card.add_widget(btn_row)
##
##            self.list_container.add_widget(card)
##
##    def show_add_dialog(self, *args):
##        dialog, fields = MobileDialog.create("Add Farm Note", [
##            {"key": "author", "hint": "Your Name", "icon": "account", "required": True, "text": self.settings.get("owner_name", "")},
##            {"key": "category", "hint": "Category (General/Task/Reminder)", "icon": "tag", "text": "General"},
##            {"key": "message", "hint": "Message / Note", "icon": "message-text", "required": True, "multiline": True},
##        ], on_save=self.save_note)
##        self.dialog = dialog
##        self.dialog.open()
##
##    def save_note(self, values):
##        self.data.setdefault("chats", []).insert(0, {
##            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
##            "author": values["author"],
##            "category": values["category"],
##            "message": values["message"],
##        })
##        save_data(self.data)
##        log_activity(self.data, "Note added", values["author"])
##        show_snackbar("Note saved!")
##        self.refresh()
##
##    def edit_note(self, index, *args):
##        item = self.data["chats"][index]
##        dialog, fields = MobileDialog.create("Edit Note", [
##            {"key": "author", "hint": "Author", "icon": "account", "text": item.get("author", "")},
##            {"key": "category", "hint": "Category", "icon": "tag", "text": item.get("category", "")},
##            {"key": "message", "hint": "Message", "icon": "message-text", "text": item.get("message", ""), "multiline": True},
##        ], on_save=lambda v: self.update_note(index, v))
##        self.dialog = dialog
##        self.dialog.open()
##
##    def update_note(self, index, values):
##        self.data["chats"][index].update({
##            "author": values["author"],
##            "category": values["category"],
##            "message": values["message"],
##        })
##        save_data(self.data)
##        show_snackbar("Note updated!")
##        self.refresh()
##
##    def delete_note(self, index, *args):
##        def confirm():
##            self.data["chats"].pop(index)
##            save_data(self.data)
##            show_snackbar("Note deleted")
##            self.refresh()
##        ConfirmDialog.show("Delete Note?", "Remove this note permanently?", confirm)
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### GRAPHS / ANALYTICS SCREEN
### ============================================================
##
##class GraphsScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "GRAPHS & ANALYTICS")
##
##        income = self._calc_income()
##        expenses = self._calc_expenses()
##        net = income - expenses
##
##        content.add_widget(KPICard("NET P&L", format_currency(net), THEME["success"] if net >= 0 else THEME["danger"]))
##
##        # Monthly Income vs Expense text bars
##        content.add_widget(MDLabel(text="MONTHLY INCOME vs EXPENSE (KES)", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
##        max_val = 1
##        monthly_data = []
##        for i in range(12):
##            m_income = (sum(safe_float(x.get("amount")) for x in self.data["milk_sales"] if self._get_month(x.get("date", "")) == i) +
##                       sum(safe_float(x.get("amount")) for x in self.data["other_income"] if self._get_month(x.get("date", "")) == i))
##            m_exp = (sum(safe_float(x.get("amount")) for x in self.data["feed_expenses"] if self._get_month(x.get("date", "")) == i) +
##                    sum(safe_float(x.get("cost")) for x in self.data["vet_health"] if self._get_month(x.get("date", "")) == i) +
##                    sum(safe_float(x.get("amount")) for x in self.data["labour"] if self._get_month(x.get("date", "")) == i) +
##                    sum(safe_float(x.get("amount")) for x in self.data["operations"] if self._get_month(x.get("date", "")) == i))
##            monthly_data.append((months[i], m_income, m_exp))
##            max_val = max(max_val, m_income, m_exp)
##
##        for month, inc, exp in monthly_data:
##            row = MDBoxLayout(size_hint_y=None, height=dp(50), spacing=dp(4))
##
##            label_col = MDBoxLayout(orientation="vertical", size_hint_x=0.15)
##            label_col.add_widget(MDLabel(text=month, theme_text_color="Custom",
##                text_color=THEME["text"], font_style="Caption", halign="center"))
##            row.add_widget(label_col)
##
##            bars_col = MDBoxLayout(orientation="vertical", size_hint_x=0.85, spacing=dp(2))
##
##            # Income bar
##            inc_pct = (inc / max_val * 100) if max_val > 0 else 0
##            inc_bar = MDBoxLayout(size_hint_y=None, height=dp(16))
##            inc_bar.add_widget(MDLabel(text=f"Inc: {format_currency(inc)}", theme_text_color="Custom",
##                text_color=THEME["success"], font_style="Caption", size_hint_x=0.4))
##            inc_fill = MDBoxLayout(size_hint_x=0.6)
##            with inc_fill.canvas:
##                Color(*THEME["success"])
##                Rectangle(pos=(0, 0), size=(1, 1))  # Will be sized properly
##            # Use a simpler approach - just show the percentage as text
##            inc_bar.add_widget(MDLabel(text=f"{inc_pct:.0f}%", theme_text_color="Custom",
##                text_color=THEME["success"], font_style="Caption", halign="right", size_hint_x=0.2))
##            bars_col.add_widget(inc_bar)
##
##            # Expense bar
##            exp_pct = (exp / max_val * 100) if max_val > 0 else 0
##            exp_bar = MDBoxLayout(size_hint_y=None, height=dp(16))
##            exp_bar.add_widget(MDLabel(text=f"Exp: {format_currency(exp)}", theme_text_color="Custom",
##                text_color=THEME["danger"], font_style="Caption", size_hint_x=0.4))
##            exp_bar.add_widget(MDLabel(text=f"{exp_pct:.0f}%", theme_text_color="Custom",
##                text_color=THEME["danger"], font_style="Caption", halign="right", size_hint_x=0.2))
##            bars_col.add_widget(exp_bar)
##
##            row.add_widget(bars_col)
##            content.add_widget(row)
##
##        # Expense breakdown
##        content.add_widget(MDLabel(text="EXPENSE BREAKDOWN", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        feed = sum(safe_float(x.get("amount")) for x in self.data["feed_expenses"])
##        vet = sum(safe_float(x.get("cost")) for x in self.data["vet_health"])
##        lab = sum(safe_float(x.get("amount")) for x in self.data["labour"])
##        ops = sum(safe_float(x.get("amount")) for x in self.data["operations"])
##        total_exp = feed + vet + lab + ops
##
##        breakdown = [("Feed", feed, THEME["warning"]), ("Vet & Health", vet, THEME["danger"]),
##                     ("Labour", lab, THEME["info"]), ("Operations", ops, THEME["text_secondary"])]
##        for name, amount, color in breakdown:
##            pct = (amount / total_exp * 100) if total_exp > 0 else 0
##            content.add_widget(MDLabel(
##                text=f"{name}: {format_currency(amount)} ({pct:.1f}%)",
##                theme_text_color="Custom", text_color=color,
##                font_style="Body1", size_hint_y=None, height=dp(26)
##            ))
##
##        # Milk production summary
##        content.add_widget(MDLabel(text="MILK PRODUCTION SUMMARY", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        total_milk = sum(safe_float(x.get("quantity")) for x in self.data["milk_production"])
##        avg_daily = total_milk / 30 if total_milk > 0 else 0  # rough estimate
##        content.add_widget(MDLabel(text=f"Total Recorded: {total_milk:.1f} L", theme_text_color="Custom",
##            text_color=THEME["text"], font_style="Body1", size_hint_y=None, height=dp(24)))
##        content.add_widget(MDLabel(text=f"Records: {len(self.data['milk_production'])}", theme_text_color="Custom",
##            text_color=THEME["text_secondary"], font_style="Body1", size_hint_y=None, height=dp(24)))
##
##        layout.add_widget(scroll)
##        self.add_widget(layout)
##
##    def _calc_income(self):
##        return (sum(safe_float(x.get("amount")) for x in self.data.get("milk_sales", [])) +
##                sum(safe_float(x.get("amount")) for x in self.data.get("other_income", [])))
##
##    def _calc_expenses(self):
##        return (sum(safe_float(x.get("amount")) for x in self.data.get("feed_expenses", [])) +
##                sum(safe_float(x.get("cost")) for x in self.data.get("vet_health", [])) +
##                sum(safe_float(x.get("amount")) for x in self.data.get("labour", [])) +
##                sum(safe_float(x.get("amount")) for x in self.data.get("operations", [])))
##
##    def _get_month(self, date_str):
##        try:
##            return datetime.strptime(str(date_str), "%Y-%m-%d").month - 1
##        except:
##            return -1
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
### ============================================================
### REPORTS SCREEN
### ============================================================
##
##class ReportsScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "REPORTS & ANALYTICS")
##
##        income = self._calc_income()
##        expenses = self._calc_expenses()
##        net = income - expenses
##
##        content.add_widget(KPICard("NET PROFIT/LOSS", format_currency(net), THEME["success"] if net >= 0 else THEME["danger"]))
##
##        # Expense breakdown pie chart
##        feed = sum(safe_float(x.get("amount")) for x in self.data["feed_expenses"])
##        vet = sum(safe_float(x.get("cost")) for x in self.data["vet_health"])
##        lab = sum(safe_float(x.get("amount")) for x in self.data["labour"])
##        ops = sum(safe_float(x.get("amount")) for x in self.data["operations"])
##
##        if CHARTS_AVAILABLE and expenses > 0:
##            labels = ["Feed", "Vet", "Labour", "Ops"]
##            values = [feed, vet, lab, ops]
##            colors = ['#f39c12', '#e74c3c', '#3498db', '#7f8c8d']
##            chart = create_pie_chart(labels, values, "Expense Breakdown", colors)
##            if chart:
##                try:
##                    chart.size_hint_y = None
##                    chart.height = dp(240)
##                    content.add_widget(chart)
##                except Exception as e:
##                    print(f"Chart display error: {e}")
##
##        content.add_widget(MDLabel(text="EXPENSE BREAKDOWN", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        breakdown = [("Feed", feed, THEME["warning"]), ("Vet & Health", vet, THEME["danger"]),
##                     ("Labour", lab, THEME["info"]), ("Operations", ops, THEME["text_secondary"])]
##        for name, amount, color in breakdown:
##            pct = (amount / expenses * 100) if expenses > 0 else 0
##            content.add_widget(MDLabel(
##                text=f"{name}: {format_currency(amount)} ({pct:.1f}%)",
##                theme_text_color="Custom", text_color=color,
##                font_style="Body1", size_hint_y=None, height=dp(26)
##            ))
##
##        # Monthly income vs expense chart
##        if CHARTS_AVAILABLE:
##            months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
##            inc_vals = []
##            exp_vals = []
##            for i in range(12):
##                m_income = sum(safe_float(x.get("amount")) for x in self.data["milk_sales"] if self._get_month(x.get("date", "")) == i)
##                m_income += sum(safe_float(x.get("amount")) for x in self.data["other_income"] if self._get_month(x.get("date", "")) == i)
##                m_exp = (sum(safe_float(x.get("amount")) for x in self.data["feed_expenses"] if self._get_month(x.get("date", "")) == i) +
##                         sum(safe_float(x.get("cost")) for x in self.data["vet_health"] if self._get_month(x.get("date", "")) == i) +
##                         sum(safe_float(x.get("amount")) for x in self.data["labour"] if self._get_month(x.get("date", "")) == i) +
##                         sum(safe_float(x.get("amount")) for x in self.data["operations"] if self._get_month(x.get("date", "")) == i))
##                inc_vals.append(m_income)
##                exp_vals.append(m_exp)
##
##            try:
##                fig, ax = plt.subplots(figsize=(3.2, 2.4), dpi=100)
##                fig.patch.set_alpha(0)
##                ax.set_facecolor('none')
##                x = range(12)
##                ax.bar([i - 0.2 for i in x], inc_vals, 0.4, label='Income', color='#27ae60', alpha=0.8)
##                ax.bar([i + 0.2 for i in x], exp_vals, 0.4, label='Expense', color='#e74c3c', alpha=0.8)
##                ax.set_xticks(x)
##                ax.set_xticklabels(months, fontsize=6, color='white')
##                ax.tick_params(colors='white', labelsize=7)
##                ax.set_title("Monthly Income vs Expense", color='white', fontsize=10)
##                ax.legend(facecolor='none', edgecolor='white', labelcolor='white', fontsize=7)
##                ax.spines['bottom'].set_color('white')
##                ax.spines['left'].set_color('white')
##                ax.spines['top'].set_visible(False)
##                ax.spines['right'].set_visible(False)
##                ax.grid(True, alpha=0.3, color='white', axis='y')
##                plt.tight_layout()
##                chart = FigureCanvasKivyAgg(fig)
##                chart.size_hint_y = None
##                chart.height = dp(200)
##                content.add_widget(chart)
##            except Exception as e:
##                print(f"Chart display error: {e}")
##
##        # Herd Summary
##        content.add_widget(MDLabel(text="HERD SUMMARY", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        breeds = defaultdict(int)
##        for cow in self.data["herd"]:
##            breeds[cow.get("breed", "Unknown")] += 1
##        if not breeds:
##            content.add_widget(MDLabel(text="No herd data.", theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(40)))
##        else:
##            for breed, count in breeds.items():
##                content.add_widget(MDLabel(
##                    text=f"{breed}: {count} cow(s)",
##                    theme_text_color="Custom", text_color=THEME["text"],
##                    font_style="Body1", size_hint_y=None, height=dp(24)
##                ))
##
##        layout.add_widget(scroll)
##        self.add_widget(layout)
##
##    def _calc_income(self):
##        return (sum(safe_float(x.get("amount")) for x in self.data.get("milk_sales", [])) +
##                sum(safe_float(x.get("amount")) for x in self.data.get("other_income", [])))
##
##    def _calc_expenses(self):
##        return (sum(safe_float(x.get("amount")) for x in self.data.get("feed_expenses", [])) +
##                sum(safe_float(x.get("cost")) for x in self.data.get("vet_health", [])) +
##                sum(safe_float(x.get("amount")) for x in self.data.get("labour", [])) +
##                sum(safe_float(x.get("amount")) for x in self.data.get("operations", [])))
##
##    def _get_month(self, date_str):
##        try:
##            return datetime.strptime(str(date_str), "%Y-%m-%d").month - 1
##        except:
##            return -1
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### SETTINGS SCREEN
### ============================================================
##
##class SettingsScreen(BaseScreen):
##    def __init__(self, **kwargs):
##        super().__init__(**kwargs)
##        self.build_ui()
##
##    def build_ui(self):
##        layout = MDBoxLayout(orientation="vertical")
##        scroll, content = self.create_scroll_layout()
##        self.add_screen_title(content, "SETTINGS & TOOLS")
##
##        # Farm Profile
##        content.add_widget(MDLabel(text="FARM PROFILE", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        profile_card = ThemedCard(orientation="vertical", size_hint_y=None, height=dp(200), spacing=dp(8))
##        self.farm_name_field = MDTextField(hint_text="Farm Name", text=self.settings.get("farm_name", ""), mode="rectangle")
##        self.farm_loc_field = MDTextField(hint_text="Location", text=self.settings.get("farm_location", ""), mode="rectangle")
##        self.owner_field = MDTextField(hint_text="Owner Name", text=self.settings.get("owner_name", ""), mode="rectangle")
##        profile_card.add_widget(self.farm_name_field)
##        profile_card.add_widget(self.farm_loc_field)
##        profile_card.add_widget(self.owner_field)
##        profile_card.add_widget(MDRaisedButton(text="SAVE PROFILE", md_bg_color=THEME["success"],
##            on_release=self.save_profile, size_hint=(1, None), height=dp(40)))
##        content.add_widget(profile_card)
##
##        # Defaults
##        content.add_widget(MDLabel(text="DEFAULTS", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        defaults_card = ThemedCard(orientation="vertical", size_hint_y=None, height=dp(160), spacing=dp(8))
##        self.price_field = MDTextField(hint_text="Default Milk Price/Liter (KES)",
##            text=self.settings.get("milk_price_default", "60"), mode="rectangle", input_filter="float")
##        self.rate_field = MDTextField(hint_text="Default Daily Labour Rate (KES)",
##            text=self.settings.get("daily_rate_default", "500"), mode="rectangle", input_filter="float")
##        defaults_card.add_widget(self.price_field)
##        defaults_card.add_widget(self.rate_field)
##        defaults_card.add_widget(MDRaisedButton(text="SAVE DEFAULTS", md_bg_color=THEME["success"],
##            on_release=self.save_defaults, size_hint=(1, None), height=dp(40)))
##        content.add_widget(defaults_card)
##
##        # Data Management
##        content.add_widget(MDLabel(text="DATA MANAGEMENT", theme_text_color="Custom",
##            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))
##
##        data_card = ThemedCard(orientation="vertical", size_hint_y=None, height=dp(220), spacing=dp(8))
##        data_card.add_widget(MDRaisedButton(text="EXPORT TO EXCEL", md_bg_color=THEME["info"],
##            on_release=self.export_excel, size_hint=(1, None), height=dp(44)))
##        data_card.add_widget(MDRaisedButton(text="VIEW BACKUPS", md_bg_color=THEME["accent"],
##            on_release=self.view_backups, size_hint=(1, None), height=dp(44)))
##        data_card.add_widget(MDRaisedButton(text="VIEW BACKUPS", md_bg_color=THEME["accent"],
##            on_release=self.view_backups, size_hint=(1, None), height=dp(44)))
##        data_card.add_widget(MDRaisedButton(text="CLEAR ALL DATA", md_bg_color=THEME["danger"],
##            on_release=self.clear_all_data, size_hint=(1, None), height=dp(44)))
##        content.add_widget(data_card)
##
##        # App Info
##        content.add_widget(MDLabel(text="DAIRY FARM MS PRO v2.0", theme_text_color="Custom",
##            text_color=THEME["text_secondary"], font_style="Caption",
##            halign="center", size_hint_y=None, height=dp(30)))
##
##        layout.add_widget(scroll)
##        self.add_widget(layout)
##
##    def save_profile(self, *args):
##        self.settings["farm_name"] = self.farm_name_field.text
##        self.settings["farm_location"] = self.farm_loc_field.text
##        self.settings["owner_name"] = self.owner_field.text
##        save_settings(self.settings)
##        show_snackbar("Profile saved!")
##
##    def save_defaults(self, *args):
##        self.settings["milk_price_default"] = self.price_field.text or "60"
##        self.settings["daily_rate_default"] = self.rate_field.text or "500"
##        save_settings(self.settings)
##        show_snackbar("Defaults saved!")
##
##    def export_excel(self, *args):
##        if not EXCEL_AVAILABLE:
##            show_snackbar("Install openpyxl: pip install openpyxl", THEME["danger"], 3)
##            return
##        app = MDApp.get_running_app()
##        app.do_export("excel")
##
##    def view_backups(self, *args):
##        backups = sorted([f for f in os.listdir(get_backup_dir()) if f.startswith("backup_")], reverse=True)
##        content = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(4))
##        content.height = dp(20) + len(backups) * dp(36)
##
##        if not backups:
##            content.add_widget(MDLabel(text="No backups found.", theme_text_color="Hint",
##                halign="center", size_hint_y=None, height=dp(40)))
##        else:
##            for b in backups[:20]:
##                content.add_widget(MDLabel(text=b, theme_text_color="Custom",
##                    text_color=THEME["text"], font_style="Caption", size_hint_y=None, height=dp(32)))
##
##        dialog = MDDialog(title="Backups", type="custom", content_cls=content,
##            buttons=[MDFlatButton(text="CLOSE", on_release=lambda x: dialog.dismiss())])
##        dialog.open()
##
##    def clear_all_data(self, *args):
##        def confirm():
##            self.data = {
##                "farm": {"name": "My Dairy Farm", "location": "", "owner": ""},
##                "herd": [], "milk_production": [], "milk_sales": [],
##                "other_income": [], "feed_expenses": [], "vet_health": [],
##                "labour": [], "operations": [], "assets": [], "breeding": [], "activity_log": [],
##            }
##            save_data(self.data)
##            show_snackbar("All data cleared!")
##            self.refresh()
##        ConfirmDialog.show("CLEAR ALL DATA?", "This will delete EVERYTHING. This cannot be undone!", confirm, "CLEAR", THEME["danger"])
##
##    def refresh(self):
##        self.clear_widgets()
##        self.reload_data()
##        self.build_ui()
##
##
### ============================================================
### MAIN APP
### ============================================================
##
##class DairyFarmApp(MDApp):
##    def build(self):
##        self.theme_cls.primary_palette = "Green"
##        self.theme_cls.theme_style = "Light"
##
##        # Load settings for theme
##        self.settings = load_settings()
##        if self.settings.get("dark_mode"):
##            self.theme_cls.theme_style = "Dark"
##            global THEME
##            THEME = DARK_THEME
##
##        sm = MDScreenManager(transition=SlideTransition(duration=0.2))
##        sm.add_widget(DashboardScreen(name="dashboard"))
##        sm.add_widget(HerdScreen(name="herd"))
##        sm.add_widget(MilkProductionScreen(name="milk_production"))
##        sm.add_widget(MilkSalesScreen(name="milk_sales"))
##        sm.add_widget(FeedExpensesScreen(name="feed_expenses"))
##        sm.add_widget(VetHealthScreen(name="vet_health"))
##        sm.add_widget(LabourScreen(name="labour"))
##        sm.add_widget(OperationsScreen(name="operations"))
##        sm.add_widget(ReportsScreen(name="reports"))
##        sm.add_widget(GraphsScreen(name="graphs"))
##        sm.add_widget(ChatsScreen(name="chats"))
##        sm.add_widget(SettingsScreen(name="settings"))
##
##        self.sm = sm
##
##        # Main layout with bottom navigation
##        layout = MDBoxLayout(orientation="vertical")
##
##        # Toolbar
##        self.toolbar = MDTopAppBar(
##            title="Dairy Farm MS",
##            elevation=4,
##            pos_hint={"top": 1},
##            md_bg_color=THEME["primary"],
##            left_action_items=[["menu", lambda x: self.nav_drawer.set_state("open")]],
##            right_action_items=[
##                ["chart-bar", lambda x: self.switch_screen("graphs")],
##                ["export-variant", lambda x: self.show_export_menu()],
##            ],
##        )
##        layout.add_widget(self.toolbar)
##        layout.add_widget(sm)
##
##        # Custom Bottom Navigation Bar (reliable across KivyMD versions)
##        self.bottom_bar = MDBoxLayout(
##            orientation="horizontal",
##            size_hint_y=None,
##            height=dp(56),
##            md_bg_color=THEME["card"],
##            padding=[dp(4), dp(4), dp(4), dp(4)],
##            spacing=dp(2),
##        )
##
##        # Add shadow line above bottom bar
##        with self.bottom_bar.canvas.before:
##            Color(*THEME["divider"])
##            Rectangle(pos=(0, dp(56)), size=(Window.width, dp(1)))
##
##        self.nav_buttons = {}
##        nav_items = [
##            ("dashboard", "Dashboard", "chart-bar"),
##            ("herd", "Herd", "cow"),
##            ("milk_production", "Milk", "cup-water"),
##            ("milk_sales", "Sales", "cash-multiple"),
##            ("reports", "Reports", "file-chart"),
##        ]
##
##        for screen_name, label, icon in nav_items:
##            btn = MDBoxLayout(orientation="vertical", size_hint_x=1)
##            btn_icon = MDIconButton(
##                icon=icon,
##                theme_text_color="Custom",
##                text_color=THEME["text_secondary"],
##                pos_hint={"center_x": 0.5},
##                icon_size=dp(22),
##            )
##            btn_label = MDLabel(
##                text=label,
##                theme_text_color="Custom",
##                text_color=THEME["text_secondary"],
##                font_style="Caption",
##                halign="center",
##                size_hint_y=None,
##                height=dp(16),
##            )
##            btn.add_widget(btn_icon)
##            btn.add_widget(btn_label)
##
##            # Make the whole box layout clickable
##            btn.bind(on_touch_down=partial(self.on_bottom_button_touch, screen_name, btn_icon, btn_label))
##
##            self.nav_buttons[screen_name] = (btn, btn_icon, btn_label)
##            self.bottom_bar.add_widget(btn)
##
##        layout.add_widget(self.bottom_bar)
##        self.bottom_nav = self.bottom_bar  # alias for compatibility
##
##        # Side Navigation Drawer (for less-used screens)
##        self.nav_drawer = MDNavigationDrawer(
##            md_bg_color=THEME["bg"],
##            elevation=4,
##        )
##
##        nav_content = MDBoxLayout(orientation="vertical", padding=dp(8), spacing=dp(4))
##        nav_content.add_widget(MDLabel(
##            text="MENU",
##            theme_text_color="Custom",
##            text_color=THEME["primary"],
##            font_style="H6",
##            size_hint_y=None,
##            height=dp(40),
##            halign="center"
##        ))
##
##        drawer_items = [
##            ("feed_expenses", "Feed Expenses", "corn"),
##            ("vet_health", "Vet & Health", "needle"),
##            ("labour", "Labour", "account-hard-hat"),
##            ("operations", "Operations", "cogs"),
##            ("chats", "Farm Notes", "message-text"),
##            ("settings", "Settings & Tools", "cog"),
##        ]
##
##        nav_list = MDList()
##        for screen_name, label, icon in drawer_items:
##            item = OneLineIconListItem(
##                IconLeftWidget(icon=icon),
##                text=label,
##                on_release=lambda x, s=screen_name: self.switch_screen(s)
##            )
##            nav_list.add_widget(item)
##
##        nav_content.add_widget(nav_list)
##        self.nav_drawer.add_widget(nav_content)
##
##        root = MDScreen()
##        root.add_widget(layout)
##        root.add_widget(self.nav_drawer)
##
##        return root
##
##    def show_export_menu(self, *args):
##        menu_items = [
##            {"text": "Export to Excel", "icon": "microsoft-excel", "on_release": lambda: self.do_export("excel")},
##        ]
##        self.export_menu = MDDropdownMenu(
##            caller=self.toolbar,
##            items=menu_items,
##            width_mult=3,
##        )
##        self.export_menu.open()
##
##    def do_export(self, fmt):
##        self.export_menu.dismiss()
##        if not EXCEL_AVAILABLE:
##            show_snackbar("Install openpyxl: pip install openpyxl", THEME["danger"], 3)
##            return
##
##        data = load_data()
##        export_dir = os.path.join(get_data_dir(), "exports")
##        os.makedirs(export_dir, exist_ok=True)
##        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
##        filename = os.path.join(export_dir, f"DairyFarm_Report_{timestamp}.xlsx")
##
##        try:
##            wb = Workbook()
##
##            # Helper styles
##            header_font = Font(bold=True, color="FFFFFF")
##            header_fill = PatternFill(start_color="1a5c3a", end_color="1a5c3a", fill_type="solid")
##            header_align = Alignment(horizontal="center", vertical="center")
##            thin_border = Border(
##                left=Side(style='thin'), right=Side(style='thin'),
##                top=Side(style='thin'), bottom=Side(style='thin')
##            )
##
##            def add_sheet(wb, title, headers, rows):
##                ws = wb.create_sheet(title=title)
##                ws.append(headers)
##                for cell in ws[1]:
##                    cell.font = header_font
##                    cell.fill = header_fill
##                    cell.alignment = header_align
##                    cell.border = thin_border
##                for row in rows:
##                    ws.append(row)
##                    for cell in ws[ws.max_row]:
##                        cell.border = thin_border
##                # Auto-adjust column widths
##                for col in ws.columns:
##                    max_length = 0
##                    col_letter = col[0].column_letter
##                    for cell in col:
##                        try:
##                            if cell.value:
##                                max_length = max(max_length, len(str(cell.value)))
##                        except:
##                            pass
##                    ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
##                return ws
##
##            # 1. SUMMARY Sheet
##            total_income = (sum(safe_float(x.get("amount")) for x in data.get("milk_sales", [])) +
##                           sum(safe_float(x.get("amount")) for x in data.get("other_income", [])))
##            total_feed = sum(safe_float(x.get("amount")) for x in data.get("feed_expenses", []))
##            total_vet = sum(safe_float(x.get("cost")) for x in data.get("vet_health", []))
##            total_labour = sum(safe_float(x.get("amount")) for x in data.get("labour", []))
##            total_ops = sum(safe_float(x.get("amount")) for x in data.get("operations", []))
##            total_exp = total_feed + total_vet + total_labour + total_ops
##
##            ws_summary = wb.create_sheet(title="Summary", index=0)
##            ws_summary.append(["DAIRY FARM MANAGEMENT REPORT"])
##            ws_summary.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
##            ws_summary.append(["Farm:", data.get("farm", {}).get("name", "My Dairy Farm")])
##            ws_summary.append([])
##            ws_summary.append(["FINANCIAL SUMMARY"])
##            ws_summary.append(["Total Income", format_currency(total_income)])
##            ws_summary.append(["Total Expenses", format_currency(total_exp)])
##            ws_summary.append(["Net Profit/Loss", format_currency(total_income - total_exp)])
##            ws_summary.append([])
##            ws_summary.append(["EXPENSE BREAKDOWN"])
##            ws_summary.append(["Feed", format_currency(total_feed)])
##            ws_summary.append(["Vet & Health", format_currency(total_vet)])
##            ws_summary.append(["Labour", format_currency(total_labour)])
##            ws_summary.append(["Operations", format_currency(total_ops)])
##            ws_summary.append([])
##            ws_summary.append(["HERD SUMMARY"])
##            ws_summary.append(["Total Cows", len(data.get("herd", []))])
##            ws_summary.append(["Milking", len([c for c in data.get("herd", []) if c.get("status", "").lower() == "milking"])])
##            ws_summary.append(["Dry", len([c for c in data.get("herd", []) if c.get("status", "").lower() == "dry"])])
##            ws_summary.append(["Heifer", len([c for c in data.get("herd", []) if c.get("status", "").lower() == "heifer"])])
##            ws_summary.append([])
##            ws_summary.append(["MILK SUMMARY"])
##            total_milk = sum(safe_float(x.get("quantity")) for x in data.get("milk_production", []))
##            total_liters_sold = sum(safe_float(x.get("liters")) for x in data.get("milk_sales", []))
##            ws_summary.append(["Total Production", f"{total_milk:.1f} L"])
##            ws_summary.append(["Total Sold", f"{total_liters_sold:.1f} L"])
##
##            # 2. HERD Sheet
##            herd_rows = []
##            for cow in data.get("herd", []):
##                herd_rows.append([
##                    cow.get("tag_no", ""), cow.get("name", ""), cow.get("breed", ""),
##                    cow.get("status", ""), cow.get("dob", ""), cow.get("purchase_date", ""),
##                    cow.get("purchase_cost", ""), cow.get("notes", "")
##                ])
##            add_sheet(wb, "Herd", ["Tag No", "Name", "Breed", "Status", "DOB", "Purchase Date", "Cost", "Notes"], herd_rows)
##
##            # 3. Milk Production Sheet
##            prod_rows = []
##            for rec in data.get("milk_production", []):
##                prod_rows.append([
##                    rec.get("date", ""), rec.get("session", ""), rec.get("cow_tag", ""),
##                    rec.get("quantity", ""), rec.get("quality", ""), rec.get("notes", ""), rec.get("recorded_at", "")
##                ])
##            add_sheet(wb, "Milk Production", ["Date", "Session", "Cow Tag", "Quantity (L)", "Quality", "Notes", "Recorded At"], prod_rows)
##
##            # 4. Milk Sales Sheet
##            sales_rows = []
##            for rec in data.get("milk_sales", []):
##                sales_rows.append([
##                    rec.get("date", ""), rec.get("buyer", ""), rec.get("liters", ""),
##                    rec.get("price_per_liter", ""), rec.get("amount", "")
##                ])
##            add_sheet(wb, "Milk Sales", ["Date", "Buyer", "Liters", "Price/Liter", "Total Amount"], sales_rows)
##
##            # 5. Other Income Sheet
##            income_rows = []
##            for rec in data.get("other_income", []):
##                income_rows.append([
##                    rec.get("date", ""), rec.get("source", ""), rec.get("amount", ""), rec.get("notes", "")
##                ])
##            add_sheet(wb, "Other Income", ["Date", "Source", "Amount", "Notes"], income_rows)
##
##            # 6. Feed Expenses Sheet
##            feed_rows = []
##            for rec in data.get("feed_expenses", []):
##                feed_rows.append([
##                    rec.get("date", ""), rec.get("feed_type", ""), rec.get("supplier", ""),
##                    rec.get("quantity", ""), rec.get("unit", ""), rec.get("amount", "")
##                ])
##            add_sheet(wb, "Feed Expenses", ["Date", "Feed Type", "Supplier", "Quantity", "Unit", "Amount"], feed_rows)
##
##            # 7. Vet & Health Sheet
##            vet_rows = []
##            for rec in data.get("vet_health", []):
##                vet_rows.append([
##                    rec.get("date", ""), rec.get("cow_tag", ""), rec.get("treatment", ""),
##                    rec.get("vet_name", ""), rec.get("cost", ""), rec.get("notes", "")
##                ])
##            add_sheet(wb, "Vet & Health", ["Date", "Cow Tag", "Treatment", "Vet Name", "Cost", "Notes"], vet_rows)
##
##            # 8. Labour Sheet
##            labour_rows = []
##            for rec in data.get("labour", []):
##                labour_rows.append([
##                    rec.get("date", ""), rec.get("worker_name", ""), rec.get("task", ""),
##                    rec.get("days", ""), rec.get("daily_rate", ""), rec.get("amount", "")
##                ])
##            add_sheet(wb, "Labour", ["Date", "Worker", "Task", "Days", "Daily Rate", "Total Pay"], labour_rows)
##
##            # 9. Operations Sheet
##            ops_rows = []
##            for rec in data.get("operations", []):
##                ops_rows.append([
##                    rec.get("date", ""), rec.get("category", ""), rec.get("description", ""), rec.get("amount", "")
##                ])
##            add_sheet(wb, "Operations", ["Date", "Category", "Description", "Amount"], ops_rows)
##
##            # 10. Farm Notes Sheet
##            chat_rows = []
##            for rec in data.get("chats", []):
##                chat_rows.append([
##                    rec.get("timestamp", ""), rec.get("author", ""), rec.get("category", ""), rec.get("message", "")
##                ])
##            add_sheet(wb, "Farm Notes", ["Timestamp", "Author", "Category", "Message"], chat_rows)
##
##            wb.save(filename)
##            show_snackbar(f"Excel exported: {filename}", duration=3)
##        except Exception as e:
##            show_snackbar(f"Export failed: {str(e)}", THEME["danger"], 3)
##
##    def on_bottom_button_touch(self, screen_name, btn_icon, btn_label, widget, touch):
##        if widget.collide_point(*touch.pos):
##            if touch.button == "left" and not touch.is_double_tap:
##                self.switch_screen(screen_name)
##                return True
##        return False
##
##    def update_bottom_nav_highlight(self, active_name):
##        """Highlight the active bottom nav button."""
##        for name, (btn, icon, label) in self.nav_buttons.items():
##            if name == active_name:
##                icon.text_color = THEME["primary"]
##                label.text_color = THEME["primary"]
##            else:
##                icon.text_color = THEME["text_secondary"]
##                label.text_color = THEME["text_secondary"]
##
##    def switch_screen(self, name):
##        # Refresh the target screen
##        for screen in self.sm.screens:
##            if screen.name == name and hasattr(screen, "refresh"):
##                screen.refresh()
##
##        self.sm.current = name
##        self.toolbar.title = name.replace("_", " ").title()
##        self.nav_drawer.set_state("close")
##
##        # Update bottom nav highlight
##        self.update_bottom_nav_highlight(name)
##
##
##if __name__ == "__main__":
##    DairyFarmApp().run()

##############################################################################################################

"""
Dairy Farm MS Pro - Fully Functional Mobile Version
====================================================
Production-ready dairy farm management for Android/iOS.

Features:
- Bottom Navigation (mobile-native)
- Data Validation & Error Handling
- Search & Filter across all modules
- Charts & Analytics (Matplotlib)
- CSV/JSON Export & Share
- Farm Settings & Profile
- Dark Mode Support
- Auto-backup
- Cow Detail Profiles
- Production Trends
- Expense Breakdown Charts
- Income vs Expense Analytics
- Recent Activity Feed
- Responsive Touch Targets

Dependencies:
    pip install kivymd matplotlib

Build Android:
    buildozer android debug

Run Desktop:
    python dairy_farm_mobile_pro.py
"""

import json
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
import os
import csv
import io
import base64
from datetime import datetime, timedelta
from functools import partial
from collections import defaultdict

from kivy.config import Config
nConfig = Config  # alias to avoid issues
nConfig.set('graphics', 'width', '360')
nConfig.set('graphics', 'height', '640')

from kivy.core.window import Window
from kivy.metrics import dp
from kivy.properties import ObjectProperty, StringProperty, BooleanProperty
from kivy.clock import Clock
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.anchorlayout import AnchorLayout
from kivy.animation import Animation
from kivy.graphics import Color, Rectangle, RoundedRectangle

from kivymd.app import MDApp
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.button import (
    MDRaisedButton, MDIconButton, MDFlatButton, 
    MDFloatingActionButton, MDRoundFlatButton
)
from kivymd.uix.card import MDCard
from kivymd.uix.chip import MDChip
from kivymd.uix.dialog import MDDialog
from kivymd.uix.label import MDLabel
from kivymd.uix.list import (
    MDList, OneLineIconListItem, TwoLineAvatarIconListItem,
    IconLeftWidget, IconRightWidget
)
from kivymd.uix.menu import MDDropdownMenu
from kivymd.uix.pickers import MDDatePicker
from kivymd.uix.screen import MDScreen
from kivymd.uix.screenmanager import MDScreenManager
from kivy.uix.screenmanager import SlideTransition
from kivymd.uix.textfield import MDTextField
from kivymd.uix.toolbar import MDTopAppBar
# from kivymd.uix.bottomnavigation import MDBottomNavigation, MDBottomNavigationItem  # Using custom bottom bar
from kivymd.uix.navigationdrawer import MDNavigationDrawer
from kivymd.uix.snackbar import MDSnackbar
from kivymd.uix.selectioncontrol import MDCheckbox
from kivymd.uix.progressbar import MDProgressBar
from kivymd.uix.gridlayout import MDGridLayout

# Chart imports
# Charts disabled: kivy-garden matplotlib is incompatible with Kivy 2.3.1
# All chart functions return None gracefully
CHARTS_AVAILABLE = False

# Keep stub imports for the chart helper functions to reference
# but they will never create actual widgets
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.dates import DateFormatter
except Exception:
    pass

# ============================================================
# THEME & CONFIG
# ============================================================

THEME = {
    "primary": [0.102, 0.361, 0.227, 1],      # #1a5c3a
    "accent": [0.176, 0.541, 0.369, 1],       # #2d8a5e
    "light": [0.302, 0.722, 0.478, 1],        # #4db87a
    "bg": [0.941, 0.957, 0.941, 1],           # #f0f4f0
    "card": [1, 1, 1, 1],
    "danger": [0.906, 0.298, 0.235, 1],       # #e74c3c
    "warning": [0.953, 0.612, 0.071, 1],      # #f39c12
    "success": [0.153, 0.682, 0.376, 1],      # #27ae60
    "info": [0.204, 0.596, 0.859, 1],         # #3498db
    "text": [0.173, 0.243, 0.314, 1],         # #2c3e50
    "text_secondary": [0.498, 0.549, 0.553, 1], # #7f8c8d
    "divider": [0.85, 0.85, 0.85, 1],
}

DARK_THEME = {
    "primary": [0.302, 0.722, 0.478, 1],
    "accent": [0.176, 0.541, 0.369, 1],
    "light": [0.102, 0.361, 0.227, 1],
    "bg": [0.12, 0.12, 0.12, 1],
    "card": [0.18, 0.18, 0.18, 1],
    "danger": [0.906, 0.298, 0.235, 1],
    "warning": [0.953, 0.612, 0.071, 1],
    "success": [0.153, 0.682, 0.376, 1],
    "info": [0.204, 0.596, 0.859, 1],
    "text": [0.9, 0.9, 0.9, 1],
    "text_secondary": [0.6, 0.6, 0.6, 1],
    "divider": [0.3, 0.3, 0.3, 1],
}

CURRENCY = "KES"

# ============================================================
# DATA LAYER (Robust & Validated)
# ============================================================

def get_data_dir():
    try:
        app = MDApp.get_running_app()
        if app:
            return app.user_data_dir
    except:
        pass
    base = os.path.expanduser("~/.local/share/DairyFarmMSPro")
    os.makedirs(base, exist_ok=True)
    return base


def get_settings_path():
    return os.path.join(get_data_dir(), "settings.json")


def get_data_path():
    return os.path.join(get_data_dir(), "dairy_farm_data.json")


def get_backup_dir():
    bd = os.path.join(get_data_dir(), "backups")
    os.makedirs(bd, exist_ok=True)
    return bd


def load_settings():
    path = get_settings_path()
    defaults = {
        "farm_name": "My Dairy Farm",
        "farm_location": "",
        "owner_name": "",
        "currency": "KES",
        "dark_mode": False,
        "auto_backup": True,
        "milk_price_default": "60",
        "daily_rate_default": "500",
    }
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                loaded = json.load(f)
                defaults.update(loaded)
        except:
            pass
    return defaults


def save_settings(settings):
    with open(get_settings_path(), "w", encoding="utf-8") as f:
        json.dump(settings, f, indent=2)


def load_data():
    path = get_data_path()
    defaults = {
        "farm": {"name": "My Dairy Farm", "location": "", "owner": ""},
        "herd": [],
        "milk_production": [],
        "milk_sales": [],
        "other_income": [],
        "feed_expenses": [],
        "vet_health": [],
        "labour": [],
        "operations": [],
        "assets": [],
        "breeding": [],
        "activity_log": [],
    }
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                loaded = json.load(f)
                # Merge with defaults for new fields
                for k, v in defaults.items():
                    if k not in loaded:
                        loaded[k] = v
                return loaded
        except Exception as e:
            print(f"Data load error: {e}")
    return defaults


def save_data(data):
    try:
        with open(get_data_path(), "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)
        # Auto-backup
        settings = load_settings()
        if settings.get("auto_backup", True):
            backup_file = os.path.join(get_backup_dir(), f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
            try:
                with open(backup_file, "w", encoding="utf-8") as f:
                    json.dump(data, f, indent=2, default=str)
                # Keep only last 10 backups
                backups = sorted([f for f in os.listdir(get_backup_dir()) if f.startswith("backup_")])
                for old in backups[:-10]:
                    os.remove(os.path.join(get_backup_dir(), old))
            except:
                pass
        return True
    except Exception as e:
        print(f"Save error: {e}")
        return False


def log_activity(data, action, details=""):
    data.setdefault("activity_log", []).insert(0, {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "action": action,
        "details": details,
    })
    data["activity_log"] = data["activity_log"][:100]  # Keep last 100
    save_data(data)


def validate_date(date_str):
    try:
        datetime.strptime(str(date_str), "%Y-%m-%d")
        return True
    except:
        return False


def safe_float(val, default=0.0):
    try:
        return float(val) if val not in (None, "", "None") else default
    except:
        return default


def _get_month(date_str):
    """Extract month index (0-11) from YYYY-MM-DD string."""
    try:
        return datetime.strptime(str(date_str), "%Y-%m-%d").month - 1
    except:
        return -1


def format_currency(amount):
    try:
        return f"{CURRENCY} {float(amount):,.0f}"
    except:
        return f"{CURRENCY} 0"


# ============================================================
# CHART HELPERS
# ============================================================

def create_pie_chart(labels, values, title, colors_list=None):
    if not CHARTS_AVAILABLE:
        return None
    try:
        fig, ax = plt.subplots(figsize=(3.2, 3.2), dpi=100)
        fig.patch.set_alpha(0)
        ax.set_facecolor('none')
        if colors_list:
            ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90, colors=colors_list,
                   textprops={'color': 'white', 'fontsize': 8})
        else:
            ax.pie(values, labels=labels, autopct='%1.1f%%', startangle=90,
                   textprops={'color': 'white', 'fontsize': 8})
        ax.set_title(title, color='white', fontsize=10, pad=10)
        plt.tight_layout()
        canvas = FigureCanvasKivyAgg(fig)
        return canvas
    except Exception as e:
        print(f"Chart error: {e}")
        return None


def create_line_chart(dates, values, title, ylabel="Amount (KES)"):
    if not CHARTS_AVAILABLE or not dates:
        return None
    try:
        fig, ax = plt.subplots(figsize=(3.2, 2.4), dpi=100)
        fig.patch.set_alpha(0)
        ax.set_facecolor('none')
        ax.plot(dates, values, marker='o', linewidth=2, markersize=4, color='#4db87a')
        ax.set_title(title, color='white', fontsize=10)
        ax.set_ylabel(ylabel, color='white', fontsize=8)
        ax.tick_params(colors='white', labelsize=7)
        ax.spines['bottom'].set_color('white')
        ax.spines['left'].set_color('white')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(True, alpha=0.3, color='white')
        plt.tight_layout()
        return FigureCanvasKivyAgg(fig)
    except Exception as e:
        print(f"Chart error: {e}")
        return None


def create_bar_chart(labels, values, title, color='#2d8a5e'):
    if not CHARTS_AVAILABLE or not labels:
        return None
    try:
        fig, ax = plt.subplots(figsize=(3.2, 2.4), dpi=100)
        fig.patch.set_alpha(0)
        ax.set_facecolor('none')
        bars = ax.bar(labels, values, color=color, alpha=0.8)
        ax.set_title(title, color='white', fontsize=10)
        ax.tick_params(colors='white', labelsize=7)
        ax.spines['bottom'].set_color('white')
        ax.spines['left'].set_color('white')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(True, alpha=0.3, color='white', axis='y')
        plt.tight_layout()
        return FigureCanvasKivyAgg(fig)
    except Exception as e:
        print(f"Chart error: {e}")
        return None


# ============================================================
# UI HELPERS
# ============================================================

class ThemedCard(MDCard):
    """Enhanced card with theme support."""
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.elevation = 2
        self.radius = [dp(12),]
        self.padding = dp(12)
        self.md_bg_color = THEME["card"]


class KPICard(ThemedCard):
    def __init__(self, title, value, color, subtitle="", **kwargs):
        super().__init__(**kwargs)
        self.orientation = "vertical"
        self.size_hint_y = None
        self.height = dp(100)
        self.add_widget(MDLabel(
            text=title, theme_text_color="Custom",
            text_color=THEME["text_secondary"], font_style="Caption",
            halign="center", size_hint_y=None, height=dp(18)
        ))
        self.add_widget(MDLabel(
            text=value, theme_text_color="Custom",
            text_color=color, font_style="H5",
            halign="center", bold=True, size_hint_y=None, height=dp(32)
        ))
        if subtitle:
            self.add_widget(MDLabel(
                text=subtitle, theme_text_color="Custom",
                text_color=THEME["text_secondary"], font_style="Caption",
                halign="center", size_hint_y=None, height=dp(18)
            ))


class ActionCard(ThemedCard):
    def __init__(self, title, subtitle, on_edit=None, on_delete=None, on_tap=None, **kwargs):
        super().__init__(**kwargs)
        self.orientation = "vertical"
        self.size_hint_y = None
        self.height = dp(90)

        main = MDBoxLayout()
        text_area = MDBoxLayout(orientation="vertical", size_hint_x=0.78)
        text_area.add_widget(MDLabel(
            text=title, theme_text_color="Custom",
            text_color=THEME["primary"], font_style="Subtitle1",
            bold=True, size_hint_y=None, height=dp(22)
        ))
        text_area.add_widget(MDLabel(
            text=subtitle, theme_text_color="Secondary",
            font_style="Caption", size_hint_y=None, height=dp(36)
        ))
        main.add_widget(text_area)

        btn_area = MDBoxLayout(size_hint_x=0.22, spacing=dp(2))
        if on_edit:
            btn_area.add_widget(MDIconButton(
                icon="pencil", theme_text_color="Custom",
                text_color=THEME["info"], on_release=on_edit,
                icon_size=dp(18)
            ))
        if on_delete:
            btn_area.add_widget(MDIconButton(
                icon="delete", theme_text_color="Custom",
                text_color=THEME["danger"], on_release=on_delete,
                icon_size=dp(18)
            ))
        main.add_widget(btn_area)
        self.add_widget(main)

        if on_tap:
            self.bind(on_release=on_tap)


class SearchBar(MDBoxLayout):
    def __init__(self, on_search, **kwargs):
        super().__init__(**kwargs)
        self.size_hint_y = None
        self.height = dp(50)
        self.padding = [dp(4), dp(4), dp(4), dp(4)]
        self.md_bg_color = THEME["card"]
        self.radius = [dp(8),]

        self.search_field = MDTextField(
            hint_text="Search...",
            icon_right="magnify",
            mode="rectangle",
            size_hint_x=0.85,
            height=dp(40)
        )
        self.search_field.bind(text=lambda inst, val: on_search(val))
        self.add_widget(self.search_field)

        clear_btn = MDIconButton(
            icon="close-circle", theme_text_color="Custom",
            text_color=THEME["text_secondary"],
            on_release=lambda x: [setattr(self.search_field, "text", ""), on_search("")],
            size_hint_x=0.15
        )
        self.add_widget(clear_btn)


class MobileDialog:
    @staticmethod
    def create(title, fields, on_save, on_cancel=None, size_hint=(0.92, None)):
        content = MDBoxLayout(orientation="vertical", spacing=dp(6), size_hint_y=None,
                              padding=[dp(8), dp(8), dp(8), dp(8)])
        field_refs = {}
        total_h = dp(16)

        for field in fields:
            h = dp(70) if field.get("multiline") else dp(55)
            tf = MDTextField(
                hint_text=field.get("hint", ""),
                icon_right=field.get("icon", ""),
                multiline=field.get("multiline", False),
                mode="rectangle",
                size_hint_y=None, height=h,
                text=field.get("text", ""),
                input_filter=field.get("input_filter", None),
            )
            field_refs[field.get("key", field["hint"])] = tf
            content.add_widget(tf)
            total_h += h + dp(6)

        content.height = total_h

        def _save(*args):
            vals = {k: v.text for k, v in field_refs.items()}
            # Validation
            errors = []
            for field in fields:
                key = field.get("key", field["hint"])
                if field.get("required") and not vals.get(key, "").strip():
                    errors.append(f"{field['hint']} is required")
                if field.get("is_date") and vals.get(key) and not validate_date(vals[key]):
                    errors.append(f"{field['hint']} must be YYYY-MM-DD")
                if field.get("is_number") and vals.get(key):
                    try:
                        float(vals[key])
                    except:
                        errors.append(f"{field['hint']} must be a number")

            if errors:
                show_snackbar("; ".join(errors[:2]), THEME["danger"])
                return
            on_save(vals)
            dialog.dismiss()

        def _cancel(*args):
            if on_cancel: on_cancel()
            dialog.dismiss()

        dialog = MDDialog(
            title=title, type="custom", content_cls=content,
            size_hint=size_hint,
            buttons=[
                MDFlatButton(text="CANCEL", on_release=_cancel),
                MDRaisedButton(text="SAVE", md_bg_color=THEME["success"], on_release=_save)
            ],
        )
        return dialog, field_refs


def show_snackbar(text, color=THEME["success"], duration=2):
    try:
        # KivyMD 2.0+ style
        sb = MDSnackbar(
            MDLabel(text=text, theme_text_color="Custom", text_color=[1,1,1,1], font_style="Body2"),
            md_bg_color=color, duration=duration,
        )
    except TypeError:
        # KivyMD 1.2.0 style
        sb = MDSnackbar(text=text, md_bg_color=color, duration=duration)
    sb.open()


class ConfirmDialog:
    @staticmethod
    def show(title, text, on_confirm, confirm_text="DELETE", confirm_color=None):
        if confirm_color is None:
            confirm_color = THEME["danger"]
        dialog = MDDialog(
            title=title, text=text,
            buttons=[
                MDFlatButton(text="CANCEL", on_release=lambda x: dialog.dismiss()),
                MDRaisedButton(text=confirm_text, md_bg_color=confirm_color,
                               on_release=lambda x: [dialog.dismiss(), on_confirm()])
            ]
        )
        dialog.open()


# ============================================================
# BASE SCREEN
# ============================================================

class BaseScreen(MDScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.data = load_data()
        self.settings = load_settings()
        self.dialog = None
        self.search_query = ""

    def reload_data(self):
        self.data = load_data()
        self.settings = load_settings()

    def refresh(self):
        pass

    def create_scroll_layout(self):
        scroll = ScrollView()
        content = MDBoxLayout(
            orientation="vertical", size_hint_y=None,
            padding=[dp(12), dp(8), dp(12), dp(80)],
            spacing=dp(8)
        )
        content.bind(minimum_height=content.setter("height"))
        scroll.add_widget(content)
        return scroll, content

    def add_screen_title(self, content, title):
        content.add_widget(MDLabel(
            text=title, theme_text_color="Custom",
            text_color=THEME["primary"], font_style="H5",
            halign="center", size_hint_y=None, height=dp(40), bold=True
        ))

    def add_fab(self, layout, callback, icon="plus"):
        fab = MDFloatingActionButton(
            icon=icon, md_bg_color=THEME["accent"],
            pos_hint={"right": 0.95, "y": 0.04},
            on_release=callback, elevation=4
        )
        layout.add_widget(fab)
        return fab

    def filter_items(self, items, query, keys):
        if not query:
            return items
        q = query.lower()
        return [item for item in items if any(q in str(item.get(k, "")).lower() for k in keys)]


# ============================================================
# DASHBOARD SCREEN
# ============================================================

class DashboardScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        scroll, content = self.create_scroll_layout()

        # Header
        farm_name = self.settings.get("farm_name", "My Dairy Farm")
        content.add_widget(MDLabel(
            text=farm_name.upper(), theme_text_color="Custom",
            text_color=THEME["primary"], font_style="H4",
            halign="center", size_hint_y=None, height=dp(44), bold=True
        ))
        content.add_widget(MDLabel(
            text=datetime.now().strftime("%A, %d %B %Y"),
            theme_text_color="Custom", text_color=THEME["text_secondary"],
            font_style="Caption", halign="center", size_hint_y=None, height=dp(22)
        ))

        # KPIs
        income = self._calc_income()
        expenses = self._calc_expenses()
        net = income - expenses
        milking = len([c for c in self.data["herd"] if c.get("status", "").lower() == "milking"])
        total_herd = len(self.data["herd"])

        grid = MDGridLayout(cols=2, spacing=dp(8), size_hint_y=None, height=dp(220))
        grid.add_widget(KPICard("INCOME", format_currency(income), THEME["success"]))
        grid.add_widget(KPICard("EXPENSES", format_currency(expenses), THEME["danger"]))
        net_color = THEME["success"] if net >= 0 else THEME["danger"]
        grid.add_widget(KPICard("NET P&L", format_currency(net), net_color))
        grid.add_widget(KPICard("HERD", f"{milking}/{total_herd}", THEME["accent"], "Milking/Total"))
        content.add_widget(grid)

        # Today's Milk
        today = datetime.now().strftime("%Y-%m-%d")
        today_milk = sum(safe_float(x.get("quantity")) for x in self.data["milk_production"] if x.get("date", "") == today)
        content.add_widget(KPICard("TODAY'S MILK", f"{today_milk:.1f} L", THEME["info"], f"{len([x for x in self.data['milk_production'] if x.get('date')==today])} records"))

        # Quick Actions
        content.add_widget(MDLabel(text="QUICK ACTIONS", theme_text_color="Custom",
            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))

        actions = MDBoxLayout(size_hint_y=None, height=dp(44), spacing=dp(6))
        actions.add_widget(MDRaisedButton(text="+ Milk", md_bg_color=THEME["info"],
            on_release=lambda x: setattr(self.manager, "current", "milk_production")))
        actions.add_widget(MDRaisedButton(text="+ Sale", md_bg_color=THEME["success"],
            on_release=lambda x: setattr(self.manager, "current", "milk_sales")))
        actions.add_widget(MDRaisedButton(text="+ Cow", md_bg_color=THEME["accent"],
            on_release=lambda x: setattr(self.manager, "current", "herd")))
        actions.add_widget(MDRaisedButton(text="+ Expense", md_bg_color=THEME["danger"],
            on_release=lambda x: setattr(self.manager, "current", "feed_expenses")))
        content.add_widget(actions)

        # Recent Activity
        content.add_widget(MDLabel(text="RECENT ACTIVITY", theme_text_color="Custom",
            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))

        logs = self.data.get("activity_log", [])[:10]
        if not logs:
            content.add_widget(MDLabel(text="No recent activity.", theme_text_color="Hint",
                halign="center", size_hint_y=None, height=dp(50)))
        for log in logs:
            card = ThemedCard(orientation="horizontal", size_hint_y=None, height=dp(50), padding=dp(8))
            card.add_widget(MDLabel(
                text=f"{log.get('timestamp', '-')[11:16]}  {log.get('action', '')}",
                theme_text_color="Custom", text_color=THEME["text"],
                font_style="Caption", size_hint_x=0.7
            ))
            card.add_widget(MDLabel(
                text=log.get('details', '')[:20],
                theme_text_color="Custom", text_color=THEME["text_secondary"],
                font_style="Caption", halign="right", size_hint_x=0.3
            ))
            content.add_widget(card)

        layout.add_widget(scroll)
        self.add_widget(layout)

    def _calc_income(self):
        milk = sum(safe_float(x.get("amount")) for x in self.data.get("milk_sales", []))
        other = sum(safe_float(x.get("amount")) for x in self.data.get("other_income", []))
        return milk + other

    def _calc_expenses(self):
        feed = sum(safe_float(x.get("amount")) for x in self.data.get("feed_expenses", []))
        vet = sum(safe_float(x.get("cost")) for x in self.data.get("vet_health", []))
        lab = sum(safe_float(x.get("amount")) for x in self.data.get("labour", []))
        ops = sum(safe_float(x.get("amount")) for x in self.data.get("operations", []))
        return feed + vet + lab + ops

    def refresh(self):
        self.clear_widgets()
        self.reload_data()
        self.build_ui()


# ============================================================
# HERD SCREEN
# ============================================================

class HerdScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        scroll, content = self.create_scroll_layout()
        self.add_screen_title(content, "HERD REGISTER")

        # Stats
        total = len(self.data["herd"])
        milking = len([c for c in self.data["herd"] if c.get("status", "").lower() == "milking"])
        dry = len([c for c in self.data["herd"] if c.get("status", "").lower() == "dry"])
        heifer = len([c for c in self.data["herd"] if c.get("status", "").lower() == "heifer"])

        stats = MDBoxLayout(size_hint_y=None, height=dp(32), spacing=dp(6))
        for label, count, color in [("Total", total, THEME["primary"]), 
                                     ("Milking", milking, THEME["success"]),
                                     ("Dry", dry, THEME["warning"]),
                                     ("Heifer", heifer, THEME["info"])]:
            chip = MDChip(text=f"{label}: {count}", md_bg_color=color)
            chip.text_color = [1, 1, 1, 1]
            stats.add_widget(chip)
        content.add_widget(stats)

        # Search
        content.add_widget(SearchBar(on_search=self.on_search))

        # List
        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
        self.list_container.bind(minimum_height=self.list_container.setter("height"))
        content.add_widget(self.list_container)
        self.refresh_list()

        layout.add_widget(scroll)
        self.add_fab(layout, self.show_add_dialog)
        self.add_widget(layout)

    def on_search(self, query):
        self.search_query = query
        self.refresh_list()

    def refresh_list(self):
        self.list_container.clear_widgets()
        items = self.filter_items(self.data["herd"], self.search_query, ["tag_no", "name", "breed", "status"])

        if not items:
            msg = "No cows found." if self.search_query else "No cows registered yet.\nTap + to add your first cow."
            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
                halign="center", size_hint_y=None, height=dp(100)))
            return

        for i, cow in enumerate(items):
            orig_idx = self.data["herd"].index(cow)
            status_colors = {
                "milking": THEME["success"], "dry": THEME["warning"],
                "heifer": THEME["info"], "sold": THEME["danger"], "dead": THEME["danger"]
            }
            status_color = status_colors.get(cow.get("status", "").lower(), THEME["text_secondary"])

            # Calculate age
            age_text = ""
            if cow.get("dob"):
                try:
                    dob = datetime.strptime(cow["dob"], "%Y-%m-%d")
                    age_days = (datetime.now() - dob).days
                    age_text = f"Age: {age_days // 365}y {(age_days % 365) // 30}m"
                except:
                    pass

            card = ActionCard(
                title=f"{cow.get('tag_no', 'N/A')} - {cow.get('name', 'Unknown')}",
                subtitle=f"Breed: {cow.get('breed', '-')} | Status: {cow.get('status', '-')} | {age_text}",
                on_edit=partial(self.edit_cow, orig_idx),
                on_delete=partial(self.delete_cow, orig_idx),
                on_tap=partial(self.view_cow, orig_idx)
            )
            self.list_container.add_widget(card)

    def show_add_dialog(self, *args):
        dialog, fields = MobileDialog.create("Add New Cow", [
            {"key": "tag", "hint": "Tag Number *", "icon": "identifier", "required": True},
            {"key": "name", "hint": "Cow Name", "icon": "cow"},
            {"key": "breed", "hint": "Breed", "icon": "dna"},
            {"key": "status", "hint": "Status (Milking/Dry/Heifer)", "icon": "information", "text": "Milking"},
            {"key": "dob", "hint": "Date of Birth (YYYY-MM-DD)", "icon": "calendar", "is_date": True},
            {"key": "cost", "hint": "Purchase Cost", "icon": "cash", "input_filter": "float"},
            {"key": "notes", "hint": "Notes", "icon": "note-text", "multiline": True},
        ], on_save=self.save_cow)
        self.dialog = dialog
        self.dialog.open()

    def save_cow(self, values):
        self.data["herd"].append({
            "tag_no": values["tag"], "name": values["name"],
            "breed": values["breed"], "status": values["status"] or "Milking",
            "dob": values["dob"], "purchase_date": datetime.now().strftime("%Y-%m-%d"),
            "purchase_cost": values["cost"] or "0", "notes": values["notes"],
        })
        save_data(self.data)
        log_activity(self.data, "Added cow", values["tag"])
        show_snackbar(f"Cow {values['tag']} added!")
        self.refresh()

    def edit_cow(self, index, *args):
        cow = self.data["herd"][index]
        dialog, fields = MobileDialog.create("Edit Cow", [
            {"key": "tag", "hint": "Tag Number", "icon": "identifier", "text": cow.get("tag_no", "")},
            {"key": "name", "hint": "Cow Name", "icon": "cow", "text": cow.get("name", "")},
            {"key": "breed", "hint": "Breed", "icon": "dna", "text": cow.get("breed", "")},
            {"key": "status", "hint": "Status", "icon": "information", "text": cow.get("status", "Milking")},
            {"key": "dob", "hint": "Date of Birth", "icon": "calendar", "text": cow.get("dob", ""), "is_date": True},
            {"key": "cost", "hint": "Purchase Cost", "icon": "cash", "text": str(cow.get("purchase_cost", "")), "input_filter": "float"},
            {"key": "notes", "hint": "Notes", "icon": "note-text", "text": cow.get("notes", ""), "multiline": True},
        ], on_save=lambda v: self.update_cow(index, v))
        self.dialog = dialog
        self.dialog.open()

    def update_cow(self, index, values):
        self.data["herd"][index].update({
            "tag_no": values["tag"], "name": values["name"],
            "breed": values["breed"], "status": values["status"],
            "dob": values["dob"], "purchase_cost": values["cost"] or "0",
            "notes": values["notes"],
        })
        save_data(self.data)
        log_activity(self.data, "Updated cow", values["tag"])
        show_snackbar("Cow updated!")
        self.refresh()

    def delete_cow(self, index, *args):
        tag = self.data["herd"][index].get("tag_no", "Unknown")
        def confirm():
            self.data["herd"].pop(index)
            save_data(self.data)
            log_activity(self.data, "Deleted cow", tag)
            show_snackbar(f"Cow {tag} removed")
            self.refresh()
        ConfirmDialog.show("Delete Cow?", f"Remove {tag} permanently?", confirm)

    def view_cow(self, index, *args):
        cow = self.data["herd"][index]
        # Production stats
        cow_milk = sum(safe_float(x.get("quantity")) for x in self.data["milk_production"] if x.get("cow_tag") == cow.get("tag_no"))
        cow_vet = sum(safe_float(x.get("cost")) for x in self.data["vet_health"] if x.get("cow_tag") == cow.get("tag_no"))

        content = MDBoxLayout(orientation="vertical", spacing=dp(8), size_hint_y=None)
        content.height = dp(200)
        info_lines = [
            f"Tag: {cow.get('tag_no', '-')}",
            f"Name: {cow.get('name', '-')}",
            f"Breed: {cow.get('breed', '-')}",
            f"Status: {cow.get('status', '-')}",
            f"DOB: {cow.get('dob', '-')}",
            f"Purchased: {cow.get('purchase_date', '-')}",
            f"Total Milk: {cow_milk:.1f} L",
            f"Vet Costs: {format_currency(cow_vet)}",
        ]
        for line in info_lines:
            content.add_widget(MDLabel(text=line, theme_text_color="Custom",
                text_color=THEME["text"], font_style="Body1", size_hint_y=None, height=dp(22)))

        dialog = MDDialog(title="Cow Profile", type="custom", content_cls=content,
            buttons=[MDFlatButton(text="CLOSE", on_release=lambda x: dialog.dismiss())])
        dialog.open()

    def refresh(self):
        self.clear_widgets()
        self.reload_data()
        self.build_ui()


# ============================================================
# MILK PRODUCTION SCREEN
# ============================================================

class MilkProductionScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        scroll, content = self.create_scroll_layout()
        self.add_screen_title(content, "MILK PRODUCTION")

        today = datetime.now().strftime("%Y-%m-%d")
        today_recs = [x for x in self.data["milk_production"] if x.get("date", "") == today]
        today_total = sum(safe_float(x.get("quantity")) for x in today_recs)

        content.add_widget(KPICard("TODAY'S TOTAL", f"{today_total:.1f} L", THEME["info"], f"{len(today_recs)} records"))

        # 7-day trend chart
        if CHARTS_AVAILABLE:
            dates = []
            amounts = []
            for i in range(6, -1, -1):
                d = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
                amt = sum(safe_float(x.get("quantity")) for x in self.data["milk_production"] if x.get("date") == d)
                dates.append(d[5:])  # MM-DD
                amounts.append(amt)
            chart = create_bar_chart(dates, amounts, "Last 7 Days (L)", color='#4db87a')
            if chart:
                try:
                    chart.size_hint_y = None
                    chart.height = dp(200)
                    content.add_widget(chart)
                except Exception as e:
                    print(f"Chart display error: {e}")

        content.add_widget(SearchBar(on_search=self.on_search))

        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
        self.list_container.bind(minimum_height=self.list_container.setter("height"))
        content.add_widget(self.list_container)
        self.refresh_list()

        layout.add_widget(scroll)
        self.add_fab(layout, self.show_add_dialog)
        self.add_widget(layout)

    def on_search(self, query):
        self.search_query = query
        self.refresh_list()

    def refresh_list(self):
        self.list_container.clear_widgets()
        all_items = sorted(self.data["milk_production"], key=lambda x: x.get("date", ""), reverse=True)
        items = self.filter_items(all_items, self.search_query, ["date", "cow_tag", "session", "quality"])

        if not items:
            msg = "No records found." if self.search_query else "No milk production records yet."
            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
                halign="center", size_hint_y=None, height=dp(80)))
            return

        for i, rec in enumerate(items[:50]):  # Limit to 50 for performance
            orig_idx = self.data["milk_production"].index(rec)
            card = ActionCard(
                title=f"{rec.get('date', '-')} | {rec.get('session', 'AM')}",
                subtitle=f"Cow: {rec.get('cow_tag', '-')} | {rec.get('quantity', '0')} L | {rec.get('quality', 'Good')}",
                on_edit=partial(self.edit_record, orig_idx),
                on_delete=partial(self.delete_record, orig_idx)
            )
            self.list_container.add_widget(card)

    def show_add_dialog(self, *args):
        cow_tags = [c.get("tag_no", "") for c in self.data["herd"] if c.get("tag_no")]
        cow_hint = f"Cow Tag ({', '.join(cow_tags[:3])}...)" if cow_tags else "Cow Tag"
        dialog, fields = MobileDialog.create("Record Milk Production", [
            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
            {"key": "session", "hint": "Session (AM/PM/Evening)", "icon": "clock", "text": "AM", "required": True},
            {"key": "cow", "hint": cow_hint, "icon": "cow", "required": True},
            {"key": "qty", "hint": "Quantity (Liters)", "icon": "cup-water", "input_filter": "float", "required": True, "is_number": True},
            {"key": "quality", "hint": "Quality", "icon": "star", "text": "Good"},
            {"key": "notes", "hint": "Notes", "icon": "note-text", "multiline": True},
        ], on_save=self.save_record)
        self.dialog = dialog
        self.dialog.open()

    def save_record(self, values):
        self.data["milk_production"].append({
            "date": values["date"], "session": values["session"],
            "cow_tag": values["cow"], "quantity": values["qty"],
            "quality": values["quality"], "notes": values["notes"],
            "recorded_at": datetime.now().strftime("%Y-%m-%d %H:%M")
        })
        save_data(self.data)
        log_activity(self.data, "Milk recorded", f"{values['cow']} - {values['qty']}L")
        show_snackbar("Milk record saved!")
        self.refresh()

    def edit_record(self, index, *args):
        rec = self.data["milk_production"][index]
        dialog, fields = MobileDialog.create("Edit Milk Record", [
            {"key": "date", "hint": "Date", "icon": "calendar", "text": rec.get("date", ""), "required": True, "is_date": True},
            {"key": "session", "hint": "Session", "icon": "clock", "text": rec.get("session", "")},
            {"key": "cow", "hint": "Cow Tag", "icon": "cow", "text": rec.get("cow_tag", "")},
            {"key": "qty", "hint": "Quantity", "icon": "cup-water", "text": str(rec.get("quantity", "")), "input_filter": "float", "is_number": True},
            {"key": "quality", "hint": "Quality", "icon": "star", "text": rec.get("quality", "")},
            {"key": "notes", "hint": "Notes", "icon": "note-text", "text": rec.get("notes", ""), "multiline": True},
        ], on_save=lambda v: self.update_record(index, v))
        self.dialog = dialog
        self.dialog.open()

    def update_record(self, index, values):
        self.data["milk_production"][index].update({
            "date": values["date"], "session": values["session"],
            "cow_tag": values["cow"], "quantity": values["qty"],
            "quality": values["quality"], "notes": values["notes"],
        })
        save_data(self.data)
        show_snackbar("Record updated!")
        self.refresh()

    def delete_record(self, index, *args):
        rec = self.data["milk_production"][index]
        def confirm():
            self.data["milk_production"].pop(index)
            save_data(self.data)
            show_snackbar("Record deleted")
            self.refresh()
        ConfirmDialog.show("Delete Record?", f"Remove {rec.get('date', '')} - {rec.get('cow_tag', '')}?", confirm)

    def refresh(self):
        self.clear_widgets()
        self.reload_data()
        self.build_ui()


# ============================================================
# MILK SALES SCREEN
# ============================================================

class MilkSalesScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.showing = "sales"
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        scroll, content = self.create_scroll_layout()
        self.add_screen_title(content, "MILK SALES & INCOME")

        total_sales = sum(safe_float(x.get("amount")) for x in self.data["milk_sales"])
        total_other = sum(safe_float(x.get("amount")) for x in self.data["other_income"])
        total_liters = sum(safe_float(x.get("liters")) for x in self.data["milk_sales"])

        content.add_widget(KPICard("TOTAL SALES", format_currency(total_sales), THEME["success"], f"{total_liters:.0f} liters"))
        content.add_widget(KPICard("OTHER INCOME", format_currency(total_other), THEME["info"]))

        # Tabs
        tab_bar = MDBoxLayout(size_hint_y=None, height=dp(40), spacing=dp(6))
        self.sales_btn = MDRaisedButton(text="Milk Sales", md_bg_color=THEME["primary"],
            on_release=lambda x: self.switch_tab("sales"))
        self.other_btn = MDFlatButton(text="Other Income", on_release=lambda x: self.switch_tab("other"))
        tab_bar.add_widget(self.sales_btn)
        tab_bar.add_widget(self.other_btn)
        content.add_widget(tab_bar)

        content.add_widget(SearchBar(on_search=self.on_search))

        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
        self.list_container.bind(minimum_height=self.list_container.setter("height"))
        content.add_widget(self.list_container)
        self.refresh_list()

        layout.add_widget(scroll)
        self.add_fab(layout, self.show_add_dialog)
        self.add_widget(layout)

    def switch_tab(self, tab):
        self.showing = tab
        self.search_query = ""
        if tab == "sales":
            self.sales_btn.md_bg_color = THEME["primary"]
            self.other_btn.md_bg_color = [0.9, 0.9, 0.9, 1]
        else:
            self.other_btn.md_bg_color = THEME["primary"]
            self.sales_btn.md_bg_color = [0.9, 0.9, 0.9, 1]
        self.refresh_list()

    def on_search(self, query):
        self.search_query = query
        self.refresh_list()

    def refresh_list(self):
        self.list_container.clear_widgets()

        if self.showing == "sales":
            all_items = sorted(self.data["milk_sales"], key=lambda x: x.get("date", ""), reverse=True)
            items = self.filter_items(all_items, self.search_query, ["date", "buyer", "amount"])
            if not items:
                self.list_container.add_widget(MDLabel(text="No sales records.", theme_text_color="Hint",
                    halign="center", size_hint_y=None, height=dp(60)))
            for item in items[:50]:
                idx = self.data["milk_sales"].index(item)
                card = ActionCard(
                    title=f"{item.get('date', '-')} | {item.get('buyer', 'Unknown')}",
                    subtitle=f"{item.get('liters', '0')}L @ KES {item.get('price_per_liter', '0')}/L = {format_currency(item.get('amount', '0'))}",
                    on_edit=partial(self.edit_item, idx, "sales"),
                    on_delete=partial(self.delete_item, idx, "sales")
                )
                self.list_container.add_widget(card)
        else:
            all_items = sorted(self.data["other_income"], key=lambda x: x.get("date", ""), reverse=True)
            items = self.filter_items(all_items, self.search_query, ["date", "source", "amount"])
            if not items:
                self.list_container.add_widget(MDLabel(text="No other income records.", theme_text_color="Hint",
                    halign="center", size_hint_y=None, height=dp(60)))
            for item in items[:50]:
                idx = self.data["other_income"].index(item)
                card = ActionCard(
                    title=f"{item.get('date', '-')} | {item.get('source', 'Unknown')}",
                    subtitle=f"{format_currency(item.get('amount', '0'))} | {item.get('notes', '-')}",
                    on_edit=partial(self.edit_item, idx, "other"),
                    on_delete=partial(self.delete_item, idx, "other")
                )
                self.list_container.add_widget(card)

    def show_add_dialog(self, *args):
        if self.showing == "sales":
            default_price = self.settings.get("milk_price_default", "60")
            dialog, fields = MobileDialog.create("Record Milk Sale", [
                {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
                {"key": "buyer", "hint": "Buyer Name", "icon": "account", "required": True},
                {"key": "liters", "hint": "Liters Sold", "icon": "cup-water", "input_filter": "float", "required": True, "is_number": True},
                {"key": "price", "hint": "Price per Liter (KES)", "icon": "cash", "input_filter": "float", "text": default_price, "is_number": True},
            ], on_save=self.save_sale)
        else:
            dialog, fields = MobileDialog.create("Record Other Income", [
                {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
                {"key": "source", "hint": "Income Source", "icon": "tag", "required": True},
                {"key": "amount", "hint": "Amount (KES)", "icon": "cash", "input_filter": "float", "required": True, "is_number": True},
                {"key": "notes", "hint": "Notes", "icon": "note-text", "multiline": True},
            ], on_save=self.save_other)
        self.dialog = dialog
        self.dialog.open()

    def save_sale(self, values):
        liters = safe_float(values["liters"])
        price = safe_float(values["price"])
        self.data["milk_sales"].append({
            "date": values["date"], "buyer": values["buyer"],
            "liters": str(liters), "price_per_liter": str(price),
            "amount": str(liters * price),
        })
        save_data(self.data)
        log_activity(self.data, "Milk sale", f"{values['buyer']} - {liters}L")
        show_snackbar("Sale recorded!")
        self.refresh()

    def save_other(self, values):
        self.data["other_income"].append({
            "date": values["date"], "source": values["source"],
            "amount": values["amount"], "notes": values["notes"],
        })
        save_data(self.data)
        log_activity(self.data, "Other income", f"{values['source']} - {values['amount']}")
        show_snackbar("Income recorded!")
        self.refresh()

    def edit_item(self, index, type, *args):
        if type == "sales":
            item = self.data["milk_sales"][index]
            dialog, fields = MobileDialog.create("Edit Sale", [
                {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "required": True, "is_date": True},
                {"key": "buyer", "hint": "Buyer", "icon": "account", "text": item.get("buyer", "")},
                {"key": "liters", "hint": "Liters", "icon": "cup-water", "text": str(item.get("liters", "")), "input_filter": "float", "is_number": True},
                {"key": "price", "hint": "Price/Liter", "icon": "cash", "text": str(item.get("price_per_liter", "")), "input_filter": "float", "is_number": True},
            ], on_save=lambda v: self.update_sale(index, v))
        else:
            item = self.data["other_income"][index]
            dialog, fields = MobileDialog.create("Edit Income", [
                {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "required": True, "is_date": True},
                {"key": "source", "hint": "Source", "icon": "tag", "text": item.get("source", "")},
                {"key": "amount", "hint": "Amount", "icon": "cash", "text": str(item.get("amount", "")), "input_filter": "float", "is_number": True},
                {"key": "notes", "hint": "Notes", "icon": "note-text", "text": item.get("notes", ""), "multiline": True},
            ], on_save=lambda v: self.update_other(index, v))
        self.dialog = dialog
        self.dialog.open()

    def update_sale(self, index, values):
        liters = safe_float(values["liters"])
        price = safe_float(values["price"])
        self.data["milk_sales"][index].update({
            "date": values["date"], "buyer": values["buyer"],
            "liters": str(liters), "price_per_liter": str(price),
            "amount": str(liters * price),
        })
        save_data(self.data)
        show_snackbar("Sale updated!")
        self.refresh()

    def update_other(self, index, values):
        self.data["other_income"][index].update({
            "date": values["date"], "source": values["source"],
            "amount": values["amount"], "notes": values["notes"],
        })
        save_data(self.data)
        show_snackbar("Income updated!")
        self.refresh()

    def delete_item(self, index, type, *args):
        key = "milk_sales" if type == "sales" else "other_income"
        item = self.data[key][index]
        def confirm():
            self.data[key].pop(index)
            save_data(self.data)
            show_snackbar("Deleted")
            self.refresh()
        ConfirmDialog.show("Delete?", "Remove this record permanently?", confirm)

    def refresh(self):
        self.clear_widgets()
        self.reload_data()
        self.build_ui()


# ============================================================
# FEED EXPENSES SCREEN
# ============================================================

class FeedExpensesScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        scroll, content = self.create_scroll_layout()
        self.add_screen_title(content, "FEED EXPENSES")

        total = sum(safe_float(x.get("amount")) for x in self.data["feed_expenses"])
        content.add_widget(KPICard("TOTAL FEED COSTS", format_currency(total), THEME["danger"]))

        content.add_widget(SearchBar(on_search=self.on_search))

        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
        self.list_container.bind(minimum_height=self.list_container.setter("height"))
        content.add_widget(self.list_container)
        self.refresh_list()

        layout.add_widget(scroll)
        self.add_fab(layout, self.show_add_dialog)
        self.add_widget(layout)

    def on_search(self, query):
        self.search_query = query
        self.refresh_list()

    def refresh_list(self):
        self.list_container.clear_widgets()
        all_items = sorted(self.data["feed_expenses"], key=lambda x: x.get("date", ""), reverse=True)
        items = self.filter_items(all_items, self.search_query, ["date", "feed_type", "supplier"])

        if not items:
            msg = "No records found." if self.search_query else "No feed expenses recorded."
            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
                halign="center", size_hint_y=None, height=dp(80)))
            return

        for item in items[:50]:
            idx = self.data["feed_expenses"].index(item)
            card = ActionCard(
                title=f"{item.get('date', '-')} | {item.get('feed_type', 'Unknown')}",
                subtitle=f"{item.get('supplier', '-')} | {item.get('quantity', '-')} {item.get('unit', 'kg')} | {format_currency(item.get('amount', '0'))}",
                on_edit=partial(self.edit_item, idx),
                on_delete=partial(self.delete_item, idx)
            )
            self.list_container.add_widget(card)

    def show_add_dialog(self, *args):
        dialog, fields = MobileDialog.create("Add Feed Expense", [
            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
            {"key": "type", "hint": "Feed Type (e.g. Dairy Meal)", "icon": "corn", "required": True},
            {"key": "supplier", "hint": "Supplier Name", "icon": "truck-delivery"},
            {"key": "qty", "hint": "Quantity", "icon": "numeric", "input_filter": "float", "is_number": True},
            {"key": "unit", "hint": "Unit (kg/bale/bag)", "icon": "scale", "text": "kg"},
            {"key": "amount", "hint": "Total Cost (KES)", "icon": "cash", "input_filter": "float", "required": True, "is_number": True},
        ], on_save=self.save_item)
        self.dialog = dialog
        self.dialog.open()

    def save_item(self, values):
        self.data["feed_expenses"].append({
            "date": values["date"], "feed_type": values["type"],
            "supplier": values["supplier"], "quantity": values["qty"],
            "unit": values["unit"], "amount": values["amount"],
        })
        save_data(self.data)
        log_activity(self.data, "Feed expense", f"{values['type']} - {values['amount']}")
        show_snackbar("Feed expense saved!")
        self.refresh()

    def edit_item(self, index, *args):
        item = self.data["feed_expenses"][index]
        dialog, fields = MobileDialog.create("Edit Feed Expense", [
            {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "is_date": True},
            {"key": "type", "hint": "Feed Type", "icon": "corn", "text": item.get("feed_type", "")},
            {"key": "supplier", "hint": "Supplier", "icon": "truck-delivery", "text": item.get("supplier", "")},
            {"key": "qty", "hint": "Quantity", "icon": "numeric", "text": str(item.get("quantity", "")), "input_filter": "float", "is_number": True},
            {"key": "unit", "hint": "Unit", "icon": "scale", "text": item.get("unit", "")},
            {"key": "amount", "hint": "Amount", "icon": "cash", "text": str(item.get("amount", "")), "input_filter": "float", "is_number": True},
        ], on_save=lambda v: self.update_item(index, v))
        self.dialog = dialog
        self.dialog.open()

    def update_item(self, index, values):
        self.data["feed_expenses"][index].update({
            "date": values["date"], "feed_type": values["type"],
            "supplier": values["supplier"], "quantity": values["qty"],
            "unit": values["unit"], "amount": values["amount"],
        })
        save_data(self.data)
        show_snackbar("Updated!")
        self.refresh()

    def delete_item(self, index, *args):
        item = self.data["feed_expenses"][index]
        def confirm():
            self.data["feed_expenses"].pop(index)
            save_data(self.data)
            show_snackbar("Deleted")
            self.refresh()
        ConfirmDialog.show("Delete?", f"Remove {item.get('feed_type', 'this')} expense?", confirm)

    def refresh(self):
        self.clear_widgets()
        self.reload_data()
        self.build_ui()


# ============================================================
# VET & HEALTH SCREEN
# ============================================================

class VetHealthScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        scroll, content = self.create_scroll_layout()
        self.add_screen_title(content, "VET & HEALTH")

        total = sum(safe_float(x.get("cost")) for x in self.data["vet_health"])
        content.add_widget(KPICard("TOTAL VET COSTS", format_currency(total), THEME["danger"]))

        content.add_widget(SearchBar(on_search=self.on_search))

        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
        self.list_container.bind(minimum_height=self.list_container.setter("height"))
        content.add_widget(self.list_container)
        self.refresh_list()

        layout.add_widget(scroll)
        self.add_fab(layout, self.show_add_dialog)
        self.add_widget(layout)

    def on_search(self, query):
        self.search_query = query
        self.refresh_list()

    def refresh_list(self):
        self.list_container.clear_widgets()
        all_items = sorted(self.data["vet_health"], key=lambda x: x.get("date", ""), reverse=True)
        items = self.filter_items(all_items, self.search_query, ["date", "cow_tag", "treatment", "vet_name"])

        if not items:
            msg = "No records found." if self.search_query else "No veterinary records."
            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
                halign="center", size_hint_y=None, height=dp(80)))
            return

        for item in items[:50]:
            idx = self.data["vet_health"].index(item)
            card = ActionCard(
                title=f"{item.get('date', '-')} | {item.get('cow_tag', 'Unknown')}",
                subtitle=f"{item.get('treatment', '-')} | {item.get('vet_name', '-')} | {format_currency(item.get('cost', '0'))}",
                on_edit=partial(self.edit_item, idx),
                on_delete=partial(self.delete_item, idx)
            )
            self.list_container.add_widget(card)

    def show_add_dialog(self, *args):
        dialog, fields = MobileDialog.create("Add Vet/Health Record", [
            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
            {"key": "cow", "hint": "Cow Tag Number", "icon": "cow", "required": True},
            {"key": "treatment", "hint": "Treatment/Diagnosis", "icon": "needle", "required": True},
            {"key": "vet", "hint": "Veterinarian Name", "icon": "doctor"},
            {"key": "cost", "hint": "Cost (KES)", "icon": "cash", "input_filter": "float", "is_number": True},
            {"key": "notes", "hint": "Notes/Medications", "icon": "note-text", "multiline": True},
        ], on_save=self.save_item)
        self.dialog = dialog
        self.dialog.open()

    def save_item(self, values):
        self.data["vet_health"].append({
            "date": values["date"], "cow_tag": values["cow"],
            "treatment": values["treatment"], "vet_name": values["vet"],
            "cost": values["cost"], "notes": values["notes"],
        })
        save_data(self.data)
        log_activity(self.data, "Vet record", f"{values['cow']} - {values['treatment']}")
        show_snackbar("Health record saved!")
        self.refresh()

    def edit_item(self, index, *args):
        item = self.data["vet_health"][index]
        dialog, fields = MobileDialog.create("Edit Health Record", [
            {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "is_date": True},
            {"key": "cow", "hint": "Cow Tag", "icon": "cow", "text": item.get("cow_tag", "")},
            {"key": "treatment", "hint": "Treatment", "icon": "needle", "text": item.get("treatment", "")},
            {"key": "vet", "hint": "Vet Name", "icon": "doctor", "text": item.get("vet_name", "")},
            {"key": "cost", "hint": "Cost", "icon": "cash", "text": str(item.get("cost", "")), "input_filter": "float", "is_number": True},
            {"key": "notes", "hint": "Notes", "icon": "note-text", "text": item.get("notes", ""), "multiline": True},
        ], on_save=lambda v: self.update_item(index, v))
        self.dialog = dialog
        self.dialog.open()

    def update_item(self, index, values):
        self.data["vet_health"][index].update({
            "date": values["date"], "cow_tag": values["cow"],
            "treatment": values["treatment"], "vet_name": values["vet"],
            "cost": values["cost"], "notes": values["notes"],
        })
        save_data(self.data)
        show_snackbar("Updated!")
        self.refresh()

    def delete_item(self, index, *args):
        item = self.data["vet_health"][index]
        def confirm():
            self.data["vet_health"].pop(index)
            save_data(self.data)
            show_snackbar("Deleted")
            self.refresh()
        ConfirmDialog.show("Delete?", f"Remove health record for {item.get('cow_tag', 'cow')}?", confirm)

    def refresh(self):
        self.clear_widgets()
        self.reload_data()
        self.build_ui()


# ============================================================
# LABOUR SCREEN
# ============================================================

class LabourScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        scroll, content = self.create_scroll_layout()
        self.add_screen_title(content, "LABOUR MANAGEMENT")

        total = sum(safe_float(x.get("amount")) for x in self.data["labour"])
        content.add_widget(KPICard("TOTAL LABOUR COSTS", format_currency(total), THEME["danger"]))

        content.add_widget(SearchBar(on_search=self.on_search))

        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
        self.list_container.bind(minimum_height=self.list_container.setter("height"))
        content.add_widget(self.list_container)
        self.refresh_list()

        layout.add_widget(scroll)
        self.add_fab(layout, self.show_add_dialog)
        self.add_widget(layout)

    def on_search(self, query):
        self.search_query = query
        self.refresh_list()

    def refresh_list(self):
        self.list_container.clear_widgets()
        all_items = sorted(self.data["labour"], key=lambda x: x.get("date", ""), reverse=True)
        items = self.filter_items(all_items, self.search_query, ["date", "worker_name", "task"])

        if not items:
            msg = "No records found." if self.search_query else "No labour records."
            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
                halign="center", size_hint_y=None, height=dp(80)))
            return

        for item in items[:50]:
            idx = self.data["labour"].index(item)
            card = ActionCard(
                title=f"{item.get('date', '-')} | {item.get('worker_name', 'Unknown')}",
                subtitle=f"{item.get('task', '-')} | {item.get('days', '-')} days @ KES {item.get('daily_rate', '0')}/day = {format_currency(item.get('amount', '0'))}",
                on_edit=partial(self.edit_item, idx),
                on_delete=partial(self.delete_item, idx)
            )
            self.list_container.add_widget(card)

    def show_add_dialog(self, *args):
        default_rate = self.settings.get("daily_rate_default", "500")
        dialog, fields = MobileDialog.create("Add Labour Record", [
            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
            {"key": "worker", "hint": "Worker Name", "icon": "account", "required": True},
            {"key": "task", "hint": "Task/Role", "icon": "hammer-wrench"},
            {"key": "days", "hint": "Days Worked", "icon": "calendar-clock", "input_filter": "float", "is_number": True},
            {"key": "rate", "hint": "Daily Rate (KES)", "icon": "cash", "input_filter": "float", "text": default_rate, "is_number": True},
        ], on_save=self.save_item)
        self.dialog = dialog
        self.dialog.open()

    def save_item(self, values):
        days = safe_float(values["days"])
        rate = safe_float(values["rate"])
        self.data["labour"].append({
            "date": values["date"], "worker_name": values["worker"],
            "task": values["task"], "days": str(days),
            "daily_rate": str(rate), "amount": str(days * rate),
        })
        save_data(self.data)
        log_activity(self.data, "Labour", f"{values['worker']} - {days} days")
        show_snackbar("Labour record saved!")
        self.refresh()

    def edit_item(self, index, *args):
        item = self.data["labour"][index]
        dialog, fields = MobileDialog.create("Edit Labour Record", [
            {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "is_date": True},
            {"key": "worker", "hint": "Worker", "icon": "account", "text": item.get("worker_name", "")},
            {"key": "task", "hint": "Task", "icon": "hammer-wrench", "text": item.get("task", "")},
            {"key": "days", "hint": "Days", "icon": "calendar-clock", "text": str(item.get("days", "")), "input_filter": "float", "is_number": True},
            {"key": "rate", "hint": "Daily Rate", "icon": "cash", "text": str(item.get("daily_rate", "")), "input_filter": "float", "is_number": True},
        ], on_save=lambda v: self.update_item(index, v))
        self.dialog = dialog
        self.dialog.open()

    def update_item(self, index, values):
        days = safe_float(values["days"])
        rate = safe_float(values["rate"])
        self.data["labour"][index].update({
            "date": values["date"], "worker_name": values["worker"],
            "task": values["task"], "days": str(days),
            "daily_rate": str(rate), "amount": str(days * rate),
        })
        save_data(self.data)
        show_snackbar("Updated!")
        self.refresh()

    def delete_item(self, index, *args):
        item = self.data["labour"][index]
        def confirm():
            self.data["labour"].pop(index)
            save_data(self.data)
            show_snackbar("Deleted")
            self.refresh()
        ConfirmDialog.show("Delete?", f"Remove record for {item.get('worker_name', 'worker')}?", confirm)

    def refresh(self):
        self.clear_widgets()
        self.reload_data()
        self.build_ui()


# ============================================================
# OPERATIONS SCREEN
# ============================================================

class OperationsScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        scroll, content = self.create_scroll_layout()
        self.add_screen_title(content, "OPERATIONS")

        total = sum(safe_float(x.get("amount")) for x in self.data["operations"])
        content.add_widget(KPICard("TOTAL OPERATIONS", format_currency(total), THEME["warning"]))

        content.add_widget(SearchBar(on_search=self.on_search))

        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(6))
        self.list_container.bind(minimum_height=self.list_container.setter("height"))
        content.add_widget(self.list_container)
        self.refresh_list()

        layout.add_widget(scroll)
        self.add_fab(layout, self.show_add_dialog)
        self.add_widget(layout)

    def on_search(self, query):
        self.search_query = query
        self.refresh_list()

    def refresh_list(self):
        self.list_container.clear_widgets()
        all_items = sorted(self.data["operations"], key=lambda x: x.get("date", ""), reverse=True)
        items = self.filter_items(all_items, self.search_query, ["date", "category", "description"])

        if not items:
            msg = "No records found." if self.search_query else "No operation records."
            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
                halign="center", size_hint_y=None, height=dp(80)))
            return

        for item in items[:50]:
            idx = self.data["operations"].index(item)
            card = ActionCard(
                title=f"{item.get('date', '-')} | {item.get('category', 'General')}",
                subtitle=f"{item.get('description', '-')} | {format_currency(item.get('amount', '0'))}",
                on_edit=partial(self.edit_item, idx),
                on_delete=partial(self.delete_item, idx)
            )
            self.list_container.add_widget(card)

    def show_add_dialog(self, *args):
        dialog, fields = MobileDialog.create("Add Operation Expense", [
            {"key": "date", "hint": "Date (YYYY-MM-DD)", "icon": "calendar", "text": datetime.now().strftime("%Y-%m-%d"), "required": True, "is_date": True},
            {"key": "category", "hint": "Category (Fuel/Repair/Utilities/Equipment)", "icon": "folder", "required": True},
            {"key": "desc", "hint": "Description", "icon": "text"},
            {"key": "amount", "hint": "Amount (KES)", "icon": "cash", "input_filter": "float", "required": True, "is_number": True},
        ], on_save=self.save_item)
        self.dialog = dialog
        self.dialog.open()

    def save_item(self, values):
        self.data["operations"].append({
            "date": values["date"], "category": values["category"],
            "description": values["desc"], "amount": values["amount"],
        })
        save_data(self.data)
        log_activity(self.data, "Operation", f"{values['category']} - {values['amount']}")
        show_snackbar("Operation saved!")
        self.refresh()

    def edit_item(self, index, *args):
        item = self.data["operations"][index]
        dialog, fields = MobileDialog.create("Edit Operation", [
            {"key": "date", "hint": "Date", "icon": "calendar", "text": item.get("date", ""), "is_date": True},
            {"key": "category", "hint": "Category", "icon": "folder", "text": item.get("category", "")},
            {"key": "desc", "hint": "Description", "icon": "text", "text": item.get("description", "")},
            {"key": "amount", "hint": "Amount", "icon": "cash", "text": str(item.get("amount", "")), "input_filter": "float", "is_number": True},
        ], on_save=lambda v: self.update_item(index, v))
        self.dialog = dialog
        self.dialog.open()

    def update_item(self, index, values):
        self.data["operations"][index].update({
            "date": values["date"], "category": values["category"],
            "description": values["desc"], "amount": values["amount"],
        })
        save_data(self.data)
        show_snackbar("Updated!")
        self.refresh()

    def delete_item(self, index, *args):
        item = self.data["operations"][index]
        def confirm():
            self.data["operations"].pop(index)
            save_data(self.data)
            show_snackbar("Deleted")
            self.refresh()
        ConfirmDialog.show("Delete?", f"Remove {item.get('category', 'this')} operation?", confirm)

    def refresh(self):
        self.clear_widgets()
        self.reload_data()
        self.build_ui()



# ============================================================
# CHATS / NOTES SCREEN
# ============================================================

class ChatsScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        scroll, content = self.create_scroll_layout()
        self.add_screen_title(content, "FARM NOTES & CHATS")

        total_notes = len(self.data.get("chats", []))
        content.add_widget(KPICard("TOTAL NOTES", str(total_notes), THEME["info"]))

        content.add_widget(SearchBar(on_search=self.on_search))

        self.list_container = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(8))
        self.list_container.bind(minimum_height=self.list_container.setter("height"))
        content.add_widget(self.list_container)
        self.refresh_list()

        layout.add_widget(scroll)
        self.add_fab(layout, self.show_add_dialog, icon="message-plus")
        self.add_widget(layout)

    def on_search(self, query):
        self.search_query = query
        self.refresh_list()

    def refresh_list(self):
        self.list_container.clear_widgets()
        all_items = sorted(self.data.get("chats", []), key=lambda x: x.get("timestamp", ""), reverse=True)
        items = self.filter_items(all_items, self.search_query, ["author", "message", "category"])

        if not items:
            msg = "No notes found." if self.search_query else "No notes yet.\nTap + to add a farm note or chat."
            self.list_container.add_widget(MDLabel(text=msg, theme_text_color="Hint",
                halign="center", size_hint_y=None, height=dp(100)))
            return

        for item in items:
            idx = self.data["chats"].index(item)
            card = ThemedCard(orientation="vertical", size_hint_y=None, height=dp(120), padding=dp(12))

            header = MDBoxLayout(size_hint_y=None, height=dp(24))
            header.add_widget(MDLabel(
                text=item.get("author", "Unknown"),
                theme_text_color="Custom", text_color=THEME["primary"],
                font_style="Subtitle1", bold=True, size_hint_x=0.6
            ))
            header.add_widget(MDLabel(
                text=item.get("timestamp", "")[:16],
                theme_text_color="Custom", text_color=THEME["text_secondary"],
                font_style="Caption", halign="right", size_hint_x=0.4
            ))
            card.add_widget(header)

            if item.get("category"):
                card.add_widget(MDLabel(
                    text=f"Category: {item['category']}",
                    theme_text_color="Custom", text_color=THEME["accent"],
                    font_style="Caption", size_hint_y=None, height=dp(18)
                ))

            card.add_widget(MDLabel(
                text=item.get("message", ""),
                theme_text_color="Custom", text_color=THEME["text"],
                font_style="Body1", size_hint_y=None, height=dp(50)
            ))

            btn_row = MDBoxLayout(size_hint_y=None, height=dp(32), spacing=dp(4))
            btn_row.add_widget(Widget(size_hint_x=0.6))
            btn_row.add_widget(MDIconButton(
                icon="pencil", theme_text_color="Custom", text_color=THEME["info"],
                on_release=partial(self.edit_note, idx), icon_size=dp(18)
            ))
            btn_row.add_widget(MDIconButton(
                icon="delete", theme_text_color="Custom", text_color=THEME["danger"],
                on_release=partial(self.delete_note, idx), icon_size=dp(18)
            ))
            card.add_widget(btn_row)

            self.list_container.add_widget(card)

    def show_add_dialog(self, *args):
        dialog, fields = MobileDialog.create("Add Farm Note", [
            {"key": "author", "hint": "Your Name", "icon": "account", "required": True, "text": self.settings.get("owner_name", "")},
            {"key": "category", "hint": "Category (General/Task/Reminder)", "icon": "tag", "text": "General"},
            {"key": "message", "hint": "Message / Note", "icon": "message-text", "required": True, "multiline": True},
        ], on_save=self.save_note)
        self.dialog = dialog
        self.dialog.open()

    def save_note(self, values):
        self.data.setdefault("chats", []).insert(0, {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "author": values["author"],
            "category": values["category"],
            "message": values["message"],
        })
        save_data(self.data)
        log_activity(self.data, "Note added", values["author"])
        show_snackbar("Note saved!")
        self.refresh()

    def edit_note(self, index, *args):
        item = self.data["chats"][index]
        dialog, fields = MobileDialog.create("Edit Note", [
            {"key": "author", "hint": "Author", "icon": "account", "text": item.get("author", "")},
            {"key": "category", "hint": "Category", "icon": "tag", "text": item.get("category", "")},
            {"key": "message", "hint": "Message", "icon": "message-text", "text": item.get("message", ""), "multiline": True},
        ], on_save=lambda v: self.update_note(index, v))
        self.dialog = dialog
        self.dialog.open()

    def update_note(self, index, values):
        self.data["chats"][index].update({
            "author": values["author"],
            "category": values["category"],
            "message": values["message"],
        })
        save_data(self.data)
        show_snackbar("Note updated!")
        self.refresh()

    def delete_note(self, index, *args):
        def confirm():
            self.data["chats"].pop(index)
            save_data(self.data)
            show_snackbar("Note deleted")
            self.refresh()
        ConfirmDialog.show("Delete Note?", "Remove this note permanently?", confirm)

    def refresh(self):
        self.clear_widgets()
        self.reload_data()
        self.build_ui()


# ============================================================
# GRAPHS / ANALYTICS SCREEN
# ============================================================

class GraphsScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        scroll, content = self.create_scroll_layout()
        self.add_screen_title(content, "GRAPHS & ANALYTICS")

        income = self._calc_income()
        expenses = self._calc_expenses()
        net = income - expenses

        content.add_widget(KPICard("NET P&L", format_currency(net), THEME["success"] if net >= 0 else THEME["danger"]))

        # Monthly Income vs Expense text bars
        content.add_widget(MDLabel(text="MONTHLY INCOME vs EXPENSE (KES)", theme_text_color="Custom",
            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))

        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        max_val = 1
        monthly_data = []
        for i in range(12):
            m_income = (sum(safe_float(x.get("amount")) for x in self.data["milk_sales"] if _get_month(x.get("date", "")) == i) +
                       sum(safe_float(x.get("amount")) for x in self.data["other_income"] if _get_month(x.get("date", "")) == i))
            m_exp = (sum(safe_float(x.get("amount")) for x in self.data["feed_expenses"] if _get_month(x.get("date", "")) == i) +
                    sum(safe_float(x.get("cost")) for x in self.data["vet_health"] if _get_month(x.get("date", "")) == i) +
                    sum(safe_float(x.get("amount")) for x in self.data["labour"] if _get_month(x.get("date", "")) == i) +
                    sum(safe_float(x.get("amount")) for x in self.data["operations"] if _get_month(x.get("date", "")) == i))
            monthly_data.append((months[i], m_income, m_exp))
            max_val = max(max_val, m_income, m_exp)

        for month, inc, exp in monthly_data:
            row = MDBoxLayout(size_hint_y=None, height=dp(50), spacing=dp(4))

            label_col = MDBoxLayout(orientation="vertical", size_hint_x=0.15)
            label_col.add_widget(MDLabel(text=month, theme_text_color="Custom",
                text_color=THEME["text"], font_style="Caption", halign="center"))
            row.add_widget(label_col)

            bars_col = MDBoxLayout(orientation="vertical", size_hint_x=0.85, spacing=dp(2))

            # Income bar
            inc_pct = (inc / max_val * 100) if max_val > 0 else 0
            inc_bar = MDBoxLayout(size_hint_y=None, height=dp(16))
            inc_bar.add_widget(MDLabel(text=f"Inc: {format_currency(inc)}", theme_text_color="Custom",
                text_color=THEME["success"], font_style="Caption", size_hint_x=0.4))
            inc_fill = MDBoxLayout(size_hint_x=0.6)
            with inc_fill.canvas:
                Color(*THEME["success"])
                Rectangle(pos=(0, 0), size=(1, 1))  # Will be sized properly
            # Use a simpler approach - just show the percentage as text
            inc_bar.add_widget(MDLabel(text=f"{inc_pct:.0f}%", theme_text_color="Custom",
                text_color=THEME["success"], font_style="Caption", halign="right", size_hint_x=0.2))
            bars_col.add_widget(inc_bar)

            # Expense bar
            exp_pct = (exp / max_val * 100) if max_val > 0 else 0
            exp_bar = MDBoxLayout(size_hint_y=None, height=dp(16))
            exp_bar.add_widget(MDLabel(text=f"Exp: {format_currency(exp)}", theme_text_color="Custom",
                text_color=THEME["danger"], font_style="Caption", size_hint_x=0.4))
            exp_bar.add_widget(MDLabel(text=f"{exp_pct:.0f}%", theme_text_color="Custom",
                text_color=THEME["danger"], font_style="Caption", halign="right", size_hint_x=0.2))
            bars_col.add_widget(exp_bar)

            row.add_widget(bars_col)
            content.add_widget(row)

        # Expense breakdown
        content.add_widget(MDLabel(text="EXPENSE BREAKDOWN", theme_text_color="Custom",
            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))

        feed = sum(safe_float(x.get("amount")) for x in self.data["feed_expenses"])
        vet = sum(safe_float(x.get("cost")) for x in self.data["vet_health"])
        lab = sum(safe_float(x.get("amount")) for x in self.data["labour"])
        ops = sum(safe_float(x.get("amount")) for x in self.data["operations"])
        total_exp = feed + vet + lab + ops

        breakdown = [("Feed", feed, THEME["warning"]), ("Vet & Health", vet, THEME["danger"]),
                     ("Labour", lab, THEME["info"]), ("Operations", ops, THEME["text_secondary"])]
        for name, amount, color in breakdown:
            pct = (amount / total_exp * 100) if total_exp > 0 else 0
            content.add_widget(MDLabel(
                text=f"{name}: {format_currency(amount)} ({pct:.1f}%)",
                theme_text_color="Custom", text_color=color,
                font_style="Body1", size_hint_y=None, height=dp(26)
            ))

        # Milk production summary
        content.add_widget(MDLabel(text="MILK PRODUCTION SUMMARY", theme_text_color="Custom",
            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))

        total_milk = sum(safe_float(x.get("quantity")) for x in self.data["milk_production"])
        avg_daily = total_milk / 30 if total_milk > 0 else 0  # rough estimate
        content.add_widget(MDLabel(text=f"Total Recorded: {total_milk:.1f} L", theme_text_color="Custom",
            text_color=THEME["text"], font_style="Body1", size_hint_y=None, height=dp(24)))
        content.add_widget(MDLabel(text=f"Records: {len(self.data['milk_production'])}", theme_text_color="Custom",
            text_color=THEME["text_secondary"], font_style="Body1", size_hint_y=None, height=dp(24)))

        layout.add_widget(scroll)
        self.add_widget(layout)

    def _calc_income(self):
        return (sum(safe_float(x.get("amount")) for x in self.data.get("milk_sales", [])) +
                sum(safe_float(x.get("amount")) for x in self.data.get("other_income", [])))

    def _calc_expenses(self):
        return (sum(safe_float(x.get("amount")) for x in self.data.get("feed_expenses", [])) +
                sum(safe_float(x.get("cost")) for x in self.data.get("vet_health", [])) +
                sum(safe_float(x.get("amount")) for x in self.data.get("labour", [])) +
                sum(safe_float(x.get("amount")) for x in self.data.get("operations", [])))

    def refresh(self):
        self.clear_widgets()
        self.reload_data()
        self.build_ui()

# ============================================================
# REPORTS SCREEN
# ============================================================

class ReportsScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        scroll, content = self.create_scroll_layout()
        self.add_screen_title(content, "REPORTS & ANALYTICS")

        income = self._calc_income()
        expenses = self._calc_expenses()
        net = income - expenses

        content.add_widget(KPICard("NET PROFIT/LOSS", format_currency(net), THEME["success"] if net >= 0 else THEME["danger"]))

        # Expense breakdown pie chart
        feed = sum(safe_float(x.get("amount")) for x in self.data["feed_expenses"])
        vet = sum(safe_float(x.get("cost")) for x in self.data["vet_health"])
        lab = sum(safe_float(x.get("amount")) for x in self.data["labour"])
        ops = sum(safe_float(x.get("amount")) for x in self.data["operations"])

        if CHARTS_AVAILABLE and expenses > 0:
            labels = ["Feed", "Vet", "Labour", "Ops"]
            values = [feed, vet, lab, ops]
            colors = ['#f39c12', '#e74c3c', '#3498db', '#7f8c8d']
            chart = create_pie_chart(labels, values, "Expense Breakdown", colors)
            if chart:
                try:
                    chart.size_hint_y = None
                    chart.height = dp(240)
                    content.add_widget(chart)
                except Exception as e:
                    print(f"Chart display error: {e}")

        content.add_widget(MDLabel(text="EXPENSE BREAKDOWN", theme_text_color="Custom",
            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))

        breakdown = [("Feed", feed, THEME["warning"]), ("Vet & Health", vet, THEME["danger"]),
                     ("Labour", lab, THEME["info"]), ("Operations", ops, THEME["text_secondary"])]
        for name, amount, color in breakdown:
            pct = (amount / expenses * 100) if expenses > 0 else 0
            content.add_widget(MDLabel(
                text=f"{name}: {format_currency(amount)} ({pct:.1f}%)",
                theme_text_color="Custom", text_color=color,
                font_style="Body1", size_hint_y=None, height=dp(26)
            ))

        # Monthly income vs expense chart
        if CHARTS_AVAILABLE:
            months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            inc_vals = []
            exp_vals = []
            for i in range(12):
                m_income = sum(safe_float(x.get("amount")) for x in self.data["milk_sales"] if _get_month(x.get("date", "")) == i)
                m_income += sum(safe_float(x.get("amount")) for x in self.data["other_income"] if _get_month(x.get("date", "")) == i)
                m_exp = (sum(safe_float(x.get("amount")) for x in self.data["feed_expenses"] if _get_month(x.get("date", "")) == i) +
                         sum(safe_float(x.get("cost")) for x in self.data["vet_health"] if _get_month(x.get("date", "")) == i) +
                         sum(safe_float(x.get("amount")) for x in self.data["labour"] if _get_month(x.get("date", "")) == i) +
                         sum(safe_float(x.get("amount")) for x in self.data["operations"] if _get_month(x.get("date", "")) == i))
                inc_vals.append(m_income)
                exp_vals.append(m_exp)

            try:
                fig, ax = plt.subplots(figsize=(3.2, 2.4), dpi=100)
                fig.patch.set_alpha(0)
                ax.set_facecolor('none')
                x = range(12)
                ax.bar([i - 0.2 for i in x], inc_vals, 0.4, label='Income', color='#27ae60', alpha=0.8)
                ax.bar([i + 0.2 for i in x], exp_vals, 0.4, label='Expense', color='#e74c3c', alpha=0.8)
                ax.set_xticks(x)
                ax.set_xticklabels(months, fontsize=6, color='white')
                ax.tick_params(colors='white', labelsize=7)
                ax.set_title("Monthly Income vs Expense", color='white', fontsize=10)
                ax.legend(facecolor='none', edgecolor='white', labelcolor='white', fontsize=7)
                ax.spines['bottom'].set_color('white')
                ax.spines['left'].set_color('white')
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
                ax.grid(True, alpha=0.3, color='white', axis='y')
                plt.tight_layout()
                chart = FigureCanvasKivyAgg(fig)
                chart.size_hint_y = None
                chart.height = dp(200)
                content.add_widget(chart)
            except Exception as e:
                print(f"Chart display error: {e}")

        # Herd Summary
        content.add_widget(MDLabel(text="HERD SUMMARY", theme_text_color="Custom",
            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))

        breeds = defaultdict(int)
        for cow in self.data["herd"]:
            breeds[cow.get("breed", "Unknown")] += 1
        if not breeds:
            content.add_widget(MDLabel(text="No herd data.", theme_text_color="Hint",
                halign="center", size_hint_y=None, height=dp(40)))
        else:
            for breed, count in breeds.items():
                content.add_widget(MDLabel(
                    text=f"{breed}: {count} cow(s)",
                    theme_text_color="Custom", text_color=THEME["text"],
                    font_style="Body1", size_hint_y=None, height=dp(24)
                ))

        layout.add_widget(scroll)
        self.add_widget(layout)

    def _calc_income(self):
        return (sum(safe_float(x.get("amount")) for x in self.data.get("milk_sales", [])) +
                sum(safe_float(x.get("amount")) for x in self.data.get("other_income", [])))

    def _calc_expenses(self):
        return (sum(safe_float(x.get("amount")) for x in self.data.get("feed_expenses", [])) +
                sum(safe_float(x.get("cost")) for x in self.data.get("vet_health", [])) +
                sum(safe_float(x.get("amount")) for x in self.data.get("labour", [])) +
                sum(safe_float(x.get("amount")) for x in self.data.get("operations", [])))

    def refresh(self):
        self.clear_widgets()
        self.reload_data()
        self.build_ui()


# ============================================================
# SETTINGS SCREEN
# ============================================================

class SettingsScreen(BaseScreen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.build_ui()

    def build_ui(self):
        layout = MDBoxLayout(orientation="vertical")
        scroll, content = self.create_scroll_layout()
        self.add_screen_title(content, "SETTINGS & TOOLS")

        # Farm Profile
        content.add_widget(MDLabel(text="FARM PROFILE", theme_text_color="Custom",
            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))

        profile_card = ThemedCard(orientation="vertical", size_hint_y=None, height=dp(200), spacing=dp(8))
        self.farm_name_field = MDTextField(hint_text="Farm Name", text=self.settings.get("farm_name", ""), mode="rectangle")
        self.farm_loc_field = MDTextField(hint_text="Location", text=self.settings.get("farm_location", ""), mode="rectangle")
        self.owner_field = MDTextField(hint_text="Owner Name", text=self.settings.get("owner_name", ""), mode="rectangle")
        profile_card.add_widget(self.farm_name_field)
        profile_card.add_widget(self.farm_loc_field)
        profile_card.add_widget(self.owner_field)
        profile_card.add_widget(MDRaisedButton(text="SAVE PROFILE", md_bg_color=THEME["success"],
            on_release=self.save_profile, size_hint=(1, None), height=dp(40)))
        content.add_widget(profile_card)

        # Defaults
        content.add_widget(MDLabel(text="DEFAULTS", theme_text_color="Custom",
            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))

        defaults_card = ThemedCard(orientation="vertical", size_hint_y=None, height=dp(160), spacing=dp(8))
        self.price_field = MDTextField(hint_text="Default Milk Price/Liter (KES)",
            text=self.settings.get("milk_price_default", "60"), mode="rectangle", input_filter="float")
        self.rate_field = MDTextField(hint_text="Default Daily Labour Rate (KES)",
            text=self.settings.get("daily_rate_default", "500"), mode="rectangle", input_filter="float")
        defaults_card.add_widget(self.price_field)
        defaults_card.add_widget(self.rate_field)
        defaults_card.add_widget(MDRaisedButton(text="SAVE DEFAULTS", md_bg_color=THEME["success"],
            on_release=self.save_defaults, size_hint=(1, None), height=dp(40)))
        content.add_widget(defaults_card)

        # Data Management
        content.add_widget(MDLabel(text="DATA MANAGEMENT", theme_text_color="Custom",
            text_color=THEME["primary"], font_style="H6", size_hint_y=None, height=dp(36), bold=True))

        data_card = ThemedCard(orientation="vertical", size_hint_y=None, height=dp(220), spacing=dp(8))
        data_card.add_widget(MDRaisedButton(text="EXPORT TO EXCEL", md_bg_color=THEME["info"],
            on_release=self.export_excel, size_hint=(1, None), height=dp(44)))
        data_card.add_widget(MDRaisedButton(text="VIEW BACKUPS", md_bg_color=THEME["accent"],
            on_release=self.view_backups, size_hint=(1, None), height=dp(44)))
        data_card.add_widget(MDRaisedButton(text="VIEW BACKUPS", md_bg_color=THEME["accent"],
            on_release=self.view_backups, size_hint=(1, None), height=dp(44)))
        data_card.add_widget(MDRaisedButton(text="CLEAR ALL DATA", md_bg_color=THEME["danger"],
            on_release=self.clear_all_data, size_hint=(1, None), height=dp(44)))
        content.add_widget(data_card)

        # App Info
        content.add_widget(MDLabel(text="DAIRY FARM MS PRO v2.0", theme_text_color="Custom",
            text_color=THEME["text_secondary"], font_style="Caption",
            halign="center", size_hint_y=None, height=dp(30)))

        layout.add_widget(scroll)
        self.add_widget(layout)

    def save_profile(self, *args):
        self.settings["farm_name"] = self.farm_name_field.text
        self.settings["farm_location"] = self.farm_loc_field.text
        self.settings["owner_name"] = self.owner_field.text
        save_settings(self.settings)
        show_snackbar("Profile saved!")

    def save_defaults(self, *args):
        self.settings["milk_price_default"] = self.price_field.text or "60"
        self.settings["daily_rate_default"] = self.rate_field.text or "500"
        save_settings(self.settings)
        show_snackbar("Defaults saved!")

    def export_excel(self, *args):
        if not EXCEL_AVAILABLE:
            show_snackbar("Install openpyxl: pip install openpyxl", THEME["danger"], 3)
            return
        app = MDApp.get_running_app()
        app.do_export("excel")

    def view_backups(self, *args):
        backups = sorted([f for f in os.listdir(get_backup_dir()) if f.startswith("backup_")], reverse=True)
        content = MDBoxLayout(orientation="vertical", size_hint_y=None, spacing=dp(4))
        content.height = dp(20) + len(backups) * dp(36)

        if not backups:
            content.add_widget(MDLabel(text="No backups found.", theme_text_color="Hint",
                halign="center", size_hint_y=None, height=dp(40)))
        else:
            for b in backups[:20]:
                content.add_widget(MDLabel(text=b, theme_text_color="Custom",
                    text_color=THEME["text"], font_style="Caption", size_hint_y=None, height=dp(32)))

        dialog = MDDialog(title="Backups", type="custom", content_cls=content,
            buttons=[MDFlatButton(text="CLOSE", on_release=lambda x: dialog.dismiss())])
        dialog.open()

    def clear_all_data(self, *args):
        def confirm():
            self.data = {
                "farm": {"name": "My Dairy Farm", "location": "", "owner": ""},
                "herd": [], "milk_production": [], "milk_sales": [],
                "other_income": [], "feed_expenses": [], "vet_health": [],
                "labour": [], "operations": [], "assets": [], "breeding": [], "activity_log": [],
            }
            save_data(self.data)
            show_snackbar("All data cleared!")
            self.refresh()
        ConfirmDialog.show("CLEAR ALL DATA?", "This will delete EVERYTHING. This cannot be undone!", confirm, "CLEAR", THEME["danger"])

    def refresh(self):
        self.clear_widgets()
        self.reload_data()
        self.build_ui()


# ============================================================
# MAIN APP
# ============================================================

class DairyFarmApp(MDApp):
    def build(self):
        self.theme_cls.primary_palette = "Green"
        self.theme_cls.theme_style = "Light"

        # Load settings for theme
        self.settings = load_settings()
        if self.settings.get("dark_mode"):
            self.theme_cls.theme_style = "Dark"
            global THEME
            THEME = DARK_THEME

        sm = MDScreenManager(transition=SlideTransition(duration=0.2))
        sm.add_widget(DashboardScreen(name="dashboard"))
        sm.add_widget(HerdScreen(name="herd"))
        sm.add_widget(MilkProductionScreen(name="milk_production"))
        sm.add_widget(MilkSalesScreen(name="milk_sales"))
        sm.add_widget(FeedExpensesScreen(name="feed_expenses"))
        sm.add_widget(VetHealthScreen(name="vet_health"))
        sm.add_widget(LabourScreen(name="labour"))
        sm.add_widget(OperationsScreen(name="operations"))
        sm.add_widget(ReportsScreen(name="reports"))
        sm.add_widget(GraphsScreen(name="graphs"))
        sm.add_widget(ChatsScreen(name="chats"))
        sm.add_widget(SettingsScreen(name="settings"))

        self.sm = sm

        # Main layout with bottom navigation
        layout = MDBoxLayout(orientation="vertical")

        # Toolbar
        self.toolbar = MDTopAppBar(
            title="Dairy Farm MS",
            elevation=4,
            pos_hint={"top": 1},
            md_bg_color=THEME["primary"],
            left_action_items=[["menu", lambda x: self.nav_drawer.set_state("open")]],
            right_action_items=[
                ["chart-bar", lambda x: self.switch_screen("graphs")],
                ["export-variant", lambda x: self.show_export_menu()],
            ],
        )
        layout.add_widget(self.toolbar)
        layout.add_widget(sm)

        # Custom Bottom Navigation Bar (reliable across KivyMD versions)
        self.bottom_bar = MDBoxLayout(
            orientation="horizontal",
            size_hint_y=None,
            height=dp(56),
            md_bg_color=THEME["card"],
            padding=[dp(4), dp(4), dp(4), dp(4)],
            spacing=dp(2),
        )

        # Add shadow line above bottom bar
        with self.bottom_bar.canvas.before:
            Color(*THEME["divider"])
            Rectangle(pos=(0, dp(56)), size=(Window.width, dp(1)))

        self.nav_buttons = {}
        nav_items = [
            ("dashboard", "Dashboard", "chart-bar"),
            ("herd", "Herd", "cow"),
            ("milk_production", "Milk", "cup-water"),
            ("milk_sales", "Sales", "cash-multiple"),
            ("reports", "Reports", "file-chart"),
        ]

        for screen_name, label, icon in nav_items:
            btn = MDBoxLayout(orientation="vertical", size_hint_x=1)
            btn_icon = MDIconButton(
                icon=icon,
                theme_text_color="Custom",
                text_color=THEME["text_secondary"],
                pos_hint={"center_x": 0.5},
                icon_size=dp(22),
            )
            btn_label = MDLabel(
                text=label,
                theme_text_color="Custom",
                text_color=THEME["text_secondary"],
                font_style="Caption",
                halign="center",
                size_hint_y=None,
                height=dp(16),
            )
            btn.add_widget(btn_icon)
            btn.add_widget(btn_label)

            # Make the whole box layout clickable
            btn.bind(on_touch_down=partial(self.on_bottom_button_touch, screen_name, btn_icon, btn_label))

            self.nav_buttons[screen_name] = (btn, btn_icon, btn_label)
            self.bottom_bar.add_widget(btn)

        layout.add_widget(self.bottom_bar)
        self.bottom_nav = self.bottom_bar  # alias for compatibility

        # Side Navigation Drawer (for less-used screens)
        self.nav_drawer = MDNavigationDrawer(
            md_bg_color=THEME["bg"],
            elevation=4,
        )

        nav_content = MDBoxLayout(orientation="vertical", padding=dp(8), spacing=dp(4))
        nav_content.add_widget(MDLabel(
            text="MENU",
            theme_text_color="Custom",
            text_color=THEME["primary"],
            font_style="H6",
            size_hint_y=None,
            height=dp(40),
            halign="center"
        ))

        drawer_items = [
            ("feed_expenses", "Feed Expenses", "corn"),
            ("vet_health", "Vet & Health", "needle"),
            ("labour", "Labour", "account-hard-hat"),
            ("operations", "Operations", "cogs"),
            ("chats", "Farm Notes", "message-text"),
            ("settings", "Settings & Tools", "cog"),
        ]

        nav_list = MDList()
        for screen_name, label, icon in drawer_items:
            item = OneLineIconListItem(
                IconLeftWidget(icon=icon),
                text=label,
                on_release=lambda x, s=screen_name: self.switch_screen(s)
            )
            nav_list.add_widget(item)

        nav_content.add_widget(nav_list)
        self.nav_drawer.add_widget(nav_content)

        root = MDScreen()
        root.add_widget(layout)
        root.add_widget(self.nav_drawer)

        return root

    def show_export_menu(self, *args):
        menu_items = [
            {"text": "Export to Excel", "icon": "microsoft-excel", "on_release": lambda: self.do_export("excel")},
        ]
        self.export_menu = MDDropdownMenu(
            caller=self.toolbar,
            items=menu_items,
            width_mult=3,
        )
        self.export_menu.open()

    def do_export(self, fmt):
        self.export_menu.dismiss()
        if not EXCEL_AVAILABLE:
            show_snackbar("Install openpyxl: pip install openpyxl", THEME["danger"], 3)
            return

        data = load_data()
        settings = load_settings()
        export_dir = os.path.join(get_data_dir(), "exports")
        os.makedirs(export_dir, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = os.path.join(export_dir, f"DairyFarm_Report_{timestamp}.xlsx")

        try:
            wb = Workbook()

            # ===== STYLES =====
            header_font = Font(bold=True, color="1a5c3a", size=11)
            header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

            section_font = Font(bold=True, color="FFFFFF", size=12)
            section_fill = PatternFill(start_color="2d8a5e", end_color="2d8a5e", fill_type="solid")
            section_align = Alignment(horizontal="left", vertical="center")

            title_font = Font(bold=True, color="1a5c3a", size=16)
            title_align = Alignment(horizontal="center", vertical="center")

            data_font = Font(color="2c3e50", size=10)
            data_align = Alignment(horizontal="left", vertical="center")
            number_align = Alignment(horizontal="right", vertical="center")

            thin_border = Border(
                left=Side(style='thin', color="B0B0B0"),
                right=Side(style='thin', color="B0B0B0"),
                top=Side(style='thin', color="B0B0B0"),
                bottom=Side(style='thin', color="B0B0B0")
            )

            total_font = Font(bold=True, color="1a5c3a", size=10)
            total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

            def style_header_row(ws, row_num, num_cols):
                for col in range(1, num_cols + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = thin_border

            def style_data_row(ws, row_num, num_cols, is_number_row=False):
                for col in range(1, num_cols + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.border = thin_border
                    cell.font = data_font
                    cell.alignment = number_align if (is_number_row and col > 1) else data_align

            def style_total_row(ws, row_num, num_cols):
                for col in range(1, num_cols + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.font = total_font
                    cell.fill = total_fill
                    cell.border = thin_border
                    cell.alignment = number_align if col > 1 else data_align

            def auto_width(ws, min_width=10, max_width=45):
                from openpyxl.utils import get_column_letter
                for col_idx in range(1, ws.max_column + 1):
                    max_length = 0
                    col_letter = get_column_letter(col_idx)
                    for row in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = min(max(min_width, max_length + 3), max_width)

            def add_section_title(ws, title, num_cols):
                row = ws.max_row + 1
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
                cell = ws.cell(row=row, column=1, value=title)
                cell.font = section_font
                cell.fill = section_fill
                cell.alignment = section_align
                for c in range(1, num_cols + 1):
                    ws.cell(row=row, column=c).border = thin_border
                return row

            # ===== CALCULATIONS =====
            total_income = (sum(safe_float(x.get("amount")) for x in data.get("milk_sales", [])) +
                           sum(safe_float(x.get("amount")) for x in data.get("other_income", [])))
            total_feed = sum(safe_float(x.get("amount")) for x in data.get("feed_expenses", []))
            total_vet = sum(safe_float(x.get("cost")) for x in data.get("vet_health", []))
            total_labour = sum(safe_float(x.get("amount")) for x in data.get("labour", []))
            total_ops = sum(safe_float(x.get("amount")) for x in data.get("operations", []))
            total_exp = total_feed + total_vet + total_labour + total_ops
            net = total_income - total_exp

            herd = data.get("herd", [])
            milking = len([c for c in herd if c.get("status", "").lower() == "milking"])
            dry = len([c for c in herd if c.get("status", "").lower() == "dry"])
            heifer = len([c for c in herd if c.get("status", "").lower() == "heifer"])
            total_milk = sum(safe_float(x.get("quantity")) for x in data.get("milk_production", []))
            total_liters_sold = sum(safe_float(x.get("liters")) for x in data.get("milk_sales", []))
            avg_price = (sum(safe_float(x.get("price_per_liter")) for x in data.get("milk_sales", [])) /
                        len(data.get("milk_sales", []))) if data.get("milk_sales") else 0

            # ===== 1. COVER / SUMMARY SHEET =====
            ws = wb.active
            ws.title = "Summary"

            ws.merge_cells('A1:D1')
            ws['A1'] = settings.get("farm_name", "DAIRY FARM MANAGEMENT REPORT")
            ws['A1'].font = title_font
            ws['A1'].alignment = title_align
            ws.row_dimensions[1].height = 30

            ws.merge_cells('A2:D2')
            ws['A2'] = f"Generated: {datetime.now().strftime('%A, %d %B %Y at %I:%M %p')}"
            ws['A2'].font = Font(italic=True, color="7f8c8d", size=10)
            ws['A2'].alignment = Alignment(horizontal="center")

            ws.merge_cells('A3:D3')
            ws['A3'] = f"Owner: {settings.get('owner_name', 'N/A')}  |  Location: {settings.get('farm_location', 'N/A')}"
            ws['A3'].font = Font(italic=True, color="7f8c8d", size=10)
            ws['A3'].alignment = Alignment(horizontal="center")
            ws.append([""])

            # Financial Summary
            add_section_title(ws, "FINANCIAL SUMMARY", 4)
            ws.append(["Description", "Amount (KES)", "% of Total", "Status"])
            style_header_row(ws, ws.max_row, 4)
            ws.append(["Total Income", total_income, "", ""])
            style_data_row(ws, ws.max_row, 4, True)
            ws.append(["Total Expenses", total_exp, f"{(total_exp/total_income*100):.1f}%" if total_income > 0 else "0%", ""])
            style_data_row(ws, ws.max_row, 4, True)
            ws.append(["NET PROFIT / LOSS", net, "", "PROFIT" if net >= 0 else "LOSS"])
            style_total_row(ws, ws.max_row, 4)
            ws.append([""])

            # Expense Breakdown
            add_section_title(ws, "EXPENSE BREAKDOWN", 4)
            ws.append(["Category", "Amount (KES)", "% of Expenses", "Records"])
            style_header_row(ws, ws.max_row, 4)
            exp_items = [
                ("Feed & Nutrition", total_feed, len(data.get("feed_expenses", []))),
                ("Vet & Health", total_vet, len(data.get("vet_health", []))),
                ("Labour & Wages", total_labour, len(data.get("labour", []))),
                ("Operations & Misc", total_ops, len(data.get("operations", []))),
            ]
            for name, amt, count in exp_items:
                pct = f"{(amt/total_exp*100):.1f}%" if total_exp > 0 else "0%"
                ws.append([name, amt, pct, count])
                style_data_row(ws, ws.max_row, 4, True)
            ws.append(["TOTAL EXPENSES", total_exp, "100%", ""])
            style_total_row(ws, ws.max_row, 4)
            ws.append([""])

            # Herd Summary
            add_section_title(ws, "HERD SUMMARY", 4)
            ws.append(["Status", "Count", "% of Herd", ""])
            style_header_row(ws, ws.max_row, 4)
            herd_items = [("Milking", milking), ("Dry", dry), ("Heifer", heifer), ("Other", len(herd) - milking - dry - heifer)]
            for status, count in herd_items:
                pct = f"{(count/len(herd)*100):.1f}%" if len(herd) > 0 else "0%"
                ws.append([status, count, pct, ""])
                style_data_row(ws, ws.max_row, 4, True)
            ws.append(["TOTAL HERD", len(herd), "100%", ""])
            style_total_row(ws, ws.max_row, 4)
            ws.append([""])

            # Milk Summary
            add_section_title(ws, "MILK PRODUCTION & SALES SUMMARY", 4)
            ws.append(["Metric", "Value", "Unit", "Notes"])
            style_header_row(ws, ws.max_row, 4)
            milk_rows = [
                ("Total Production", f"{total_milk:.1f}", "Liters", f"{len(data.get('milk_production', []))} records"),
                ("Total Sold", f"{total_liters_sold:.1f}", "Liters", f"{len(data.get('milk_sales', []))} sales"),
                ("Avg Price/Liter", f"{avg_price:.2f}", "KES", "Average across all sales"),
                ("Unsold / Farm Use", f"{max(0, total_milk - total_liters_sold):.1f}", "Liters", "Estimated"),
            ]
            for metric, value, unit, notes in milk_rows:
                ws.append([metric, value, unit, notes])
                style_data_row(ws, ws.max_row, 4)
            ws.append([""])

            # Monthly P&L
            add_section_title(ws, "MONTHLY PROFIT & LOSS (KES)", 4)
            ws.append(["Month", "Income", "Expense", "Net"])
            style_header_row(ws, ws.max_row, 4)
            months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            for i, month in enumerate(months):
                m_income = (sum(safe_float(x.get("amount")) for x in data["milk_sales"] if _get_month(x.get("date", "")) == i) +
                           sum(safe_float(x.get("amount")) for x in data["other_income"] if _get_month(x.get("date", "")) == i))
                m_exp = (sum(safe_float(x.get("amount")) for x in data["feed_expenses"] if _get_month(x.get("date", "")) == i) +
                        sum(safe_float(x.get("cost")) for x in data["vet_health"] if _get_month(x.get("date", "")) == i) +
                        sum(safe_float(x.get("amount")) for x in data["labour"] if _get_month(x.get("date", "")) == i) +
                        sum(safe_float(x.get("amount")) for x in data["operations"] if _get_month(x.get("date", "")) == i))
                ws.append([month, m_income, m_exp, m_income - m_exp])
                style_data_row(ws, ws.max_row, 4, True)

            auto_width(ws)

            # ===== HELPER =====
            def build_detail_sheet(wb, title, headers, rows, subtotal_col=None, subtotal_label="TOTAL"):
                ws = wb.create_sheet(title=title)
                num_cols = len(headers)

                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
                ws['A1'] = title.upper()
                ws['A1'].font = Font(bold=True, color="1a5c3a", size=14)
                ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 25
                ws.append([""])

                ws.append(headers)
                style_header_row(ws, ws.max_row, num_cols)
                ws.row_dimensions[ws.max_row].height = 22

                subtotal = 0.0
                for row in rows:
                    ws.append(row)
                    style_data_row(ws, ws.max_row, num_cols, True)
                    if subtotal_col and len(row) > subtotal_col:
                        try:
                            subtotal += float(row[subtotal_col])
                        except:
                            pass

                if subtotal_col is not None:
                    ws.append([""])
                    total_row = [""] * num_cols
                    total_row[0] = subtotal_label
                    total_row[subtotal_col] = subtotal
                    ws.append(total_row)
                    style_total_row(ws, ws.max_row, num_cols)

                auto_width(ws)
                return ws

            # ===== 2. HERD =====
            herd_rows = []
            for cow in herd:
                age_str = ""
                if cow.get("dob"):
                    try:
                        dob = datetime.strptime(cow["dob"], "%Y-%m-%d")
                        age_days = (datetime.now() - dob).days
                        age_str = f"{age_days // 365}y {(age_days % 365) // 30}m"
                    except:
                        pass
                herd_rows.append([
                    cow.get("tag_no", ""), cow.get("name", ""), cow.get("breed", ""),
                    cow.get("status", ""), cow.get("dob", ""), age_str,
                    cow.get("purchase_date", ""), safe_float(cow.get("purchase_cost", 0)), cow.get("notes", "")
                ])
            build_detail_sheet(wb, "Herd Register",
                ["Tag No", "Name", "Breed", "Status", "Date of Birth", "Age", "Purchase Date", "Cost (KES)", "Notes"],
                herd_rows, subtotal_col=7, subtotal_label="TOTAL HERD VALUE")

            # ===== 3. MILK PRODUCTION =====
            prod_rows = []
            for rec in data.get("milk_production", []):
                prod_rows.append([
                    rec.get("date", ""), rec.get("session", ""), rec.get("cow_tag", ""),
                    safe_float(rec.get("quantity", 0)), rec.get("quality", ""), rec.get("notes", ""), rec.get("recorded_at", "")
                ])
            build_detail_sheet(wb, "Milk Production Records",
                ["Date", "Session", "Cow Tag", "Quantity (L)", "Quality", "Notes", "Recorded At"],
                prod_rows, subtotal_col=3, subtotal_label="TOTAL LITERS")

            # ===== 4. MILK SALES =====
            sales_rows = []
            for rec in data.get("milk_sales", []):
                sales_rows.append([
                    rec.get("date", ""), rec.get("buyer", ""), safe_float(rec.get("liters", 0)),
                    safe_float(rec.get("price_per_liter", 0)), safe_float(rec.get("amount", 0)),
                ])
            build_detail_sheet(wb, "Milk Sales",
                ["Date", "Buyer", "Liters", "Price/Liter (KES)", "Total Amount (KES)"],
                sales_rows, subtotal_col=2, subtotal_label="TOTAL LITERS")
            ws = wb["Milk Sales"]
            ws.append(["", "", "", "TOTAL REVENUE:", sum(safe_float(r[4]) for r in sales_rows)])
            style_total_row(ws, ws.max_row, 5)

            # ===== 5. OTHER INCOME =====
            income_rows = []
            for rec in data.get("other_income", []):
                income_rows.append([rec.get("date", ""), rec.get("source", ""), safe_float(rec.get("amount", 0)), rec.get("notes", "")])
            build_detail_sheet(wb, "Other Income",
                ["Date", "Source", "Amount (KES)", "Notes"],
                income_rows, subtotal_col=2, subtotal_label="TOTAL OTHER INCOME")

            # ===== 6. FEED EXPENSES =====
            feed_rows = []
            for rec in data.get("feed_expenses", []):
                feed_rows.append([
                    rec.get("date", ""), rec.get("feed_type", ""), rec.get("supplier", ""),
                    safe_float(rec.get("quantity", 0)), rec.get("unit", ""), safe_float(rec.get("amount", 0)),
                ])
            build_detail_sheet(wb, "Feed Expenses",
                ["Date", "Feed Type", "Supplier", "Quantity", "Unit", "Amount (KES)"],
                feed_rows, subtotal_col=5, subtotal_label="TOTAL FEED COST")

            # ===== 7. VET & HEALTH =====
            vet_rows = []
            for rec in data.get("vet_health", []):
                vet_rows.append([
                    rec.get("date", ""), rec.get("cow_tag", ""), rec.get("treatment", ""),
                    rec.get("vet_name", ""), safe_float(rec.get("cost", 0)), rec.get("notes", "")
                ])
            build_detail_sheet(wb, "Veterinary & Health",
                ["Date", "Cow Tag", "Treatment/Diagnosis", "Vet Name", "Cost (KES)", "Notes"],
                vet_rows, subtotal_col=4, subtotal_label="TOTAL VET COST")

            # ===== 8. LABOUR =====
            labour_rows = []
            for rec in data.get("labour", []):
                labour_rows.append([
                    rec.get("date", ""), rec.get("worker_name", ""), rec.get("task", ""),
                    safe_float(rec.get("days", 0)), safe_float(rec.get("daily_rate", 0)), safe_float(rec.get("amount", 0)),
                ])
            build_detail_sheet(wb, "Labour & Wages",
                ["Date", "Worker Name", "Task/Role", "Days Worked", "Daily Rate (KES)", "Total Pay (KES)"],
                labour_rows, subtotal_col=5, subtotal_label="TOTAL LABOUR COST")

            # ===== 9. OPERATIONS =====
            ops_rows = []
            for rec in data.get("operations", []):
                ops_rows.append([rec.get("date", ""), rec.get("category", ""), rec.get("description", ""), safe_float(rec.get("amount", 0))])
            build_detail_sheet(wb, "Operations & Overheads",
                ["Date", "Category", "Description", "Amount (KES)"],
                ops_rows, subtotal_col=3, subtotal_label="TOTAL OPERATIONS")

            # ===== 10. FARM NOTES =====
            chat_rows = []
            for rec in data.get("chats", []):
                chat_rows.append([rec.get("timestamp", ""), rec.get("author", ""), rec.get("category", ""), rec.get("message", "")])
            build_detail_sheet(wb, "Farm Notes & Communications",
                ["Timestamp", "Author", "Category", "Message"],
                chat_rows)

            wb.save(filename)
            show_snackbar(f"Excel report exported: {filename}", duration=4)
        except Exception as e:
            show_snackbar(f"Export failed: {str(e)}", THEME["danger"], 4)

    def on_bottom_button_touch(self, screen_name, btn_icon, btn_label, widget, touch):
        if widget.collide_point(*touch.pos):
            if touch.button == "left" and not touch.is_double_tap:
                self.switch_screen(screen_name)
                return True
        return False

    def update_bottom_nav_highlight(self, active_name):
        """Highlight the active bottom nav button."""
        for name, (btn, icon, label) in self.nav_buttons.items():
            if name == active_name:
                icon.text_color = THEME["primary"]
                label.text_color = THEME["primary"]
            else:
                icon.text_color = THEME["text_secondary"]
                label.text_color = THEME["text_secondary"]

    def switch_screen(self, name):
        # Refresh the target screen
        for screen in self.sm.screens:
            if screen.name == name and hasattr(screen, "refresh"):
                screen.refresh()

        self.sm.current = name
        self.toolbar.title = name.replace("_", " ").title()
        self.nav_drawer.set_state("close")

        # Update bottom nav highlight
        self.update_bottom_nav_highlight(name)


if __name__ == "__main__":
    DairyFarmApp().run()




























